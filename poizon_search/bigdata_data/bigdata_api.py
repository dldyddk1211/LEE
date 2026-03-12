# ==========================================
# 빅데이터 관리 - Blueprint (누적 마스터 데이터)
# ==========================================

import os
import json
import sqlite3
import io
import threading
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file, render_template_string
from config import paths

bigdata_bp = Blueprint('bigdata', __name__)

BIGDATA_DIR = os.path.dirname(os.path.abspath(__file__))

# ✅ 공유 드라이브 경로 사용
def _get_db_path():
    return paths.get('bigdata_db')

def _get_backup_dir():
    return paths.get('bigdata_backups')

# ── 필수 컬럼 (이 이름이 정확히 있어야 업로드 허용) ──
REQUIRED_COLUMNS = ['상품번호', '제품명', '크림평균가', '크림판매량', '중국노출가', '중국판매량', '현업자판매량']


def _ensure_dirs():
    os.makedirs(BIGDATA_DIR, exist_ok=True)
    os.makedirs(_get_backup_dir(), exist_ok=True)


def _get_db():
    db_path = _get_db_path()
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bigdata_master (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            상품번호      TEXT NOT NULL UNIQUE,
            제품명        TEXT DEFAULT '',
            크림평균가    TEXT DEFAULT '',
            크림판매량    TEXT DEFAULT '',
            중국노출가    TEXT DEFAULT '',
            중국판매량    TEXT DEFAULT '',
            현업자판매량  TEXT DEFAULT '',
            created_at    TEXT NOT NULL,
            updated_at    TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pno ON bigdata_master(상품번호)")
    conn.commit()
    return conn


# ── 자정 자동 백업 ────────────────────────────────────────────────────────
def _do_backup():
    _ensure_dirs()
    try:
        import openpyxl
        conn = _get_db()
        rows = conn.execute(
            "SELECT 상품번호,제품명,크림평균가,크림판매량,중국노출가,중국판매량,현업자판매량 "
            "FROM bigdata_master ORDER BY 상품번호"
        ).fetchall()
        conn.close()
        if not rows:
            print("⚠️ 빅데이터 자동 백업: 데이터 없음, 건너뜀")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '빅데이터'
        ws.append(REQUIRED_COLUMNS)
        for r in rows:
            ws.append([r[c] for c in REQUIRED_COLUMNS])
        fname = datetime.now().strftime('%Y%m%d_%H%M%S') + '_backup.xlsx'
        wb.save(os.path.join(_get_backup_dir(), fname))
        print(f"✅ 빅데이터 자동 백업 완료: {fname}")
    except Exception as e:
        print(f"❌ 빅데이터 자동 백업 실패: {e}")


def _schedule_midnight_backup():
    now = datetime.now()
    next_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    wait = (next_midnight - now).total_seconds()

    def _run():
        _do_backup()
        _schedule_midnight_backup()

    t = threading.Timer(wait, _run)
    t.daemon = True
    t.start()
    print(f"⏰ 빅데이터 자정 백업 예약 완료 ({next_midnight.strftime('%Y-%m-%d %H:%M:%S')})")


# 서버 기동 시 자정 백업 스케줄 등록
_schedule_midnight_backup()


# ── 메인 페이지 ──────────────────────────────────────────────────────────
@bigdata_bp.route('/bigdata')
def bigdata_index():
    _ensure_dirs()
    tpl_path = os.path.join(os.path.dirname(BIGDATA_DIR), 'templates', 'bigdata.html')
    return render_template_string(open(tpl_path, encoding='utf-8').read())


# ── 통계 ─────────────────────────────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/stats', methods=['GET'])
def get_stats():
    conn = _get_db()
    total = conn.execute("SELECT COUNT(*) FROM bigdata_master").fetchone()[0]
    last  = conn.execute("SELECT MAX(updated_at) FROM bigdata_master").fetchone()[0]
    conn.close()
    return jsonify({'total': total, 'last_updated': last or '-'})


# ── 데이터 조회 (페이징 + 검색) ──────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/rows', methods=['GET'])
def get_rows():
    page     = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 100))
    search   = request.args.get('search', '').strip()

    conn = _get_db()
    if search:
        like = f'%{search}%'
        total = conn.execute(
            "SELECT COUNT(*) FROM bigdata_master WHERE 상품번호 LIKE ? OR 제품명 LIKE ?",
            (like, like)
        ).fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute(
            "SELECT * FROM bigdata_master WHERE 상품번호 LIKE ? OR 제품명 LIKE ? "
            "ORDER BY 상품번호 LIMIT ? OFFSET ?",
            (like, like, per_page, offset)
        ).fetchall()
    else:
        total = conn.execute("SELECT COUNT(*) FROM bigdata_master").fetchone()[0]
        offset = (page - 1) * per_page
        rows = conn.execute(
            "SELECT * FROM bigdata_master ORDER BY 상품번호 LIMIT ? OFFSET ?",
            (per_page, offset)
        ).fetchall()
    conn.close()

    return jsonify({
        'success': True,
        'headers': REQUIRED_COLUMNS,
        'rows': [dict(r) for r in rows],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': max(1, (total + per_page - 1) // per_page)
    })


# ── 엑셀 업로드 (누적 upsert) ────────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/upload', methods=['POST'])
def upload_bigdata():
    _ensure_dirs()
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '파일이 없습니다'}), 400

    filename = file.filename or ''
    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({'success': False, 'message': '.xlsx / .xls / .csv 파일만 업로드 가능합니다'}), 400

    # ── 파일 파싱 ──
    try:
        if filename.endswith('.csv'):
            import csv
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            raw_headers = [h.strip() for h in (reader.fieldnames or [])]
            data_rows = [{k.strip(): v for k, v in row.items()} for row in reader]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers_raw = next(rows_iter, None)
            raw_headers = [str(h).strip() if h is not None else f'열{i+1}' for i, h in enumerate(headers_raw or [])]
            data_rows = []
            for row in rows_iter:
                row_dict = {}
                for col_i, val in enumerate(row):
                    key = raw_headers[col_i] if col_i < len(raw_headers) else f'열{col_i+1}'
                    row_dict[key] = str(val).strip() if val is not None else ''
                # 빈 행 스킵 (상품번호가 없는 행)
                if row_dict.get('상품번호', ''):
                    data_rows.append(row_dict)
            wb.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'파일 파싱 오류: {str(e)}'}), 400

    if not data_rows:
        return jsonify({'success': False, 'message': '데이터가 비어 있습니다 (상품번호가 있는 행 없음)'}), 400

    # ── 필수 컬럼 검증 (공백 제거 후 비교) ──
    missing = [col for col in REQUIRED_COLUMNS if col not in raw_headers]
    if missing:
        return jsonify({
            'success': False,
            'message': (
                f'❌ 필수 컬럼이 없습니다: {", ".join(missing)}\n\n'
                f'필수: {", ".join(REQUIRED_COLUMNS)}\n'
                f'파일: {", ".join(raw_headers)}'
            )
        }), 400

    # ── DB upsert (상품번호 기준으로 누적) ──
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = _get_db()
    inserted = updated = 0
    try:
        for row in data_rows:
            pno = row.get('상품번호', '').strip()
            if not pno:
                continue
            exists = conn.execute(
                "SELECT id FROM bigdata_master WHERE 상품번호=?", (pno,)
            ).fetchone()
            vals = (
                row.get('제품명', ''),
                row.get('크림평균가', ''),
                row.get('크림판매량', ''),
                row.get('중국노출가', ''),
                row.get('중국판매량', ''),
                row.get('현업자판매량', ''),
                now
            )
            if exists:
                conn.execute(
                    "UPDATE bigdata_master SET 제품명=?,크림평균가=?,크림판매량=?,중국노출가=?,중국판매량=?,현업자판매량=?,updated_at=? "
                    "WHERE 상품번호=?",
                    vals + (pno,)
                )
                updated += 1
            else:
                conn.execute(
                    "INSERT INTO bigdata_master (상품번호,제품명,크림평균가,크림판매량,중국노출가,중국판매량,현업자판매량,created_at,updated_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (pno,) + vals[:-1] + (now, now)
                )
                inserted += 1
        conn.commit()
        total = conn.execute("SELECT COUNT(*) FROM bigdata_master").fetchone()[0]
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'DB 저장 오류: {str(e)}'}), 500
    conn.close()

    return jsonify({
        'success': True,
        'inserted': inserted,
        'updated': updated,
        'total': total,
        'message': f'✅ 신규 {inserted:,}건 추가 / 업데이트 {updated:,}건 / 전체 {total:,}건'
    })


# ── 전체 삭제 ─────────────────────────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/clear', methods=['DELETE'])
def clear_all():
    conn = _get_db()
    conn.execute("DELETE FROM bigdata_master")
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '전체 데이터가 삭제되었습니다'})


# ── 엑셀 다운로드 ─────────────────────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/download', methods=['GET'])
def download_all():
    conn = _get_db()
    rows = conn.execute(
        "SELECT 상품번호,제품명,크림평균가,크림판매량,중국노출가,중국판매량,현업자판매량 "
        "FROM bigdata_master ORDER BY 상품번호"
    ).fetchall()
    conn.close()
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '빅데이터'
        ws.append(REQUIRED_COLUMNS)
        for r in rows:
            ws.append([r[c] for c in REQUIRED_COLUMNS])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = datetime.now().strftime('%Y%m%d') + '_빅데이터.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ── 수동 백업 ─────────────────────────────────────────────────────────────
@bigdata_bp.route('/bigdata/api/backup', methods=['POST'])
def manual_backup():
    _do_backup()
    return jsonify({'success': True, 'message': '백업이 완료되었습니다'})
