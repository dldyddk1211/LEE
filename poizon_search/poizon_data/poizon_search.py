import os
import sys
import json
import random
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import requests
from io import BytesIO
from datetime import datetime
import platform

# 중단 플래그
stop_flag = False

# 로그 콜백 함수
LOG_CALLBACK = None

def log(message, level='info'):
    """터미널 + GUI 로그 출력"""
    print(message)
    if LOG_CALLBACK:
        try:
            LOG_CALLBACK(message, level)
        except:
            pass

############################
# ===== USER CONFIG ===== #
############################

MODE = "REAL"
SEARCH_KEYWORD = "나이키"
HEADLESS = False
POIZON_ID = "sionejj@naver.com"
POIZON_PW = "wnaoddl1!"
LOGIN_URL = "https://seller.poizon.com/"
GOODS_SEARCH_URL = "https://seller.poizon.com/main/goods/search"

# 네이버 로그인 설정
NAVER_ID = ""  # 네이버 아이디 입력
NAVER_PW = ""  # 네이버 비밀번호 입력
NAVER_COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_cookies.json")

LOG_DIR = "logs"
SHOT_DIR = "shots"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_data")
COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "poizon_cookies.json")
TEST_MAX_PAGES = 2
REAL_MAX_PAGES = 20

############################


def setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"runlog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    log(f"LOG PATH: {log_path}")
    log(f"MODE: {MODE}")
    log(f"SEARCH_KEYWORD: {SEARCH_KEYWORD}")
    return log_path


def safe_screenshot(page, name: str):
    os.makedirs(SHOT_DIR, exist_ok=True)
    path = os.path.join(SHOT_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{name}.png")
    try:
        page.screenshot(path=path, full_page=True)
        log(f"📸 스크린샷 저장: {path}")
    except Exception as e:
        log(f"⚠️ 스크린샷 실패: {e}", 'error')


def wait_for_inputs(page):
    try:
        page.wait_for_load_state("networkidle", timeout=6000)
    except Exception:
        pass
    try:
        page.wait_for_function("document.querySelectorAll('input').length > 0", timeout=6000)
    except Exception:
        pass


def fill_first(page, selectors, value, step_name="fill"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.fill(sel, value)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def click_first(page, selectors, step_name="click"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.click(sel)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def wait_stable(page, ms=600):
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(ms)


def set_language_korean(page):
    """언어 한국어로 변경"""
    log("\n[1-1] 언어 한국어로 변경 중...")
    try:
        english_found = False
        
        try:
            if page.locator("text=English").count() > 0:
                english_found = True
                log("  → English 버튼 발견 (text)")
        except:
            pass
        
        if not english_found:
            try:
                if page.locator("button:has-text('English')").count() > 0:
                    english_found = True
                    log("  → English 버튼 발견 (button)")
            except:
                pass
        
        if not english_found:
            log("  ✅ 이미 한국어 설정됨", 'success')
            return True
        
        try:
            page.click("text=English", timeout=5000)
        except:
            try:
                page.click("button:has-text('English')", timeout=5000)
            except:
                log("  ⚠️ English 버튼 클릭 실패", 'warning')
                return False
        
        log("  → English 버튼 클릭 완료")
        wait_stable(page, 1500)

        try:
            page.click("text=한국어", timeout=5000)
        except:
            try:
                page.click("button:has-text('한국어')", timeout=5000)
            except:
                try:
                    page.evaluate("""
                        const koreanBtn = [...document.querySelectorAll('*')]
                            .find(el => el.textContent.includes('한국어'));
                        if (koreanBtn) koreanBtn.click();
                    """)
                except:
                    log("  ⚠️ 한국어 버튼 클릭 실패", 'warning')
                    return False

        log("  → 한국어 선택 완료")
        wait_stable(page, 2000)

        # ✅ 실제 언어 변경 확인 (최대 3회 재시도)
        for attempt in range(3):
            try:
                # 검색창 placeholder가 한국어인지 확인
                if page.locator("input[placeholder*='상품명']").count() > 0:
                    log("  ✅ 한국어 변경 완료 (검색창 확인)", 'success')
                    return True
            except:
                pass
            try:
                # English 버튼이 사라졌는지 확인 (언어 변경 완료 신호)
                english_still = (
                    page.locator("text=English").count() > 0 or
                    page.locator("button:has-text('English')").count() > 0
                )
                if not english_still:
                    log("  ✅ 언어 변경 완료", 'success')
                    return True
                else:
                    log(f"  ⚠️ 아직 영문 상태 ({attempt+1}/3) - 재클릭 시도", 'warning')
                    try:
                        page.click("text=English", timeout=3000)
                        wait_stable(page, 1000)
                        try:
                            page.click("text=한국어", timeout=3000)
                        except:
                            page.click("button:has-text('한국어')", timeout=3000)
                        wait_stable(page, 2000)
                    except Exception as re:
                        log(f"  ⚠️ 재클릭 실패: {re}", 'warning')
            except:
                pass

        log("  ⚠️ 언어 변경 확인 불가 (계속 진행)", 'warning')
        return True

    except Exception as e:
        log(f"  ⚠️ 언어 변경 실패: {e}", 'error')
        return False


def try_sort_descending(page):
    log("\n[4-1] 중국 시장 최근 30일 판매량 내림차순 정렬")
    try:
        sales_th = page.locator("th").nth(12)
        icon = sales_th.locator("span.anticon")
        
        if icon.count() > 0:
            icon.last.click()
            page.wait_for_timeout(600)
            page.locator("text=내림차순").first.click()
            page.wait_for_timeout(400)
            page.locator("button:has-text('확인')").first.click()
            page.wait_for_timeout(1000)
            log("  ✅ 내림차순 정렬 완료", 'success')
            return True
    except Exception as e:
        log(f"  ⚠️ 정렬 실패: {e}", 'error')
    return False


def download_image(url):
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            return BytesIO(response.content)
    except Exception:
        pass
    return None


def extract_number(text):
    """텍스트에서 숫자만 추출"""
    if not text:
        return 0
    
    import re
    numbers = re.findall(r'[\d,]+', str(text))
    if numbers:
        return int(numbers[0].replace(',', ''))
    return 0


def search_single_product(code):
    """도매 오더폼용 단건 상품 검색 - search_multiple_products 로직 참고"""
    import platform as _platform
    try:
        with sync_playwright() as p:
            current_os = _platform.system()
            if current_os == 'Darwin':
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )
            else:
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    channel='chrome',
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )

            context = browser.new_context(viewport=None, no_viewport=True)

            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                except Exception:
                    pass

            page = context.new_page()
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 2000)

            # 로그인 감지 및 자동 로그인
            current_url = page.url
            if 'login' in current_url.lower() or page.locator("input[type='password']").count() > 0:
                set_language_korean(page)
                wait_stable(page, 500)
                wait_for_inputs(page)
                try:
                    page.locator("input").nth(0).fill(POIZON_ID)
                    page.locator("input").nth(1).fill(POIZON_PW)
                    try:
                        page.click("button:has-text('로그인')", timeout=3000)
                    except Exception:
                        page.locator("input").nth(1).press("Enter")
                    wait_stable(page, 1500)
                    try:
                        cookies = context.cookies()
                        with open(COOKIE_FILE, 'w') as f:
                            json.dump(cookies, f)
                    except Exception:
                        pass
                    page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
                    wait_stable(page, 2000)
                except Exception as e:
                    browser.close()
                    return None
            else:
                set_language_korean(page)
                wait_stable(page, 2500)

            # 상품명 탭 활성화 (한국어/영문 모두 시도)
            for selector in ["text='상품명'", "span:has-text('상품명')", ".ant-tabs-tab:has-text('상품명')",
                             "text='Product name'", "span:has-text('Product name')", ".ant-tabs-tab:has-text('Product name')"]:
                try:
                    tab = page.locator(selector).first
                    if tab.count() > 0:
                        tab.click()
                        wait_stable(page, 500)
                        break
                except Exception:
                    continue

            # 검색창 클리어 및 품번 입력
            search_input = find_search_input(page)
            if not search_input:
                log("  ⚠️ 검색창 없음 - 언어 재설정 후 재시도")
                set_language_korean(page)
                wait_stable(page, 3000)
                search_input = find_search_input(page)
            if not search_input:
                browser.close()
                return None
            search_input.evaluate("el => el.value = ''")
            wait_stable(page, 200)
            search_input.click()
            wait_stable(page, 200)
            page.keyboard.press("Control+A")
            wait_stable(page, 100)
            page.keyboard.press("Backspace")
            wait_stable(page, 200)
            search_input.type(code, delay=50)
            wait_stable(page, 3000)
            page.keyboard.press("Enter")
            wait_stable(page, 2000)

            # 테이블 대기
            try:
                page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=5000)
            except Exception:
                browser.close()
                return None

            # 스크롤로 이미지 로드
            try:
                page.evaluate("window.scrollTo(0, 300)")
                wait_stable(page, 2000)
                page.evaluate("window.scrollTo(0, 0)")
                wait_stable(page, 2000)
            except Exception:
                pass

            # JS로 이미지 URL + 품명 추출
            data = page.evaluate("""
                () => {
                    const rows = document.querySelectorAll('.ant-table-tbody tr:not(.ant-table-measure-row)');
                    if (rows.length === 0) return null;
                    const row = rows[0];
                    const cells = row.querySelectorAll('td');

                    let img_url = '';
                    const img1 = cells[1]?.querySelector('img');
                    if (img1) {
                        img_url = img1.src || img1.getAttribute('data-src') || img1.getAttribute('data-lazy') || '';
                    }
                    if (!img_url || img_url.startsWith('data:')) {
                        for (const img of row.querySelectorAll('img')) {
                            const src = img.src || img.getAttribute('data-src') || '';
                            if (src && !src.startsWith('data:') && src.startsWith('http')) {
                                img_url = src; break;
                            }
                        }
                    }

                    const nameCell = cells[2]?.innerText || '';
                    const lines = nameCell.split('\\n').map(l => l.trim()).filter(l => l);
                    let name = '';
                    for (const line of lines) {
                        if (line && !line.includes('번호') && !line.toUpperCase().includes('SPU')) {
                            name = line; break;
                        }
                    }

                    return { img_url, name };
                }
            """)

            browser.close()

            if data and data.get('img_url'):
                return {'img_url': data['img_url'], 'name': data.get('name', '')}
            return None

    except Exception as e:
        log(f"search_single_product 오류: {e}", 'error')
        return None


def scrape_current_page(page):
    rows_data = []
    try:
        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
        rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
        count = rows.count()
        log(f"  → 현재 페이지 행 수: {count}")

        # ✅ 첫 번째 행의 셀 개수와 내용 모두 로드될 때까지 대기
        try:
            page.wait_for_function(
                """() => {
                    const rows = document.querySelectorAll('.ant-table-tbody tr:not(.ant-table-measure-row)');
                    if (rows.length === 0) return false;
                    const cells = rows[0].querySelectorAll('td');
                    // 최소 7개 셀이 있고, 어딘가에 숫자/원화가 있어야 함
                    if (cells.length < 7) return false;
                    for (let i = 3; i < Math.min(cells.length, 10); i++) {
                        const txt = cells[i].textContent.trim();
                        if (txt && txt !== '-') return true;
                    }
                    return false;
                }""",
                timeout=10000
            )
            log("  → 셀 데이터 로드 확인", 'info')
        except Exception:
            log("  ⚠️ 셀 로드 대기 타임아웃 - 그대로 진행", 'warning')
            page.wait_for_timeout(3000)

        # ✅ 진단: 첫 번째 행 전체 셀 구조 출력 (컬럼 인덱스 확인용)
        try:
            cell_debug = page.evaluate("""
                () => {
                    const rows = document.querySelectorAll('.ant-table-tbody tr:not(.ant-table-measure-row)');
                    if (rows.length === 0) return [];
                    const cells = rows[0].querySelectorAll('td');
                    return Array.from(cells).map((c, i) => `[${i}]${c.textContent.trim().slice(0,30)}`);
                }
            """)
            log(f"  🔍 셀 구조: {' | '.join(cell_debug)}", 'info')
        except Exception as e:
            log(f"  ⚠️ 진단 실패: {e}", 'warning')

        for i in range(count):
            try:
                row = rows.nth(i)

                # ✅ 스크롤하여 행을 뷰포트에 노출 (lazy load 트리거)
                try:
                    row.scroll_into_view_if_needed(timeout=2000)
                    page.wait_for_timeout(300)
                except:
                    pass

                # 행이 로드될 때까지 대기
                try:
                    row.wait_for(state="visible", timeout=5000)
                except:
                    pass

                cells = row.locator("td")
                
                # 셀이 로드될 때까지 대기
                try:
                    cells.first.wait_for(state="visible", timeout=3000)
                except:
                    pass
                
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception:
                    pass
                
                if not img_url:
                    try:
                        imgs = row.locator("img")
                        if imgs.count() > 0:
                            img_url = imgs.first.get_attribute("src")
                    except Exception:
                        pass

                try:
                    # 타임아웃 증가 및 안전한 접근
                    product_cell = cells.nth(2)
                    
                    # 셀이 존재하는지 먼저 확인
                    if product_cell.count() == 0:
                        log(f"    ⚠️ 행 {i+1}: 상품 정보 셀을 찾을 수 없음", 'warning')
                        continue
                    
                    # 타임아웃을 60초로 증가
                    try:
                        item_info = product_cell.inner_text(timeout=60000)
                    except Exception as e:
                        log(f"    ⚠️ 행 {i+1}: 텍스트 추출 실패 - {str(e)}", 'warning')
                        # JavaScript로 텍스트 추출 시도
                        try:
                            item_info = product_cell.evaluate("el => el.textContent")
                        except:
                            log(f"    ⚠️ 행 {i+1}: JavaScript 추출도 실패, 건너뜀", 'warning')
                            continue
                    
                    lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                    
                    style_id = ""
                    item_name = ""
                    spu_id = ""
                    
                    for idx, line in enumerate(lines):
                        line_clean = line.strip()
                        
                        if not style_id:
                            if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                                style_id = lines[idx + 1].strip()
                            elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                                style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                        
                        if not spu_id:
                            if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                                spu_id = lines[idx + 1].strip()
                            elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                                spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                        
                        if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                            item_name = line_clean
                    
                    log(f"  ✨ 최종 결과 - 상품번호:'{style_id}' / 제품명:'{item_name}' / SPU:'{spu_id}'")
                
                except Exception as e:
                    log(f"    상품 정보 파싱 오류: {e}", 'error')
                    style_id = ""
                    item_name = ""
                    spu_id = ""

                def safe_cell(idx, default='-'):
                    try:
                        txt = cells.nth(idx).inner_text(timeout=5000).strip()
                        return txt if txt else default
                    except:
                        return default

                # ✅ JS로 직접 전체 셀 텍스트 추출 (렌더링 문제 대응)
                try:
                    all_cells_js = row.evaluate("""el => {
                        return Array.from(el.querySelectorAll('td')).map(c => c.textContent.trim());
                    }""")
                except Exception:
                    all_cells_js = []

                def js_cell(idx, default='-'):
                    try:
                        val = all_cells_js[idx] if idx < len(all_cells_js) else ''
                        return val if val else default
                    except:
                        return default

                cell_count = len(all_cells_js)
                log(f"    → 셀 개수: {cell_count}, raw: {all_cells_js[:9]}", 'info')

                brand_category = js_cell(3).replace("\n", " / ")
                status = js_cell(4)
                avg_price = js_cell(5)
                cn_exposure = js_cell(6)
                cn_sales_raw = js_cell(7)
                local_sales_raw = js_cell(8)

                # ✅ JS 추출도 '-'이면 Playwright locator로 재시도
                if avg_price == '-' and cn_exposure == '-':
                    log("  ⚠️ JS 추출 실패 - Playwright로 재시도", 'warning')
                    page.wait_for_timeout(2000)
                    avg_price = safe_cell(5)
                    cn_exposure = safe_cell(6)
                    cn_sales_raw = safe_cell(7)
                    local_sales_raw = safe_cell(8)
                
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)

                rows_data.append({
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "상태": status,
                    "최근30일평균거래가": avg_price,
                    "중국노출": cn_exposure,
                    "중국시장최근30일판매량": cn_sales_num,
                    "중국시장최근30일판매량_원본": cn_sales_raw,
                    "현지판매자최근30일판매량": local_sales_num,
                    "현지판매자최근30일판매량_원본": local_sales_raw,
                })
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"DATA:{json.dumps(rows_data[-1], ensure_ascii=False)}", 'data')
                    except:
                        pass
                
                log(f"    [{i+1}/{count}] ✓ {style_id} | 평균거래가:{avg_price} | 중국노출:{cn_exposure} | 판매량:{cn_sales_num}")

            except Exception as e:
                log(f"    [{i+1}/{count}] ✗ 파싱 실패: {e}", 'error')
                continue

    except Exception as e:
        log(f"⚠️ scrape_current_page 오류: {e}", 'error')

    return rows_data


def save_to_excel(all_data, keyword):
    now = datetime.now()
    filename = f"poizon_{keyword}_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검색결과"
    
    headers = ["이미지", "상품번호", "제품명", "SPU_ID", "상태", 
               "최근30일평균거래가", "중국노출", "중국시장최근30일판매량", "현지판매자최근30일판매량"]
    ws.append(headers)
    ws.row_dimensions[1].height = 20

    row_num = 2
    for data in all_data:
        ws.append([
            "",
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("상태", ""),
            data.get("최근30일평균거래가", ""),
            data.get("중국노출", ""),
            data.get("중국시장최근30일판매량", ""),
            data.get("현지판매자최근30일판매량", ""),
        ])

        img_url = data.get("이미지URL", "")
        if img_url:
            try:
                img_data = download_image(img_url)
                if img_data:
                    img = XLImage(img_data)
                    img.width = 60
                    img.height = 60
                    ws.add_image(img, f'A{row_num}')
                    ws.row_dimensions[row_num].height = 50
            except Exception:
                pass
        row_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 25

    wb.save(filepath)
    log(f"💾 엑셀 저장 완료: {filepath} ({len(all_data)}행)", 'success')
    return filepath


def run(keyword=None, max_pages=None, callback=None, skip_login=False):
    global LOG_CALLBACK, SEARCH_KEYWORD, REAL_MAX_PAGES, stop_flag
    
    # 중단 플래그 초기화
    stop_flag = False
    
    import time as time_module
    total_start_time = time_module.time()
    
    if callback:
        LOG_CALLBACK = callback
    
    if keyword:
        SEARCH_KEYWORD = keyword
    
    if max_pages:
        REAL_MAX_PAGES = int(max_pages)
    
    setup_logging()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            channel='chrome',  # 설치된 Chrome 사용 (더 안정적)
            args=[
                '--window-size=960,648', '--window-position=0,432',  # 최대화 모드
                '--disable-blink-features=AutomationControlled',  # 자동화 감지 우회
            ]
        )
        context = browser.new_context(
            viewport=None,  # 전체 화면 사용
            no_viewport=True  # viewport 제한 없음
        )
        
        need_login = True
        if skip_login and os.path.exists(COOKIE_FILE):
            try:
                with open(COOKIE_FILE, 'r') as f:
                    cookies = json.load(f)
                context.add_cookies(cookies)
                log("\n[1] 저장된 쿠키 로드 완료", 'success')
                need_login = False
            except Exception as e:
                log(f"  ⚠️ 쿠키 로드 실패: {e}", 'error')
                need_login = True
        
        page = context.new_page()

        try:
            if need_login:
                log("\n[1] 접속: " + LOGIN_URL)
                page.goto(LOGIN_URL, wait_until="domcontentloaded")
                wait_stable(page, 500)

                lang_success = set_language_korean(page)
                if not lang_success:
                    raise Exception("언어 변경 실패")
                
                log("\n[2] 로그인 시작")
                wait_for_inputs(page)
                page.locator("input").nth(0).fill(POIZON_ID)
                log("  → ID 입력 완료")
                page.locator("input").nth(1).fill(POIZON_PW)
                log("  → PW 입력 완료")

                try:
                    click_first(page, ["button:has-text('로그인')"], "로그인")
                except Exception:
                    page.locator("input").nth(1).press("Enter")

                wait_stable(page, 3000)
                log("  ✅ 로그인 완료", 'success')

                try:
                    cookies = context.cookies()
                    with open(COOKIE_FILE, 'w') as f:
                        json.dump(cookies, f)
                    log("  💾 쿠키 저장 완료", 'success')
                except Exception as e:
                    log(f"  ⚠️ 쿠키 저장 실패: {e}", 'error')
            else:
                log("\n[2] 쿠키로 로그인 건너뛰기 - 검색 페이지로 바로 이동")
                page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
                wait_stable(page, 2000)

                # 쿠키 만료로 로그인 페이지로 리다이렉트됐는지 확인
                current_url = page.url
                if 'login' in current_url.lower() or page.locator("input[type='password']").count() > 0:
                    log("  ⚠️ 쿠키 만료 → 재로그인 진행", 'warning')
                    set_language_korean(page)
                    wait_stable(page, 1500)
                    wait_for_inputs(page)
                    page.locator("input").nth(0).fill(POIZON_ID)
                    log("  → ID 입력 완료")
                    page.locator("input").nth(1).fill(POIZON_PW)
                    log("  → PW 입력 완료")
                    try:
                        click_first(page, ["button:has-text('로그인')"], "로그인")
                    except Exception:
                        page.locator("input").nth(1).press("Enter")
                    wait_stable(page, 3000)
                    try:
                        cookies = context.cookies()
                        with open(COOKIE_FILE, 'w') as f:
                            json.dump(cookies, f)
                        log("  💾 새 쿠키 저장", 'success')
                    except Exception:
                        pass
                    page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
                    wait_stable(page, 2000)
                else:
                    set_language_korean(page)
                    wait_stable(page, 2000)
                    # ✅ 언어 실제 확인 - 아직 영문이면 재시도
                    for _retry in range(2):
                        try:
                            if page.locator("input[placeholder*='상품명']").count() > 0:
                                log("  ✅ 한국어 검색창 확인됨", 'success')
                                break
                            english_visible = (
                                page.locator("text=English").count() > 0 or
                                page.locator("button:has-text('English')").count() > 0
                            )
                            if english_visible:
                                log(f"  ⚠️ 아직 영문 상태 - 언어 재설정 ({_retry+1}/2)", 'warning')
                                set_language_korean(page)
                                wait_stable(page, 2500)
                            else:
                                break
                        except:
                            break

            log("\n[3] 검색 준비 완료")

            # 키워드 검색 탭 클릭 (기본 선택되도록)
            try:
                log("  → 키워드 검색 탭 활성화 중...")
                # 한국어 + 영문 선택자 모두 시도
                keyword_tab_selectors = [
                    "text='상품명'",
                    "span:has-text('상품명')",
                    ".ant-tabs-tab:has-text('상품명')",
                    "[role='tab']:has-text('상품명')",
                    "text='Product name'",
                    "span:has-text('Product name')",
                    ".ant-tabs-tab:has-text('Product name')",
                ]
                for selector in keyword_tab_selectors:
                    try:
                        tab = page.locator(selector).first
                        if tab.count() > 0:
                            tab.click()
                            log(f"  ✅ 키워드 검색 탭 클릭 완료 (선택자: {selector})")
                            wait_stable(page, 500)
                            break
                    except:
                        continue
            except Exception as e:
                log(f"  ⚠️ 탭 클릭 실패 (계속 진행): {e}", 'warning')

            log(f"\n[4] 키워드 입력: {SEARCH_KEYWORD}")
            wait_for_inputs(page)
                        
            # ✅ 검색창 클리어 및 입력
            try:
                search_input = find_search_input(page)
                if not search_input:
                    log("  ⚠️ 검색창 없음 - 언어 재설정 후 재시도", 'warning')
                    set_language_korean(page)
                    wait_stable(page, 3000)
                    search_input = find_search_input(page)

                if search_input:
                    search_input.evaluate("el => el.value = ''")
                    wait_stable(page, 200)
                    search_input.click()
                    wait_stable(page, 200)
                    page.keyboard.press("Control+A")
                    wait_stable(page, 100)
                    page.keyboard.press("Backspace")
                    wait_stable(page, 200)
                    search_input.type(SEARCH_KEYWORD, delay=50)
                    log(f"  ✅ 키워드 입력 완료: {SEARCH_KEYWORD}")
                else:
                    raise Exception("검색창을 찾을 수 없음")

            except Exception as e:
                log(f"  ❌ 키워드 입력 실패: {e}", 'error')
                fill_first(page, ["input[placeholder*='상품명']", "input[placeholder*='Product']", "input[type='text']"], SEARCH_KEYWORD, "키워드")
                
          
            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except Exception:
                page.keyboard.press("Enter")

            wait_stable(page, 3000)  # 검색 결과 로드 대기 (기존 1600ms → 3000ms)
            try_sort_descending(page)
            wait_stable(page, 1500)  # 정렬 후 추가 대기

            # ✅ 스크롤로 이미지/가격 lazy loading 트리거
            try:
                page.evaluate("window.scrollTo(0, 500)")
                wait_stable(page, 2000)
                page.evaluate("window.scrollTo(0, 0)")
                wait_stable(page, 1000)
                log("  → 스크롤 로드 완료", 'info')
            except Exception:
                pass

            max_pages = REAL_MAX_PAGES
            log(f"\n[5] 수집 시작 (최대 {max_pages}페이지)")
            all_data = []
            page_num = 1

            while page_num <= max_pages:
                # 중단 체크
                if stop_flag:
                    log("\n⏹️ 사용자가 수집을 중단했습니다", 'warning')
                    log(f"현재까지 수집: {len(all_data)}개", 'info')
                    break
                
                log(f"\n  📄 페이지 {page_num} 수집 중...")
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"PROGRESS:{page_num}/{max_pages}|PAGE:{page_num}/{max_pages}", 'progress')
                    except:
                        pass
                
                page_data = scrape_current_page(page)
                all_data.extend(page_data)
                log(f"  ✅ 페이지 {page_num} 완료: {len(page_data)}개 수집 (누계: {len(all_data)}개)", 'success')

                if page_num >= max_pages:
                    log(f"\n  ⏹ 최대 페이지({max_pages}) 도달")
                    break

                try:
                    next_btn = page.locator("li.ant-pagination-next:not(.ant-pagination-disabled) button")
                    if next_btn.count() == 0 or not next_btn.first.is_enabled():
                        log("\n  ⏹ 마지막 페이지 도달")
                        break

                    next_btn.first.click()
                    log(f"  → 다음 페이지로 이동 중...")
                    wait_stable(page, 1500)
                    page_num += 1
                except Exception as e:
                    log(f"\n  ⏹ 다음 페이지 오류: {e}", 'error')
                    break

            # 중단되었거나 정상 완료
            reason = "중단됨" if stop_flag else "정상완료"
            log(f"\n[6] 총 {len(all_data)}개 수집 완료 ({reason})", 'success' if not stop_flag else 'warning')
            
            # 데이터가 있으면 엑셀 저장
            if len(all_data) > 0:
                excel_path = save_to_excel(all_data, SEARCH_KEYWORD)
                safe_screenshot(page, "done")
            else:
                log("⚠️ 수집된 데이터가 없어 엑셀 저장 생략", 'warning')
                excel_path = None
            
            log(f"\n=== {'중단됨' if stop_flag else '완료'} ===", 'warning' if stop_flag else 'success')
            
            total_elapsed_sec = time_module.time() - total_start_time
            hours = int(total_elapsed_sec // 3600)
            minutes = int((total_elapsed_sec % 3600) // 60)
            seconds = int(total_elapsed_sec % 60)
            
            time_str = ""
            if hours > 0:
                time_str = f"{hours}시간 {minutes}분 {seconds}초"
            elif minutes > 0:
                time_str = f"{minutes}분 {seconds}초"
            else:
                time_str = f"{seconds}초"
            
            log(f"⏱️ 총 검색 시간: {time_str}", 'success')
            return {
                'success': True,
                'total_items': len(all_data),
                'pages': page_num,
                'file_path': os.path.basename(excel_path) if excel_path else None,
                'data': all_data,  # ✅ 추가!
                'stopped': stop_flag
            }

        except Exception as e:
            log(f"⛔ Error: {e}", 'error')
            safe_screenshot(page, "error")
            return {'success': False, 'error': str(e)}
        finally:
            page.wait_for_timeout(1000)
            browser.close()


def perform_login():
    """접속 확인용 - 로그인 후 쿠키 저장"""
    global LOG_CALLBACK
    
    try:
        # 쿠키 파일이 있고 24시간 이내면 로그인 스킵
        if os.path.exists(COOKIE_FILE):
            import time
            file_age = time.time() - os.path.getmtime(COOKIE_FILE)
            if file_age < 24 * 3600:
                log("♻️ 이미 로그인되어 있습니다 (쿠키 유효)", 'success')
                return {
                    'success': True,
                    'message': '✅ 이미 접속되어 있습니다 (쿠키 사용)'
                }

        # 쿠키 없거나 만료 → 실제 로그인 진행
        with sync_playwright() as p:
            current_os = platform.system()

            if current_os == 'Darwin':  # macOS
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    args=[
                        '--window-size=960,648', '--window-position=0,432',
                        '--disable-blink-features=AutomationControlled',
                    ]
                )
            else:  # Windows/Linux
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    channel='chrome',
                    args=[
                        '--window-size=960,648', '--window-position=0,432',
                        '--disable-blink-features=AutomationControlled',
                    ]
                )

            context = browser.new_context(
                viewport=None,
                no_viewport=True
            )

            page = context.new_page()

            log("\n[1] POIZON 접속 중...", 'info')
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            wait_stable(page, 500)

            lang_success = set_language_korean(page)
            if not lang_success:
                raise Exception("언어 변경 실패")

            log("\n[2] 로그인 진행 중...", 'info')
            wait_for_inputs(page)
            page.locator("input").nth(0).fill(POIZON_ID)
            log("  → ID 입력 완료")
            page.locator("input").nth(1).fill(POIZON_PW)
            log("  → PW 입력 완료")

            try:
                click_first(page, ["button:has-text('로그인')"], "로그인")
            except Exception:
                page.locator("input").nth(1).press("Enter")

            wait_stable(page, 3000)
            log("  ✅ 로그인 완료", 'success')

            cookies = context.cookies()
            with open(COOKIE_FILE, 'w') as f:
                json.dump(cookies, f)
            log("  💾 쿠키 저장 완료", 'success')

            page.wait_for_timeout(500)
            browser.close()

            log("\n✅ POIZON 접속 완료! 이제 데이터 수집을 시작할 수 있습니다", 'success')

        return {
            'success': True,
            'message': '✅ POIZON 접속 성공'
        }

    except Exception as e:
        log(f"⛔ 접속 오류: {e}", 'error')
        return {
            'success': False,
            'message': f'접속 실패: {str(e)}'
        }


def run_poizon_from_gui(keyword, max_pages=20, callback=None, skip_login=False):
    """GUI에서 호출하는 함수"""
    
    return run(
        keyword=keyword, 
        max_pages=max_pages, 
        callback=callback,
        skip_login=skip_login
    )


###############################################################
# ============ 엑셀 비교 기능 ============== #
###############################################################

def compare_product_price(product_code, product_name, callback=None):
    """단일 상품을 포이즌에서 검색하여 가격 정보 반환"""
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=HEADLESS,
                channel='chrome',
                args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
            )
            context = browser.new_context(viewport=None, no_viewport=True)
            
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                    log("쿠키 로드 완료")
                except:
                    pass
            
            page = context.new_page()
            
            search_query = product_code if product_code else product_name
            log(f"🔍 검색: {search_query}")
            
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)
            
            wait_for_inputs(page)
            fill_first(page, ["input[placeholder*='상품명']"], search_query, "키워드")
            
            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except:
                page.keyboard.press("Enter")
            
            wait_stable(page, 1600)
            
            try:
                page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
                rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
                
                if rows.count() == 0:
                    log(f"  ⚠️ 검색 결과 없음: {search_query}")
                    browser.close()
                    return None
                
                row = rows.nth(0)
                
                # 행이 로드될 때까지 대기
                try:
                    row.wait_for(state="visible", timeout=5000)
                except:
                    pass
                
                cells = row.locator("td")
                
                # 셀이 로드될 때까지 대기
                try:
                    cells.first.wait_for(state="visible", timeout=3000)
                except:
                    pass
                
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception as e:
                    log(f"  이미지 추출 실패: {e}")
                    pass
                
                # 타임아웃 증가 및 안전한 접근
                try:
                    product_cell = cells.nth(2)
                    
                    # 셀이 존재하는지 확인
                    if product_cell.count() == 0:
                        log(f"  ⚠️ 상품 정보 셀을 찾을 수 없음", 'warning')
                        browser.close()
                        return None
                    
                    # 타임아웃을 60초로 증가
                    try:
                        item_info = product_cell.inner_text(timeout=60000)
                    except Exception as e:
                        log(f"  ⚠️ 텍스트 추출 실패: {str(e)}", 'warning')
                        # JavaScript로 텍스트 추출 시도
                        try:
                            item_info = product_cell.evaluate("el => el.textContent")
                        except:
                            log(f"  ❌ JavaScript 추출도 실패", 'error')
                            browser.close()
                            return None
                except Exception as e:
                    log(f"  ❌ 상품 셀 접근 실패: {str(e)}", 'error')
                    browser.close()
                    return None
                
                lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                
                style_id = ""
                item_name = ""
                spu_id = ""
                
                for idx, line in enumerate(lines):
                    line_clean = line.strip()
                    
                    if not style_id:
                        if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                            style_id = lines[idx + 1].strip()
                        elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                            style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                    
                    if not spu_id:
                        if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                            spu_id = lines[idx + 1].strip()
                        elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                            spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                    
                    if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                        item_name = line_clean
                
                avg_price_raw = cells.nth(5).inner_text()
                cn_exposure_raw = cells.nth(6).inner_text()
                cn_sales_raw = cells.nth(7).inner_text()
                local_sales_raw = cells.nth(8).inner_text()
                
                cn_exposure_num = extract_number(cn_exposure_raw)
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)
                
                result = {
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "최근30일평균거래가": avg_price_raw,
                    "중국노출": cn_exposure_raw,
                    "중국노출_숫자": cn_exposure_num,
                    "중국시장최근30일판매량": cn_sales_num,
                    "현지판매자최근30일판매량": local_sales_num,
                }
                
                log(f"  ✅ 찾음: {style_id} | {item_name[:30] if item_name else ''}... | 중국노출: {cn_exposure_raw}")
                
                browser.close()
                return result
                
            except Exception as e:
                log(f"  ❌ 파싱 오류: {e}")
                browser.close()
                return None
                
    except Exception as e:
        log(f"❌ 검색 오류: {e}")
        return None


def save_comparison_to_excel(results, original_products):
    """비교 결과를 엑셀로 저장"""
    now = datetime.now()
    filename = f"poizon_comparison_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가격비교"
    
    headers = ["이미지", "엑셀_상품번호", "상품번호", "제품명", "SPU_ID", 
               "엑셀_정가", "엑셀_할인가", "엑셀_재고", "포이즌_중국노출가", "가격차이",
               "평균거래가", "판매량", "현지판매량"]
    ws.append(headers)
    
    for data in results:
        ws.append([
            "",
            data.get("엑셀_상품번호", ""),
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("엑셀_정가", 0),
            data.get("엑셀_할인가", 0),
            data.get("엑셀_재고", 0),
            data.get("중국노출_숫자", 0),
            data.get("가격차이", 0),
            data.get("최근30일평균거래가", ""),
            data.get("중국시장최근30일판매량", 0),
            data.get("현지판매자최근30일판매량", 0),
        ])
    
    wb.save(filepath)
    log(f"💾 비교 결과 저장: {filepath}")
    
    return filename


###############################################################
# ============ 구매처 검색 결과 엑셀 저장 ============== #
###############################################################

def save_sourcing_results_to_excel(results, reason="정상종료"):
    """
    구매처 검색 결과를 엑셀로 저장
    (sourcing_search.py의 Selenium 버전에서 호출)
    """
    if not results:
        log("💾 저장할 데이터가 없습니다.")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sourcing_result_{timestamp}.xlsx"
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(current_dir, '..', 'outputs')
    
    if not os.path.exists(outputs_dir):
        outputs_dir = OUTPUT_DIR
    
    filepath = os.path.join(outputs_dir, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "구매처 검색 결과"
    
    headers = ['#', '상품번호', '검색결과', '쇼핑몰', '상품명', '가격', '배송비', 'URL', '비고']
    ws.append(headers)
    
    row_num = 1
    for result in results:
        product_code = result.get('product_code', '')
        success = result.get('success', False)
        
        if success:
            products = result.get('products', [])
            
            if products:
                for product in products:
                    ws.append([
                        row_num,
                        product_code,
                        '성공',
                        product.get('mall', ''),
                        product.get('name', ''),
                        product.get('price', ''),
                        product.get('shipping', '무료배송'),
                        product.get('link', ''),
                        ''
                    ])
                    row_num += 1
            else:
                ws.append([
                    row_num,
                    product_code,
                    '검색결과 없음',
                    '', '', '', '', '',
                    '네이버 쇼핑에서 상품을 찾을 수 없습니다'
                ])
                row_num += 1
        else:
            error = result.get('error', '알 수 없는 오류')
            ws.append([
                row_num,
                product_code,
                '실패',
                '', '', '', '', '',
                f'오류: {error}'
            ])
            row_num += 1
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    column_widths = {
        'A': 8, 'B': 20, 'C': 15, 'D': 20, 'E': 50,
        'F': 15, 'G': 12, 'H': 60, 'I': 30
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(filepath)
    
    log(f"\n{'='*60}")
    log(f"✅ 엑셀 저장 완료!")
    log(f"{'='*60}")
    log(f"📁 파일명: {filename}")
    log(f"📂 경로: {filepath}")
    log(f"📊 저장 사유: {reason}")
    log(f"📦 총 {len(results)}개 상품 데이터 저장")
    log(f"{'='*60}\n")
    
    return filepath


def find_search_input(page):
    """검색창 찾기 - 한국어/영문 placeholder 모두 시도"""
    selectors = [
        "input[placeholder*='상품명']",
        "input[placeholder*='Product name']",
        "input[placeholder*='product name']",
        "input[placeholder*='Product']",
        "input[placeholder*='Search']",
        "input[placeholder*='search']",
        "input[placeholder*='Item']",
        "input[type='search']",
        "input[type='text']",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.count() > 0:
                return el
        except Exception:
            continue
    return None


def do_relogin(page, context):
    """쿠키 삭제 후 포이즌 첫 화면에서 재로그인"""
    log("🔄 쿠키 초기화 후 재로그인 시도...", 'warning')
    try:
        # 쿠키 파일 삭제
        if os.path.exists(COOKIE_FILE):
            os.remove(COOKIE_FILE)
            log("  🗑️ 쿠키 파일 삭제", 'info')
        # 컨텍스트 쿠키 초기화
        try:
            context.clear_cookies()
        except Exception:
            pass

        # 포이즌 첫 화면으로 이동
        page.goto(LOGIN_URL, wait_until="domcontentloaded")
        wait_stable(page, 2000)

        # 언어 한국어 설정
        set_language_korean(page)
        wait_stable(page, 2000)

        # 로그인
        wait_for_inputs(page)
        page.locator("input").nth(0).fill(POIZON_ID)
        page.locator("input").nth(1).fill(POIZON_PW)
        try:
            page.click("button:has-text('로그인')", timeout=3000)
        except Exception:
            page.locator("input").nth(1).press("Enter")
        wait_stable(page, 2500)

        # 새 쿠키 저장
        try:
            cookies = context.cookies()
            with open(COOKIE_FILE, 'w') as f:
                json.dump(cookies, f)
            log("  💾 새 쿠키 저장", 'success')
        except Exception:
            pass

        # 검색 페이지로 이동
        page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
        wait_stable(page, 3000)

        # 언어 다시 한국어 설정
        set_language_korean(page)
        wait_stable(page, 3000)

        log("  ✅ 재로그인 완료", 'success')
        return True
    except Exception as e:
        log(f"  ❌ 재로그인 실패: {e}", 'error')
        return False


def run_excel_comparison(products, callback=None):
    """
    엑셀 리스트의 여러 상품을 포이즌에서 검색하여 비교

    inner_text() 대신 evaluate()와 textContent 사용 (안정적!)
    """
    global LOG_CALLBACK, stop_flag
    
    if callback:
        LOG_CALLBACK = callback
    
    stop_flag = False
    
    import time as time_module
    import random
    
    try:
        results = []
        total = len(products)
        
        log(f"\n🔍 총 {total}개 상품 비교 시작", 'info')
        
        # 브라우저 열기
        with sync_playwright() as p:
            current_os = platform.system()
            
            if current_os == 'Darwin':
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )
            else:
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    channel='chrome',
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )
            
            context = browser.new_context(viewport=None, no_viewport=True)
            
            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                    log("✅ 쿠키 로드", 'success')
                except:
                    pass
            
            page = context.new_page()
            
            # 초기 설정
            log("\n[1] 포이즌 접속", 'info')
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 3000)
            
            # 한국어 설정
            set_language_korean(page)
            wait_stable(page, 3000)

            # 각 상품 검색
            log("\n[2] 상품 검색 시작", 'info')

            # 봇 감지 회피: 100~130개마다 재로그인
            relogin_interval = random.randint(100, 130)
            relogin_counter = 0
            log(f"  ℹ️ 재로그인 간격: {relogin_interval}개", 'info')

            for idx, product in enumerate(products, 1):
                if stop_flag:
                    break

                # 주기적 재로그인 (그림판 캡챠 회피)
                relogin_counter += 1
                if relogin_counter >= relogin_interval:
                    log(f"\n  🔄 [{idx}/{total}] {relogin_counter}개 검색 완료 → 재로그인 (캡챠 회피)", 'warning')
                    do_relogin(page, context)
                    relogin_counter = 0
                    relogin_interval = random.randint(100, 130)
                    log(f"  ℹ️ 다음 재로그인 간격: {relogin_interval}개", 'info')

                # 진행상황
                if callback:
                    callback(f"PROGRESS:{idx}/{total}", 'progress')
                
                product_code = product.get('code', product.get('상품번호', ''))
                product_name = product.get('name', product.get('상품명', ''))
                
                if not product_code and not product_name:
                    continue
                
                log(f"\n[{idx}/{total}] 🔍 검색: {product_code}", 'info')
                
                search_query = product_code if product_code else product_name
                                
                try:
                    # ✅ 검색창 찾기 (한국어/영문 모두 시도)
                    search_input = find_search_input(page)

                    if not search_input:
                        log(f"  ❌ 검색창 없음 - 언어 재설정 시도")
                        set_language_korean(page)
                        wait_stable(page, 3000)
                        search_input = find_search_input(page)

                    if not search_input:
                        log(f"  ❌ 검색창 없음")
                        fail_data = create_fail_data(product, '검색창 없음')
                        results.append(fail_data)
                        send_result(callback, product_code, fail_data)
                        continue

                    # ✅ 검색창 완전 클리어
                    try:
                        search_input.evaluate("el => el.value = ''")
                        wait_stable(page, 200)
                        search_input.click()
                        wait_stable(page, 200)
                        page.keyboard.press("Control+A")
                        wait_stable(page, 100)
                        page.keyboard.press("Backspace")
                        wait_stable(page, 300)
                        log(f"  → 검색창 클리어 완료")
                    except Exception as e:
                        log(f"  ⚠️ 클리어 실패: {e}")

                    # ✅ 검색어 입력
                    search_input.type(search_query, delay=50)
                    log(f"  ✅ 검색어 입력: {search_query}")
                    wait_stable(page, 200)

                    # Enter로 검색
                    page.keyboard.press("Enter")
                    log(f"  ✓ 검색 실행")
                    wait_stable(page, 2000)

                    # ✅ 첫 상품: 3초 내 결과 없으면 쿠키 삭제 후 재로그인
                    table_found = False
                    try:
                        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=3000)
                        table_found = True
                    except Exception:
                        pass

                    if not table_found and idx == 1:
                        log("  ⚠️ 첫 상품 3초 내 결과 없음 → 재로그인 후 재시도", 'warning')
                        if do_relogin(page, context):
                            # 재시도
                            search_input = find_search_input(page)
                            if search_input:
                                search_input.evaluate("el => el.value = ''")
                                search_input.click()
                                wait_stable(page, 200)
                                page.keyboard.press("Control+A")
                                page.keyboard.press("Backspace")
                                wait_stable(page, 300)
                                search_input.type(search_query, delay=50)
                                wait_stable(page, 200)
                                page.keyboard.press("Enter")
                                wait_stable(page, 2000)
                                try:
                                    page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=10000)
                                    table_found = True
                                except Exception:
                                    pass

                    if not table_found:
                        # 아직 안 나왔으면 나머지 대기 시간 사용
                        try:
                            page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=7000)
                            table_found = True
                        except Exception:
                            pass

                    if not table_found:
                        log(f"  ⚠️ 검색 결과 없음")
                        fail_data = create_fail_data(product, '검색 결과 없음')
                        results.append(fail_data)
                        send_result(callback, product_code, fail_data)
                        continue
                    
                    # evaluate로 데이터 추출 (inner_text 대신!)
                    data = page.evaluate("""
                        () => {
                            const rows = document.querySelectorAll('.ant-table-tbody tr:not(.ant-table-measure-row)');
                            if (rows.length === 0) return null;
                            
                            const row = rows[0];
                            const cells = row.querySelectorAll('td');
                            
                            // 이미지
                            let img_url = '';
                            const img = cells[1]?.querySelector('img');
                            if (img) img_url = img.src || '';
                            
                            // 상품 정보 (textContent 사용!)
                            const product_text = cells[2]?.textContent || '';
                            
                            // 나머지 정보
                            const brand = cells[3]?.textContent || '-';
                            const status = cells[4]?.textContent || '-';
                            const avg_price = cells[5]?.textContent || '-';
                            const cn_exposure = cells[6]?.textContent || '-';
                            const cn_sales = cells[7]?.textContent || '0';
                            const local_sales = cells[8]?.textContent || '0';
                            
                            return {
                                img_url: img_url,
                                product_text: product_text,
                                brand: brand,
                                status: status,
                                avg_price: avg_price,
                                cn_exposure: cn_exposure,
                                cn_sales: cn_sales,
                                local_sales: local_sales
                            };
                        }
                    """)
                    
                    if not data:
                        log(f"  ⚠️ 데이터 없음")
                        fail_data = create_fail_data(product, '데이터 없음')
                        results.append(fail_data)
                        send_result(callback, product_code, fail_data)
                        continue
                    
                    # 상품 정보 파싱 (개선!)
                    lines = [l.strip() for l in data['product_text'].split("\n") if l.strip()]
                    
                    style_id = ""
                    item_name = ""
                    spu_id = ""
                    
                    for idx_line, line in enumerate(lines):
                        line_clean = line.strip()
                        
                        # 상품번호 추출
                        if not style_id:
                            if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx_line + 1 < len(lines):
                                style_id = lines[idx_line + 1].strip()
                            elif "상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean:
                                # "상품번호: ABC123" 형식
                                temp = line_clean.replace("상품번호:", "").replace("상품번호", "").replace("货号:", "").replace("货号", "").replace("번호:", "").replace("번호", "").strip()
                                # "상품 " 같은 접두어 제거
                                if temp and not temp.startswith("상품"):
                                    style_id = temp
                        
                        # SPU_ID 추출 (데이터로만 사용, 표시 안 함)
                        if not spu_id:
                            if "SPU" in line_clean.upper():
                                spu_id = line_clean.replace("SPU_ID:", "").replace("SPU_ID：", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                        
                        # 제품명 추출 (깔끔하게!)
                        if not item_name and line_clean:
                            # 제외할 조건들
                            is_style_id = (line_clean == style_id)
                            has_keyword = any(k in line_clean for k in ["상품번호", "货号", "번호:", "SPU", "상품 "])
                            is_label = line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:", "SPU_ID:", "SPU:"]
                            
                            # 깔끔한 제품명만 추출
                            if not is_style_id and not has_keyword and not is_label:
                                item_name = line_clean
                                break  # 첫 번째 깔끔한 이름을 찾으면 멈춤
                    
                    log(f"  ✨ 상품번호: {style_id}")
                    log(f"  ✨ 제품명: {item_name}")
                    
                    # 숫자 추출
                    avg_price_num = extract_number(data['avg_price'])
                    cn_sales_num = extract_number(data['cn_sales'])
                    local_sales_num = extract_number(data['local_sales'])
                    
                    # 데이터 구성
                    combined = {
                        **product,
                        '제품명': item_name or product_name or '-',
                        '상품번호': style_id or product_code or '-',
                        'SPU_ID': spu_id or '-',
                        '브랜드카테고리': data['brand'].replace("\n", " / "),
                        '상태': data['status'],
                        '최근30일평균거래가': data['avg_price'],
                        '중국노출': data['cn_exposure'],
                        '중국노출_숫자': avg_price_num,
                        '중국시장최근30일판매량': cn_sales_num,
                        '현지판매자최근30일판매량': local_sales_num,
                        '이미지URL': data['img_url'],
                        '엑셀_상품번호': product.get('code', '-'),
                        '엑셀_할인가': product.get('sale_price', 0),
                        '엑셀_재고': product.get('stock', 0),
                        '가격차이': (product.get('sale_price', 0) - avg_price_num) if avg_price_num else 0,
                    }
                    
                    results.append(combined)
                    log(f"  ✅ 완료", 'success')
                    
                    # 실시간 전송
                    send_result(callback, product_code, combined)
                    
                except Exception as e:
                    err_msg = str(e)
                    log(f"  ❌ 오류: {err_msg}", 'error')
                    fail_data = create_fail_data(product, f'오류: {err_msg}')
                    results.append(fail_data)
                    send_result(callback, product_code, fail_data)

                    # 브라우저가 닫혔으면 루프 중단
                    if 'has been closed' in err_msg or 'browser' in err_msg.lower():
                        log("  ⛔ 브라우저가 닫혔습니다. 검색 중단.", 'warning')
                        break

                if stop_flag:
                    log("\n⏹️ 사용자가 검색을 중단했습니다", 'warning')
                    break

                # 다음 상품 대기
                time_module.sleep(random.uniform(1.0, 2.0))

            # 브라우저 종료
            try:
                browser.close()
            except:
                pass
        
        # 엑셀 저장
        if results:
            filepath = save_comparison_to_excel(results, products)
            
            log(f"\n✅ 비교 완료!", 'success')
            
            return {
                'success': True,
                'total_items': len(results),
                'file_path': os.path.basename(filepath),
                'results': results
            }
        else:
            return {'success': False, 'error': '결과 없음'}
            
    except Exception as e:
        log(f"❌ 오류: {e}", 'error')
        return {'success': False, 'error': str(e)}


def create_fail_data(product, reason):
    """실패 데이터 생성"""
    return {
        **product,
        '제품명': reason,
        '상품번호': '-',
        'SPU_ID': '-',
        '브랜드카테고리': '-',
        '상태': '-',
        '최근30일평균거래가': '-',
        '중국노출': '-',
        '중국노출_숫자': 0,
        '중국시장최근30일판매량': 0,
        '현지판매자최근30일판매량': 0,
        '이미지URL': '',
        '엑셀_상품번호': product.get('code', '-'),
        '엑셀_할인가': product.get('sale_price', 0),
        '엑셀_재고': product.get('stock', 0),
        '가격차이': 0,
    }


def send_result(callback, product_code, data):
    """실시간 전송"""
    print(f"[SEND_RESULT] 함수 시작!")  # ← 직접 print!
    print(f"[SEND_RESULT] callback 존재: {callback is not None}")
    
    if callback:
        try:
            # 🔍 디버깅: 전송할 데이터 확인
            log(f"  🔍 [DEBUG] send_result 호출됨", 'info')
            log(f"  🔍 [DEBUG] product_code: {product_code}", 'info')
            log(f"  🔍 [DEBUG] data 키: {list(data.keys())}", 'info')
            log(f"  🔍 [DEBUG] 제품명: {data.get('제품명', 'KEY_MISSING')}", 'info')
            
            json_data = json.dumps({'product_code': product_code, 'products': [data]}, ensure_ascii=False)
            log(f"  🔍 [DEBUG] JSON 생성 성공: {len(json_data)}자", 'info')
            
            print(f"[SEND_RESULT] callback 호출 직전!")  # ← 직접 print!
            callback(f"PRODUCT_RESULT:{json_data}", 'data')
            print(f"[SEND_RESULT] callback 호출 완료!")  # ← 직접 print!
            
            log(f"  ✅ [DEBUG] callback 호출 완료!", 'success')
        except Exception as e:
            print(f"[SEND_RESULT] 오류 발생: {e}")  # ← 직접 print!
            log(f"  ❌ [DEBUG] send_result 오류: {e}", 'error')
            import traceback
            log(f"  ❌ [DEBUG] traceback: {traceback.format_exc()}", 'error')
            traceback.print_exc()  # ← 직접 print!

def search_multiple_products(product_codes, progress_queue):
    """
    여러 상품을 포이즌에서 검색 (무신사→포이즌용)
    
    Args:
        product_codes: 검색할 상품번호 리스트
        progress_queue: 진행 상황 전달용 큐
    """
    import time as time_module
    import random
    
    try:
        total = len(product_codes)
        
        progress_queue.put({
            'event': 'message',
            'data': {'message': f'🟣 포이즌 검색 시작: {total}개'}
        })
        
        with sync_playwright() as p:
            # OS에 맞게 브라우저 실행
            current_os = platform.system()
            
            if current_os == 'Darwin':
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )
            else:
                browser = p.chromium.launch(
                    headless=HEADLESS,
                    channel='chrome',
                    args=['--window-size=960,648', '--window-position=0,432', '--disable-blink-features=AutomationControlled']
                )
            
            context = browser.new_context(viewport=None, no_viewport=True)
            
            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                except:
                    pass
            
            page = context.new_page()
            
            # 초기 설정

            # 초기 설정
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 2000)

            # ✅ 로그인 페이지 감지 및 자동 로그인
            current_url = page.url
            print(f"  현재 URL: {current_url}")

            if 'login' in current_url.lower() or page.locator("input[type='password']").count() > 0:
                print("  → 로그인 페이지 감지! 자동 로그인 시도...")
                progress_queue.put({
                    'event': 'message',
                    'data': {'message': '🔐 로그인 중...'}
                })
                
                set_language_korean(page)
                wait_stable(page, 500)
                wait_for_inputs(page)
                
                try:
                    page.locator("input").nth(0).fill(POIZON_ID)
                    page.locator("input").nth(1).fill(POIZON_PW)
                    
                    try:
                        page.click("button:has-text('로그인')", timeout=3000)
                    except:
                        page.locator("input").nth(1).press("Enter")
                    
                    wait_stable(page, 1500)
                    
                    # 쿠키 저장
                    try:
                        cookies = context.cookies()
                        with open(COOKIE_FILE, 'w') as f:
                            json.dump(cookies, f)
                        print("  ✅ 자동 로그인 및 쿠키 저장 완료")
                    except:
                        pass
                    
                    # 검색 페이지로 이동
                    page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
                    wait_stable(page, 2000)
                    
                except Exception as login_err:
                    print(f"  ❌ 자동 로그인 실패: {login_err}")
                    progress_queue.put({
                        'event': 'error',
                        'data': {'error': f'로그인 실패: {str(login_err)}'}
                    })
                    browser.close()
                    return  # ← 함수 종료

            else:
                set_language_korean(page)
                wait_stable(page, 1000)
                        
            # 키워드 검색 탭 활성화
            try:
                keyword_tab_selectors = [
                    "text='상품명'",
                    "span:has-text('상품명')",
                    ".ant-tabs-tab:has-text('상품명')",
                ]
                for selector in keyword_tab_selectors:
                    try:
                        tab = page.locator(selector).first
                        if tab.count() > 0:
                            tab.click()
                            wait_stable(page, 500)
                            break
                    except:
                        continue
            except:
                pass
            
            # 각 상품 검색
            for idx, product_code in enumerate(product_codes, 1):
                if stop_flag:
                    break
                
                # 진행상황
                progress_queue.put({
                    'event': 'progress',
                    'data': {
                        'current': idx,
                        'total': total
                    }
                })
                
                try:
                    # 검색창 찾기 (한국어/영문 모두 시도)
                    search_input = find_search_input(page)
                    if not search_input:
                        log("  ⚠️ 검색창 없음 - 언어 재설정 후 재시도", 'warning')
                        set_language_korean(page)
                        wait_stable(page, 3000)
                        search_input = find_search_input(page)
                    if not search_input:
                        raise Exception("검색창을 찾을 수 없음")

                    # 검색창 완전 클리어
                    search_input.evaluate("el => el.value = ''")
                    wait_stable(page, 200)
                    search_input.click()
                    wait_stable(page, 200)
                    page.keyboard.press("Control+A")
                    wait_stable(page, 100)
                    page.keyboard.press("Backspace")
                    wait_stable(page, 200)

                    # 검색어 입력
                    search_input.type(product_code, delay=50)
                    wait_stable(page, 3000)

                    # Enter로 검색
                    page.keyboard.press("Enter")
                    wait_stable(page, 2000)
                    
                   
                    # 테이블 대기
                    try:
                        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=5000)
                    except:
                        progress_queue.put({
                            'event': 'result',
                            'data': {
                                'product_code': product_code,
                                'success': False,
                                'error': '검색 결과 없음'
                            }
                        })
                        continue
                    
                    # ✅ 이미지 로드를 위해 잠깐 스크롤
                    try:
                        page.evaluate("window.scrollTo(0, 300)")
                        wait_stable(page, 2000)  # 이미지 로드 대기
                        page.evaluate("window.scrollTo(0, 0)")
                        wait_stable(page, 2000)
                    except:
                        pass

                    # 첫 번째 결과 파싱
                    data = page.evaluate("""
                        () => {
                            const rows = document.querySelectorAll('.ant-table-tbody tr:not(.ant-table-measure-row)');
                            if (rows.length === 0) return null;
                            
                            const row = rows[0];
                            const cells = row.querySelectorAll('td');
                            
                            // ✅ 이미지 URL - 여러 방법 시도
                            let img_url = '';
                            
                            // 방법1: cells[1] img src
                            const img1 = cells[1]?.querySelector('img');
                            if (img1) {
                                img_url = img1.src || img1.getAttribute('data-src') || img1.getAttribute('data-lazy') || '';
                            }
                            
                            // 방법2: cells[1] src 없으면 전체 행에서 찾기
                            if (!img_url || img_url.startsWith('data:')) {
                                const allImgs = row.querySelectorAll('img');
                                for (const img of allImgs) {
                                    const src = img.src || img.getAttribute('data-src') || img.getAttribute('data-lazy') || '';
                                    if (src && !src.startsWith('data:') && src.startsWith('http')) {
                                        img_url = src;
                                        break;
                                    }
                                }
                            }
                            
                            // 방법3: background-image 스타일 확인
                            if (!img_url) {
                                const divs = cells[1]?.querySelectorAll('div');
                                for (const div of divs || []) {
                                    const bg = window.getComputedStyle(div).backgroundImage;
                                    if (bg && bg !== 'none' && bg.includes('http')) {
                                        img_url = bg.replace(/url\(["']?/, '').replace(/["']?\)$/, '');
                                        break;
                                    }
                                }
                            }
                            
                            return {
                                img_url: img_url,
                                cn_exposure: cells[6]?.textContent?.trim() || '-',
                                cn_sales: cells[7]?.textContent?.trim() || '0',
                                local_sales: cells[8]?.textContent?.trim() || '0'
                            };
                        }
                    """)
                  
                    if data:
                        # 숫자 추출
                        cn_sales_num = extract_number(data['cn_sales'])
                        local_sales_num = extract_number(data['local_sales'])
                        
                        # ✅ 디버깅: 원본 데이터 확인
                        print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
                        print(f"📦 {product_code} 원본 데이터:")
                        print(f"  중국노출: {data.get('cn_exposure', 'NONE')}")
                        print(f"  이미지URL: {data.get('img_url', 'NONE')}")
                        print(f"  중국판매량: {data.get('cn_sales', 'NONE')} → {cn_sales_num}")
                        print(f"  현업자판매량: {data.get('local_sales', 'NONE')} → {local_sales_num}")
                        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
                        
                        result = {
                            'product_code': product_code,
                            'success': True,
                            'poizon_data': {
                                '중국노출': data['cn_exposure'],
                                '이미지URL': data.get('img_url', '-'),  # ✅ 추가!
                                '중국시장최근30일판매량': cn_sales_num,
                                '현지판매자최근30일판매량': local_sales_num
                            }
                        }
                        
                        # ✅ 디버깅: 최종 결과 확인
                        print(f"📤 {product_code} 전송 데이터:")
                        print(f"  중국노출: {result['poizon_data']['중국노출']}")
                        print(f"  이미지URL: {result['poizon_data']['이미지URL']}")
                        print(f"  중국판매량: {result['poizon_data']['중국시장최근30일판매량']:,}개")
                        print(f"  현업자판매량: {result['poizon_data']['현지판매자최근30일판매량']:,}개\n")
                        

                        # 실시간 전송
                        progress_queue.put({
                            'event': 'result',
                            'data': result
                        })
                    else:
                        progress_queue.put({
                            'event': 'result',
                            'data': {
                                'product_code': product_code,
                                'success': False,
                                'error': '데이터 파싱 실패'
                            }
                        })
                    
                except Exception as e:
                    err_msg = str(e)
                    print(f"  ❌ 상품 {product_code} 검색 오류: {err_msg}")
                    progress_queue.put({
                        'event': 'result',
                        'data': {
                            'product_code': product_code,
                            'success': False,
                            'error': err_msg
                        }
                    })

                    # 브라우저가 닫혔으면 루프 중단
                    if 'has been closed' in err_msg or 'browser' in err_msg.lower():
                        print("  ⛔ 브라우저가 닫혔습니다. 검색 중단.")
                        break

                if stop_flag:
                    print("\n⏹️ 사용자가 검색을 중단했습니다")
                    break

                # 다음 상품 대기
                time_module.sleep(random.uniform(1.0, 2.0))

            try:
                browser.close()
            except:
                pass
        
        # 완료
        progress_queue.put({
            'event': 'complete',
            'data': {
                'message': f'✅ 포이즌 검색 완료'
            }
        })
        
    except Exception as e:
        print(f"❌ search_multiple_products 오류: {e}")
        import traceback
        traceback.print_exc()
        progress_queue.put({
            'event': 'error',
            'data': {'error': str(e)}
        })

if __name__ == "__main__":
    print("=" * 60)
    print("⚠️  이 파일은 app.py를 통해 실행해주세요!")
    print("=" * 60)
    print("올바른 실행 방법:")
    print("  py -u app.py")
    print("=" * 60)
    # run()  # 자동 실행 금지! 26-02-17 04시 test
    