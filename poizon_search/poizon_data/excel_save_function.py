"""
엑셀 저장 함수 - poizon_search.py에 추가할 코드
"""

from datetime import datetime
from openpyxl import Workbook
import os

def save_sourcing_results_to_excel(results, reason="정상종료"):
    """
    구매처 검색 결과를 엑셀로 저장
    
    Args:
        results: 검색 결과 리스트
        reason: 저장 사유 (정상종료, 사용자 중단, 오류 등)
    
    Returns:
        str: 저장된 파일 경로
    """
    if not results:
        print("저장할 데이터가 없습니다.")
        return None
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sourcing_result_{timestamp}.xlsx"
    
    # 저장 경로 (poizon_data 폴더의 상위 outputs 폴더)
    current_dir = os.path.dirname(__file__)
    outputs_dir = os.path.join(current_dir, '..', 'outputs')
    
    # outputs 폴더 없으면 생성
    if not os.path.exists(outputs_dir):
        outputs_dir = current_dir  # 폴더 없으면 현재 폴더
    
    filepath = os.path.join(outputs_dir, filename)
    
    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "구매처 검색 결과"
    
    # 헤더
    headers = ['#', '상품번호', '검색결과', '쇼핑몰', '상품명', '가격', '배송비', 'URL', '비고']
    ws.append(headers)
    
    # 데이터 입력
    row_num = 1
    for result in results:
        product_code = result.get('product_code', '')
        success = result.get('success', False)
        
        if success:
            products = result.get('products', [])
            
            if products:
                # 상품이 있으면 각 상품을 한 행씩
                for product in products:
                    ws.append([
                        row_num,
                        product_code,
                        '성공',
                        product.get('mall', ''),
                        product.get('name', ''),
                        product.get('price', ''),
                        product.get('shipping', ''),
                        product.get('url', ''),
                        ''
                    ])
                    row_num += 1
            else:
                # 검색 성공했지만 상품 없음
                ws.append([
                    row_num,
                    product_code,
                    '검색결과 없음',
                    '', '', '', '', '',
                    '네이버 쇼핑에서 상품을 찾을 수 없습니다'
                ])
                row_num += 1
        else:
            # 검색 실패
            error = result.get('error', '알 수 없는 오류')
            ws.append([
                row_num,
                product_code,
                '실패',
                '', '', '', '', '',
                f'오류: {error}'
            ])
            row_num += 1
    
    # 스타일 적용
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 열 너비 자동 조정
    column_widths = {
        'A': 8,   # #
        'B': 20,  # 상품번호
        'C': 15,  # 검색결과
        'D': 20,  # 쇼핑몰
        'E': 50,  # 상품명
        'F': 15,  # 가격
        'G': 12,  # 배송비
        'H': 60,  # URL
        'I': 30   # 비고
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 파일 저장
    wb.save(filepath)
    
    print(f"\n{'='*60}")
    print(f"✅ 엑셀 저장 완료!")
    print(f"{'='*60}")
    print(f"📁 파일명: {filename}")
    print(f"📂 경로: {filepath}")
    print(f"📊 저장 사유: {reason}")
    print(f"📦 총 {len(results)}개 상품 데이터 저장")
    print(f"{'='*60}\n")
    
    return filepath
