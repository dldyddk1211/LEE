import os
import sys
import json
import random
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import requests
from io import BytesIO

# 로그 콜백 함수
LOG_CALLBACK = None

def log(message, level='info'):
    """터미널 + GUI 로그 출력"""
    print(message)
    if LOG_CALLBACK:
        try:
            LOG_CALLBACK(message, level)
        except:
            pass

############################
# ===== USER CONFIG ===== #
############################

MODE = "REAL"
SEARCH_KEYWORD = "나이키"
HEADLESS = False
POIZON_ID = "sionejj@naver.com"
POIZON_PW = "wnaoddl1!"
LOGIN_URL = "https://seller.poizon.com/"
GOODS_SEARCH_URL = "https://seller.poizon.com/main/goods/search"

# 네이버 로그인 설정
NAVER_ID = ""  # 네이버 아이디 입력
NAVER_PW = ""  # 네이버 비밀번호 입력
NAVER_COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_cookies.json")

LOG_DIR = "logs"
SHOT_DIR = "shots"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_data")
COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "poizon_cookies.json")
TEST_MAX_PAGES = 2
REAL_MAX_PAGES = 20

############################


def setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"runlog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    log(f"LOG PATH: {log_path}")
    log(f"MODE: {MODE}")
    log(f"SEARCH_KEYWORD: {SEARCH_KEYWORD}")
    return log_path


def safe_screenshot(page, name: str):
    os.makedirs(SHOT_DIR, exist_ok=True)
    path = os.path.join(SHOT_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{name}.png")
    try:
        page.screenshot(path=path, full_page=True)
        log(f"📸 스크린샷 저장: {path}")
    except Exception as e:
        log(f"⚠️ 스크린샷 실패: {e}", 'error')


def wait_for_inputs(page):
    try:
        page.wait_for_load_state("networkidle", timeout=6000)
    except Exception:
        pass
    try:
        page.wait_for_function("document.querySelectorAll('input').length > 0", timeout=6000)
    except Exception:
        pass


def fill_first(page, selectors, value, step_name="fill"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.fill(sel, value)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def click_first(page, selectors, step_name="click"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.click(sel)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def wait_stable(page, ms=600):
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(ms)


def set_language_korean(page):
    """언어 한국어로 변경"""
    log("\n[1-1] 언어 한국어로 변경 중...")
    try:
        # 여러 방법으로 English 버튼 찾기
        english_found = False
        
        # 방법 1: text=English
        try:
            if page.locator("text=English").count() > 0:
                english_found = True
                log("  → English 버튼 발견 (text)")
        except:
            pass
        
        # 방법 2: 언어 선택 버튼
        if not english_found:
            try:
                if page.locator("button:has-text('English')").count() > 0:
                    english_found = True
                    log("  → English 버튼 발견 (button)")
            except:
                pass
        
        # 이미 한국어면 종료
        if not english_found:
            log("  ✅ 이미 한국어 설정됨", 'success')
            return True
        
        # English 클릭
        try:
            page.click("text=English", timeout=5000)
        except:
            try:
                page.click("button:has-text('English')", timeout=5000)
            except:
                log("  ⚠️ English 버튼 클릭 실패", 'warning')
                return False
        
        log("  → English 버튼 클릭 완료")
        wait_stable(page, 500)
        
        # 한국어 선택
        try:
            page.click("text=한국어", timeout=5000)
        except:
            try:
                page.click("button:has-text('한국어')", timeout=5000)
            except:
                # JavaScript로 강제 클릭
                try:
                    page.evaluate("""
                        const koreanBtn = [...document.querySelectorAll('*')]
                            .find(el => el.textContent.includes('한국어'));
                        if (koreanBtn) koreanBtn.click();
                    """)
                except:
                    log("  ⚠️ 한국어 버튼 클릭 실패", 'warning')
                    return False
        
        log("  → 한국어 선택 완료")
        wait_stable(page, 800)
        
        # 언어 변경 확인
        try:
            if page.locator("text=한국어").count() > 0:
                log("  ✅ 한국어 변경 완료", 'success')
                return True
        except:
            pass
        
        log("  ✅ 언어 변경 완료 (확인 불가)", 'success')
        return True
        
    except Exception as e:
        log(f"  ⚠️ 언어 변경 실패: {e}", 'error')
        return False


def try_sort_descending(page):
    log("\n[4-1] 중국 시장 최근 30일 판매량 내림차순 정렬")
    try:
        sales_th = page.locator("th").nth(12)
        icon = sales_th.locator("span.anticon")
        
        if icon.count() > 0:
            icon.last.click()
            page.wait_for_timeout(600)
            page.locator("text=내림차순").first.click()
            page.wait_for_timeout(400)
            page.locator("button:has-text('확인')").first.click()
            page.wait_for_timeout(1000)
            log("  ✅ 내림차순 정렬 완료", 'success')
            return True
    except Exception as e:
        log(f"  ⚠️ 정렬 실패: {e}", 'error')
    return False


def download_image(url):
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            return BytesIO(response.content)
    except Exception:
        pass
    return None


def extract_number(text):
    """텍스트에서 숫자만 추출 (예: '31,000+' → 31000, '<5' → 5)"""
    if not text:
        return 0
    
    import re
    # 숫자와 콤마만 추출
    numbers = re.findall(r'[\d,]+', str(text))
    if numbers:
        # 콤마 제거 후 정수 변환
        return int(numbers[0].replace(',', ''))
    return 0


def scrape_current_page(page):
    rows_data = []
    try:
        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
        rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
        count = rows.count()
        log(f"  → 현재 페이지 행 수: {count}")

        for i in range(count):
            try:
                row = rows.nth(i)
                cells = row.locator("td")
                
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception:
                    pass
                
                if not img_url:
                    try:
                        imgs = row.locator("img")
                        if imgs.count() > 0:
                            img_url = imgs.first.get_attribute("src")
                    except Exception:
                        pass

                # td[2]: 상품 정보 파싱
                try:
                    product_cell = cells.nth(2)
                    
                    # 텍스트 파싱만 사용
                    item_info = product_cell.inner_text()
                    lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                    
                    style_id = ""
                    item_name = ""
                    spu_id = ""
                    
                    for idx, line in enumerate(lines):
                        line_clean = line.strip()
                        
                        # 상품번호 찾기
                        if not style_id:
                            if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                                style_id = lines[idx + 1].strip()
                            elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                                style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                        
                        # SPU_ID 찾기
                        if not spu_id:
                            if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                                spu_id = lines[idx + 1].strip()
                            elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                                spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                        
                        # 제품명 찾기 (상품번호와 중복 방지)
                        if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                            item_name = line_clean
                    
                    log(f"  ✨ 최종 결과 - 상품번호:'{style_id}' / 제품명:'{item_name}' / SPU:'{spu_id}'")
                
                except Exception as e:
                    log(f"    상품 정보 파싱 오류: {e}", 'error')
                    style_id = ""
                    item_name = ""
                    spu_id = ""

                brand_category = cells.nth(3).inner_text().replace("\n", " / ")
                status = cells.nth(4).inner_text()
                avg_price = cells.nth(5).inner_text()
                cn_exposure = cells.nth(6).inner_text()
                cn_sales_raw = cells.nth(7).inner_text()
                local_sales_raw = cells.nth(8).inner_text()
                
                # 판매량에서 숫자만 추출
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)

                rows_data.append({
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "상태": status,
                    "최근30일평균거래가": avg_price,
                    "중국노출": cn_exposure,
                    "중국시장최근30일판매량": cn_sales_num,
                    "중국시장최근30일판매량_원본": cn_sales_raw,
                    "현지판매자최근30일판매량": local_sales_num,
                    "현지판매자최근30일판매량_원본": local_sales_raw,
                })
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"DATA:{json.dumps(rows_data[-1], ensure_ascii=False)}", 'data')
                    except:
                        pass
                
                log(f"    [{i+1}/{count}] ✓ {style_id} | {item_name[:30]}... | 판매량:{cn_sales_num}")

            except Exception as e:
                log(f"    [{i+1}/{count}] ✗ 파싱 실패: {e}", 'error')
                continue

    except Exception as e:
        log(f"⚠️ scrape_current_page 오류: {e}", 'error')

    return rows_data


def save_to_excel(all_data, keyword):
    now = datetime.now()
    filename = f"poizon_{keyword}_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검색결과"
    
    headers = ["이미지", "상품번호", "제품명", "SPU_ID", "상태", 
               "최근30일평균거래가", "중국노출", "중국시장최근30일판매량", "현지판매자최근30일판매량"]
    ws.append(headers)
    ws.row_dimensions[1].height = 20

    row_num = 2
    for data in all_data:
        ws.append([
            "",
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("상태", ""),
            data.get("최근30일평균거래가", ""),
            data.get("중국노출", ""),
            data.get("중국시장최근30일판매량", ""),
            data.get("현지판매자최근30일판매량", ""),
        ])

        img_url = data.get("이미지URL", "")
        if img_url:
            try:
                img_data = download_image(img_url)
                if img_data:
                    img = XLImage(img_data)
                    img.width = 60
                    img.height = 60
                    ws.add_image(img, f'A{row_num}')
                    ws.row_dimensions[row_num].height = 50
            except Exception:
                pass
        row_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 25

    wb.save(filepath)
    log(f"💾 엑셀 저장 완료: {filepath} ({len(all_data)}행)", 'success')
    return filepath


def run(keyword=None, max_pages=None, callback=None, skip_login=False):
    global LOG_CALLBACK, SEARCH_KEYWORD, REAL_MAX_PAGES
    
    # 전체 수집 시작 시간
    import time as time_module
    total_start_time = time_module.time()
    
    if callback:
        LOG_CALLBACK = callback
    
    if keyword:
        SEARCH_KEYWORD = keyword
    
    if max_pages:
        REAL_MAX_PAGES = int(max_pages)
    
    setup_logging()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context()
        
        # 쿠키 로드 시도
        need_login = True
        if skip_login and os.path.exists(COOKIE_FILE):
            try:
                with open(COOKIE_FILE, 'r') as f:
                    cookies = json.load(f)
                context.add_cookies(cookies)
                log("\n[1] 저장된 쿠키 로드 완료", 'success')
                need_login = False
            except Exception as e:
                log(f"  ⚠️ 쿠키 로드 실패: {e}", 'error')
                need_login = True
        
        page = context.new_page()

        try:
            if need_login:
                log("\n[1] 접속: " + LOGIN_URL)
                page.goto(LOGIN_URL, wait_until="domcontentloaded")
                wait_stable(page, 500)

                lang_success = set_language_korean(page)
                if not lang_success:
                    raise Exception("언어 변경 실패")
                
                log("\n[2] 로그인 시작")
                wait_for_inputs(page)
                page.locator("input").nth(0).fill(POIZON_ID)
                log("  → ID 입력 완료")
                page.locator("input").nth(1).fill(POIZON_PW)
                log("  → PW 입력 완료")

                try:
                    click_first(page, ["button:has-text('로그인')"], "로그인")
                except Exception:
                    page.locator("input").nth(1).press("Enter")

                wait_stable(page, 400)
                log("  ✅ 로그인 완료", 'success')
                
                # 쿠키 저장
                try:
                    cookies = context.cookies()
                    with open(COOKIE_FILE, 'w') as f:
                        json.dump(cookies, f)
                    log("  💾 쿠키 저장 완료", 'success')
                except Exception as e:
                    log(f"  ⚠️ 쿠키 저장 실패: {e}", 'error')
            else:
                # 쿠키 사용 시 언어 확인
                log("\n[2] 쿠키로 로그인 건너뛰기")
                page.goto(LOGIN_URL, wait_until="domcontentloaded")
                wait_stable(page, 500)
                
                lang_success = set_language_korean(page)
                if not lang_success:
                    log("  ⚠️ 언어 변경 실패, 계속 진행", 'error')

            log("\n[3] 상품 검색 페이지 이동")
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)

            log(f"\n[4] 키워드 입력: {SEARCH_KEYWORD}")
            wait_for_inputs(page)
            fill_first(page, ["input[placeholder*='상품명']"], SEARCH_KEYWORD, "키워드")

            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except Exception:
                page.keyboard.press("Enter")

            wait_stable(page, 1600)
            try_sort_descending(page)

            max_pages = REAL_MAX_PAGES
            log(f"\n[5] 수집 시작 (최대 {max_pages}페이지)")
            all_data = []
            page_num = 1

            while page_num <= max_pages:
                log(f"\n  📄 페이지 {page_num} 수집 중...")
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"PROGRESS:{page_num}/{max_pages}", 'progress')
                    except:
                        pass
                
                page_data = scrape_current_page(page)
                all_data.extend(page_data)
                log(f"  ✅ 페이지 {page_num} 완료: {len(page_data)}개 수집 (누계: {len(all_data)}개)", 'success')

                if page_num >= max_pages:
                    log(f"\n  ⏹ 최대 페이지({max_pages}) 도달")
                    break

                try:
                    next_btn = page.locator("li.ant-pagination-next:not(.ant-pagination-disabled) button")
                    if next_btn.count() == 0 or not next_btn.first.is_enabled():
                        log("\n  ⏹ 마지막 페이지 도달")
                        break

                    next_btn.first.click()
                    log(f"  → 다음 페이지로 이동 중...")
                    wait_stable(page, 1500)
                    page_num += 1
                except Exception as e:
                    log(f"\n  ⏹ 다음 페이지 오류: {e}", 'error')
                    break

            log(f"\n[6] 총 {len(all_data)}개 수집 완료", 'success')
            excel_path = save_to_excel(all_data, SEARCH_KEYWORD)
            safe_screenshot(page, "done")
            log("\n=== 완료 ===", 'success')
            
            # 총 소요 시간 계산
            total_elapsed_sec = time_module.time() - total_start_time
            hours = int(total_elapsed_sec // 3600)
            minutes = int((total_elapsed_sec % 3600) // 60)
            seconds = int(total_elapsed_sec % 60)
            
            # 시간 문자열 생성
            time_str = ""
            if hours > 0:
                time_str = f"{hours}시간 {minutes}분 {seconds}초"
            elif minutes > 0:
                time_str = f"{minutes}분 {seconds}초"
            else:
                time_str = f"{seconds}초"
            
            log(f"⏱️ 총 검색 시간: {time_str}", 'success')
            
            return {
                'success': True,
                'total_items': len(all_data),
                'pages': page_num,
                'file_path': os.path.basename(excel_path)
            }

        except Exception as e:
            log(f"⛔ Error: {e}", 'error')
            safe_screenshot(page, "error")
            return {'success': False, 'error': str(e)}
        finally:
            page.wait_for_timeout(1000)
            browser.close()


def perform_login():
    """접속 확인용 - 로그인 후 쿠키 저장"""
    global LOG_CALLBACK
    
    try:
        # 쿠키가 이미 있으면
        if os.path.exists(COOKIE_FILE):
            import time
            file_age = time.time() - os.path.getmtime(COOKIE_FILE)
            if file_age < 24 * 3600:
                log("♻️ 이미 로그인되어 있습니다 (쿠키 유효)", 'success')
                return {
                    'success': True,
                    'message': '✅ 이미 접속되어 있습니다 (쿠키 사용)'
                }
        
        # 새로 로그인
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            context = browser.new_context()
            page = context.new_page()
            
            log("\n[1] POIZON 접속 중...", 'info')
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            wait_stable(page, 500)
            
            lang_success = set_language_korean(page)
            if not lang_success:
                raise Exception("언어 변경 실패")
            
            log("\n[2] 로그인 진행 중...", 'info')
            wait_for_inputs(page)
            page.locator("input").nth(0).fill(POIZON_ID)
            log("  → ID 입력 완료")
            page.locator("input").nth(1).fill(POIZON_PW)
            log("  → PW 입력 완료")

            try:
                click_first(page, ["button:has-text('로그인')"], "로그인")
            except Exception:
                page.locator("input").nth(1).press("Enter")

            wait_stable(page, 400)
            log("  ✅ 로그인 완료", 'success')
            
            # 쿠키 저장
            cookies = context.cookies()
            with open(COOKIE_FILE, 'w') as f:
                json.dump(cookies, f)
            log("  💾 쿠키 저장 완료", 'success')
            
            page.wait_for_timeout(1000)
            browser.close()
            
            log("\n✅ POIZON 접속 완료! 이제 데이터 수집을 시작할 수 있습니다", 'success')
            
            return {
                'success': True,
                'message': '✅ POIZON 접속 성공'
            }
            
    except Exception as e:
        log(f"⛔ 접속 오류: {e}", 'error')
        return {
            'success': False,
            'message': f'접속 실패: {str(e)}'
        }


def run_poizon_from_gui(keyword, max_pages=20, callback=None, skip_login=False):
    """GUI에서 호출하는 함수"""
    return run(
        keyword=keyword, 
        max_pages=max_pages, 
        callback=callback,
        skip_login=skip_login
    )


###############################################################
# ============ 엑셀 비교 기능 (여기부터 추가) ============== #
###############################################################

def compare_product_price(product_code, product_name, callback=None):
    """
    단일 상품을 포이즌에서 검색하여 가격 정보 반환
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            context = browser.new_context()
            
            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                    log("쿠키 로드 완료")
                except:
                    pass
            
            page = context.new_page()
            
            # 검색어: 상품번호 우선, 없으면 제품명
            search_query = product_code if product_code else product_name
            
            log(f"🔍 검색: {search_query}")
            
            # 상품 검색 페이지로 이동
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)
            
            # 검색어 입력
            wait_for_inputs(page)
            fill_first(page, ["input[placeholder*='상품명']"], search_query, "키워드")
            
            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except:
                page.keyboard.press("Enter")
            
            wait_stable(page, 1600)
            
            # 첫 번째 결과만 가져오기
            try:
                page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
                rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
                
                if rows.count() == 0:
                    log(f"  ⚠️ 검색 결과 없음: {search_query}")
                    browser.close()
                    return None
                
                # 첫 번째 행만 파싱
                row = rows.nth(0)
                cells = row.locator("td")
                
                # 이미지
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception as e:
                    log(f"  이미지 추출 실패: {e}")
                    pass
                
                # 상품 정보
                product_cell = cells.nth(2)
                item_info = product_cell.inner_text()
                lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                
                style_id = ""
                item_name = ""
                spu_id = ""
                
                for idx, line in enumerate(lines):
                    line_clean = line.strip()
                    
                    if not style_id:
                        if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                            style_id = lines[idx + 1].strip()
                        elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                            style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                    
                    if not spu_id:
                        if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                            spu_id = lines[idx + 1].strip()
                        elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                            spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                    
                    if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                        item_name = line_clean
                
                # 가격 정보
                avg_price_raw = cells.nth(5).inner_text()
                cn_exposure_raw = cells.nth(6).inner_text()
                cn_sales_raw = cells.nth(7).inner_text()
                local_sales_raw = cells.nth(8).inner_text()
                
                # 숫자 추출
                cn_exposure_num = extract_number(cn_exposure_raw)
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)
                
                result = {
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "최근30일평균거래가": avg_price_raw,
                    "중국노출": cn_exposure_raw,
                    "중국노출_숫자": cn_exposure_num,
                    "중국시장최근30일판매량": cn_sales_num,
                    "현지판매자최근30일판매량": local_sales_num,
                }
                
                log(f"  ✅ 찾음: {style_id} | {item_name[:30] if item_name else ''}... | 중국노출: {cn_exposure_raw}")
                
                browser.close()
                return result
                
            except Exception as e:
                log(f"  ❌ 파싱 오류: {e}")
                browser.close()
                return None
                
    except Exception as e:
        log(f"❌ 검색 오류: {e}")
        return None


###############################################################
# ============ 네이버 로그인 ============== #
###############################################################

def perform_naver_login(page, force_login=False):
    """
    네이버 자동 로그인 (메인 페이지에서 직접)
    - 로그인 버튼 존재 여부로 로그인 상태 확인
    - 로그인 필요하면 메인에서 바로 로그인
    """
    import random
    
    def log_naver(msg, level='info'):
        print(msg)
        if LOG_CALLBACK:
            try:
                LOG_CALLBACK(msg, level)
            except:
                pass
    
    # 현재 페이지 확인
    current_url = page.url
    log_naver(f"📍 현재 페이지: {current_url[:60]}...")
    
    # 로그인 버튼 존재 여부로 로그인 상태 확인
    page.wait_for_timeout(2000)
    
    login_selectors = [
        # 정확한 셀렉터 (실제 HTML 기반)
        'a.MyView-module__link_login___HpHMW',
        'a[href*="nidlogin.login"]',
        'a[class*="link_login"]',
        # 백업 셀렉터
        'a.link_login',
        'a:has-text("로그인")'
    ]
    
    login_btn = None
    for sel in login_selectors:
        try:
            count = page.locator(sel).count()
            log_naver(f"   확인: {sel} → {count}개")
            if count > 0:
                login_btn = sel
                log_naver(f"   ✓ 로그인 버튼 발견!")
                break
        except:
            continue
    
    if not login_btn and not force_login:
        log_naver("✅ 로그인 버튼 없음 → 이미 로그인되어 있습니다!")
        return True
    
    if not login_btn:
        log_naver("❌ 로그인 버튼을 찾을 수 없음")
        return False
    
    # 로그인 필요
    log_naver("\n" + "="*60)
    log_naver("🔐 네이버 로그인 시작 (메인 페이지)")
    log_naver("="*60)
    
    try:
        # 로그인 버튼 클릭
        log_naver(f"🖱️ 로그인 버튼 클릭...")
        page.click(login_btn)
        page.wait_for_timeout(random.randint(2000, 3000))
        
        # 로그인 페이지로 이동했는지 확인
        current_url = page.url
        log_naver(f"📍 로그인 페이지 이동: {current_url[:60]}...")
        
        # 아이디 입력
        log_naver("📝 아이디 입력 중...")
        page.fill('#id', 'dldyddk1211')
        log_naver("   ✓ 아이디 입력 완료")
        page.wait_for_timeout(random.randint(800, 1500))
        
        # 비밀번호 입력
        log_naver("🔑 비밀번호 입력 중...")
        page.fill('#pw', 'dhkdl4213!')
        log_naver("   ✓ 비밀번호 입력 완료")
        page.wait_for_timeout(random.randint(800, 1500))
        
        # 로그인 버튼 클릭
        log_naver("🚀 로그인 실행...")
        page.click('.btn_login')
        
        # 로그인 완료 대기
        wait_time = random.randint(5000, 8000)
        log_naver(f"⏱️ 로그인 처리 대기: {wait_time/1000:.1f}초")
        page.wait_for_timeout(wait_time)
        
        # 성공 확인
        current_url = page.url
        log_naver(f"🔍 최종 URL: {current_url[:60]}...")
        
        if "nidlogin.login" not in current_url:
            log_naver("✅ 네이버 로그인 성공!")
            
            # 쿠키 저장
            log_naver("💾 쿠키 저장 중...")
            try:
                cookies = page.context.cookies()
                with open(NAVER_COOKIE_FILE, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False, indent=2)
                log_naver(f"✅ 쿠키 저장: {NAVER_COOKIE_FILE}")
            except Exception as e:
                log_naver(f"⚠️ 쿠키 저장 실패: {e}")
            
            log_naver("="*60)
            return True
        else:
            log_naver("❌ 로그인 실패 (로그인 페이지에 남아있음)")
            return False
            
    except Exception as e:
        log_naver(f"❌ 로그인 오류: {e}")
        import traceback
        log_naver(traceback.format_exc())
        return False


###############################################################
# ============ 네이버 쇼핑 구매처 검색 ============== #
###############################################################

def search_naver_shopping(product_code, callback=None, browser_context=None):
    """
    네이버 쇼핑에서 상품 검색
    순서: 네이버 메인 → 로그인 → 쇼핑 홈 → 검색
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    def log(msg, level='info'):
        if LOG_CALLBACK:
            LOG_CALLBACK(msg, level)
    
    try:
        log(f"\n🔍 네이버 쇼핑 검색 시작: {product_code}")
        
        # 브라우저 컨텍스트 재사용 여부
        use_existing_context = browser_context is not None
        
        if not use_existing_context:
            # 새 브라우저 시작
            p = sync_playwright().start()
            browser = p.chromium.launch(
                headless=HEADLESS,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                ]
            )
            
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                locale='ko-KR',
                timezone_id='Asia/Seoul',
            )
            
            # WebDriver 감지 우회 스크립트
            context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                window.chrome = { runtime: {} };
            """)
            
            page = context.new_page()
            
            # === STEP 1: 네이버 메인 접속 ===
            log(f"  🏠 STEP 1: 네이버 메인 접속 중...")
            page.goto("https://www.naver.com/", wait_until="domcontentloaded", timeout=15000)
            
            wait_time = random.randint(3000, 5000)
            log(f"  ⏱️ 메인 페이지 대기: {wait_time/1000:.1f}초")
            page.wait_for_timeout(wait_time)
            log(f"  ✅ 네이버 메인 로딩 완료")
            
            # === STEP 2: 로그인 ===
            log(f"  🔐 STEP 2: 네이버 로그인 진행...")
            if not perform_naver_login(page):
                log("❌ 네이버 로그인 실패", 'error')
                browser.close()
                return {
                    'success': False,
                    'product_code': product_code,
                    'error': '로그인 실패'
                }
            
            # 로그인 후 네이버 메인으로 다시 이동
            log(f"  🏠 로그인 완료, 네이버 메인으로 복귀...")
            page.goto("https://www.naver.com/", wait_until="domcontentloaded", timeout=15000)
            wait_time = random.randint(2000, 4000)
            log(f"  ⏱️ 메인 페이지 대기: {wait_time/1000:.1f}초")
            page.wait_for_timeout(wait_time)
            
            skip_shopping_home = False  # 초기화
            
            # === STEP 1-2: 네이버 메인에서 상품 검색 (1차 검색) ===
            log(f"  🔍 STEP 1-2: 네이버 메인 검색창에서 검색 시도...")
            
            try:
                # 네이버 메인 검색창 셀렉터
                main_search_selectors = [
                    'input#query',
                    'input.search_input',
                    'input[name="query"]',
                    'input[type="search"]',
                    'input[placeholder*="검색"]',
                ]
                
                main_search_input = None
                for selector in main_search_selectors:
                    try:
                        if page.locator(selector).count() > 0:
                            main_search_input = selector
                            log(f"     ✓ 메인 검색창 발견: {selector}")
                            break
                    except:
                        continue
                
                if main_search_input:
                    # 메인 검색창에 입력
                    log(f"     → 메인 검색창 클릭...")
                    page.click(main_search_input)
                    page.wait_for_timeout(random.randint(500, 1000))
                    
                    log(f"     → 상품번호 입력: {product_code}")
                    page.fill(main_search_input, '')
                    page.wait_for_timeout(random.randint(300, 500))
                    
                    for char in product_code:
                        page.type(main_search_input, char, delay=random.randint(100, 200))
                    
                    page.wait_for_timeout(random.randint(800, 1500))
                    
                    # 검색 버튼 찾기
                    log(f"     → 검색 버튼 클릭...")
                    main_search_btn_selectors = [
                        'button.btn_search',
                        'button[type="submit"]',
                        'button:has-text("검색")',
                        '.search_btn'
                    ]
                    
                    btn_clicked = False
                    for btn_sel in main_search_btn_selectors:
                        try:
                            if page.locator(btn_sel).count() > 0:
                                page.wait_for_timeout(random.randint(300, 600))
                                page.click(btn_sel)
                                btn_clicked = True
                                log(f"     ✓ 메인에서 검색 실행!")
                                break
                        except:
                            continue
                    
                    if not btn_clicked:
                        page.press(main_search_input, 'Enter')
                        log(f"     ↩️ Enter로 검색 실행")
                    
                    # 검색 결과 대기
                    wait_time = random.randint(3000, 5000)
                    log(f"     ⏱️ 검색 결과 대기: {wait_time/1000:.1f}초")
                    page.wait_for_timeout(wait_time)
                    
                    current_url = page.url
                    log(f"     📍 검색 후 URL: {current_url[:60]}...")
                    
                    # 쇼핑탭 클릭 시도
                    log(f"     → 쇼핑탭 찾는 중...")
                    shopping_tab_selectors = [
                        'a:has-text("쇼핑")',
                        'a[href*="shopping"]',
                        '.tab_shopping',
                        'a.link_shopping'
                    ]
                    
                    tab_found = False
                    for tab_sel in shopping_tab_selectors:
                        try:
                            if page.locator(tab_sel).count() > 0:
                                page.wait_for_timeout(random.randint(500, 1000))
                                page.click(tab_sel)
                                log(f"     ✓ 쇼핑탭 클릭!")
                                tab_found = True
                                
                                wait_time = random.randint(3000, 5000)
                                log(f"     ⏱️ 쇼핑탭 로딩 대기: {wait_time/1000:.1f}초")
                                page.wait_for_timeout(wait_time)
                                break
                        except:
                            continue
                    
                    if tab_found:
                        log(f"  ✅ 네이버 메인에서 1차 검색 완료!")
                        # 이미 검색 결과에 있으므로 STEP 3, 4 스킵하고 STEP 5로
                        skip_shopping_home = True
                    else:
                        log(f"  ⚠️ 쇼핑탭 못 찾음, 정상 프로세스 진행")
                        skip_shopping_home = False
                else:
                    log(f"  ⚠️ 메인 검색창 못 찾음, 정상 프로세스 진행")
                    skip_shopping_home = False
                    
            except Exception as e:
                log(f"  ⚠️ 메인 검색 오류: {e}, 정상 프로세스 진행")
                skip_shopping_home = False
            
        else:
            # 기존 컨텍스트 사용
            page = browser_context['page']
            skip_shopping_home = False  # 변수 초기화
        
        # 메인 검색이 성공했으면 STEP 3, 4 스킵
        if skip_shopping_home:
            log(f"  ✅ 메인 검색 완료, 쇼핑 홈/검색 단계 스킵")
        else:
            # === STEP 3: 네이버 쇼핑 홈 접속 ===
            log(f"  🛒 STEP 3: 네이버 쇼핑 홈 접속 중...")
            page.goto("https://shopping.naver.com/ns/home", wait_until="domcontentloaded", timeout=15000)
            
            wait_time = random.randint(3000, 5000)
            log(f"  ⏱️ 쇼핑 홈 로딩 대기: {wait_time/1000:.1f}초")
            page.wait_for_timeout(wait_time)
            
            # 현재 URL 확인
            current_url = page.url
            log(f"  🔍 현재 URL: {current_url[:80]}...")
        
            if "security" in current_url.lower() or "error" in current_url.lower():
                log(f"  ❌ 쇼핑 서비스 접속 제한 감지")
            
                if not use_existing_context:
                    browser.close()
            
                return {
                    'success': False,
                    'product_code': product_code,
                    'error': '쇼핑 서비스 접속 제한'
                }
        
            log(f"  ✅ 네이버 쇼핑 홈 로딩 완료")
        
            # === STEP 4: 상품 검색 ===
            log(f"  🔎 STEP 4: 상품 검색 시작...")
        
            try:
                # 네이버 쇼핑 홈의 검색창 찾기
                log(f"  🔍 검색창 찾는 중...")
            
                search_input = None
            
                # 실제 네이버 쇼핑 검색창 셀렉터 (우선순위)
                selectors = [
                    # 쇼핑 홈 검색창
                    'input[placeholder="상품명 또는 브랜드 입력"]',
                    'input._searchInput_search_text_83jy9',
                    'input[title="검색어 입력"]',
                    'input[data-shp-area-id="input"]',
                
                    # 검색 결과 페이지 검색창
                    'input#query',
                    'input[name="query"]',
                    'input.search_input',
                    'input[placeholder="검색어를 입력해 주세요."]',
                
                    # 백업 셀렉터
                    'input[class*="_searchInput"]',
                    'input.input_text',
                    'input[type="search"]',
                    'input[placeholder*="검색"]',
                    'input[class*="search"]',
                    '.search_input_box input',
                    '#autocompleteWrapper input',
                    'form[action*="search"] input',
                    'input[placeholder*="상품"]',
                    'input[placeholder*="브랜드"]',
                    'input[type="text"]'
                ]
            
                for idx, selector in enumerate(selectors, 1):
                    try:
                        count = page.locator(selector).count()
                        if count > 0:
                            log(f"     [{idx}] {selector}: {count}개 발견 ✓")
                            search_input = selector
                            log(f"  ✅ 검색창 선택: {selector}")
                            break
                        else:
                            log(f"     [{idx}] {selector}: 0개")
                    except Exception as e:
                        log(f"     [{idx}] {selector}: 오류")
                        continue
            
                if not search_input:
                    log(f"  ❌ 모든 셀렉터 시도 실패! 검색창을 찾을 수 없음")
                
                    # 현재 페이지 URL 확인
                    current_url = page.url
                    log(f"  📍 현재 페이지: {current_url}")
                
                    # 디버깅: 페이지의 모든 input 태그 확인
                    log(f"  🔍 페이지의 모든 input 태그 분석 중...")
                    try:
                        all_inputs = page.locator('input').all()
                        log(f"  📊 총 {len(all_inputs)}개의 input 태그 발견")
                    
                        for i, inp in enumerate(all_inputs[:5], 1):  # 처음 5개만
                            try:
                                inp_type = inp.get_attribute('type', timeout=1000) or ''
                                inp_class = inp.get_attribute('class', timeout=1000) or ''
                                inp_placeholder = inp.get_attribute('placeholder', timeout=1000) or ''
                                inp_id = inp.get_attribute('id', timeout=1000) or ''
                                inp_name = inp.get_attribute('name', timeout=1000) or ''
                            
                                log(f"     Input #{i}:")
                                log(f"       - type: {inp_type}")
                                if len(inp_class) > 50:
                                    log(f"       - class: {inp_class[:50]}...")
                                else:
                                    log(f"       - class: {inp_class}")
                                log(f"       - placeholder: {inp_placeholder}")
                                log(f"       - id: {inp_id}")
                                log(f"       - name: {inp_name}")
                            except:
                                log(f"     Input #{i}: 속성 읽기 실패")
                    except Exception as e:
                        log(f"  ❌ input 태그 확인 실패: {e}")
                
                    # 스크린샷 저장
                    log(f"  📸 페이지 스크린샷 저장 중...")
                    try:
                        screenshot_path = f"/home/claude/naver_shopping_debug_{product_code}.png"
                        page.screenshot(path=screenshot_path, full_page=True)
                        log(f"  ✅ 스크린샷 저장: {screenshot_path}")
                    except Exception as e:
                        log(f"  ❌ 스크린샷 저장 실패: {e}")
                
                    # HTML 소스 저장
                    log(f"  📄 HTML 소스 저장 중...")
                    try:
                        html_path = f"/home/claude/naver_shopping_debug_{product_code}.html"
                        with open(html_path, 'w', encoding='utf-8') as f:
                            f.write(page.content())
                        log(f"  ✅ HTML 저장: {html_path}")
                    except Exception as e:
                        log(f"  ❌ HTML 저장 실패: {e}")
                
                    # 검색창 못 찾으면 직접 URL로 이동
                    log(f"  🔀 직접 검색 URL로 우회...")
                    search_url = f"https://search.shopping.naver.com/search/all?query={product_code}"
                    page.goto(search_url, wait_until="domcontentloaded", timeout=15000)
                
                    wait_time = random.randint(3000, 5000)
                    log(f"  ⏱️ 검색 결과 대기: {wait_time/1000:.1f}초")
                    page.wait_for_timeout(wait_time)
                
                    result_url = page.url
                    log(f"  🔍 검색 결과 URL: {result_url[:80]}...")
                
                else:
                    # 검색창 발견! 입력 진행
                    log(f"  ⌨️ 상품번호 입력 시작: {product_code}")
                
                    # 검색창 클릭 (포커스)
                    log(f"     → 검색창 클릭...")
                    page.click(search_input)
                    page.wait_for_timeout(random.randint(800, 1200))
                
                    # 기존 내용 지우기
                    log(f"     → 기존 내용 지우기...")
                    page.fill(search_input, '')
                    page.wait_for_timeout(random.randint(500, 800))
                
                    # 상품번호 타이핑 (한 글자씩 - 사람처럼)
                    log(f"     → 타이핑: {product_code}")
                    for char in product_code:
                        page.type(search_input, char, delay=random.randint(100, 200))
                
                    # 입력 완료 후 대기
                    page.wait_for_timeout(random.randint(1000, 2000))
                    log(f"  ✅ 입력 완료")
                
                    # Enter 대신 검색 버튼 클릭 시도
                    log(f"  🔍 검색 버튼 찾는 중...")
                    search_btn_found = False
                
                    search_btn_selectors = [
                        # 정확한 셀렉터 (실제 HTML 기반)
                        'button._searchInput_button_search_wu9xq',
                        'button[data-shp-area-id="search"]',
                        'button[class*="_searchInput_button"]',
                        
                        # 백업 셀렉터
                        'button[type="submit"]',
                        'button.btn_search',
                        'button:has-text("검색")',
                        'a.btn_search',
                        '.search_btn',
                        'button[class*="search"]',
                        'button[aria-label*="검색"]'
                    ]
                
                    for btn_selector in search_btn_selectors:
                        try:
                            if page.locator(btn_selector).count() > 0:
                                log(f"     ✓ 검색 버튼 발견: {btn_selector}")
                                page.wait_for_timeout(random.randint(500, 1000))
                                page.click(btn_selector)
                                search_btn_found = True
                                log(f"  🖱️ 검색 버튼 클릭!")
                                break
                        except Exception as e:
                            log(f"     ✗ {btn_selector}: 클릭 실패")
                            continue
                
                    if not search_btn_found:
                        # 검색 버튼 못 찾으면 Enter 사용
                        log(f"  ⚠️ 검색 버튼 못 찾음, Enter 사용")
                        page.wait_for_timeout(random.randint(500, 1000))
                        page.press(search_input, 'Enter')
                        log(f"  ↩️ Enter 검색")
                
                    # 검색 실행 후 충분한 대기 (5-8초)
                    wait_time = random.randint(5000, 8000)
                    log(f"  ⏱️ 검색 결과 대기: {wait_time/1000:.1f}초")
                    page.wait_for_timeout(wait_time)
                
                    # 검색 결과 URL 확인
                    result_url = page.url
                    log(f"  🔍 검색 결과 URL: {result_url[:80]}...")
                
                    # 접속 제한 체크
                    if "security" in result_url.lower() or "error" in result_url.lower():
                        log(f"  ❌ 쇼핑 서비스 접속 제한 감지!")
                        
                        if not use_existing_context:
                            browser.close()
                        
                        return {
                            'success': False,
                            'product_code': product_code,
                            'error': '쇼핑 서비스 접속 제한'
                        }
                    
                    elif "search.shopping.naver.com" in result_url:
                        log(f"  ✅ 검색 성공!")
                        
                    elif "shopping.naver.com" in result_url and "home" in result_url:
                        log(f"  ❌ 여전히 쇼핑 홈에 있음 - 검색 실패")
                        log(f"  💡 검색 버튼이 작동하지 않음")
                        
                        if not use_existing_context:
                            browser.close()
                        
                        return {
                            'success': False,
                            'product_code': product_code,
                            'error': '검색 실패 (쇼핑 홈에서 벗어나지 못함)'
                        }
                        
                    else:
                        log(f"  ⚠️ 예상치 못한 페이지: {result_url}")
            
            except Exception as e:
                if not skip_shopping_home:  # STEP 4 실행 시에만 오류 로그
                    log(f"  ❌ 검색 실행 오류: {e}", 'error')
                    import traceback
                log(f"  📋 상세 오류:\n{traceback.format_exc()}", 'error')
        # STEP 3, 4 조건부 블록 종료
        
        # === STEP 5: 네이버 가격비교 더보기 클릭 ===
        log(f"  🔗 STEP 5: 네이버 가격비교 더보기 찾는 중...")
        
        try:
            # "네이버 가격비교 더보기" 버튼 찾기
            more_btn_selectors = [
                'text=네이버 가격비교 더보기',
                'a:has-text("네이버 가격비교 더보기")',
                'button:has-text("네이버 가격비교 더보기")',
                'a:has-text("가격비교")',
                '.more_btn',
                'a[href*="search/all"]'
            ]
            
            more_btn_found = False
            for selector in more_btn_selectors:
                try:
                    if page.locator(selector).count() > 0:
                        log(f"  ✅ 더보기 버튼 발견: {selector}")
                        
                        # 클릭 전 대기
                        page.wait_for_timeout(random.randint(1000, 2000))
                        
                        # 클릭
                        log(f"  🖱️ '네이버 가격비교 더보기' 클릭...")
                        page.click(selector)
                        more_btn_found = True
                        
                        # 클릭 후 대기
                        wait_time = random.randint(3000, 5000)
                        log(f"  ⏱️ 페이지 전환 대기: {wait_time/1000:.1f}초")
                        page.wait_for_timeout(wait_time)
                        
                        # 전환된 URL 확인
                        new_url = page.url
                        log(f"  🔍 전환된 URL: {new_url[:80]}...")
                        log(f"  ✅ 네이버 가격비교 페이지 이동 완료!")
                        
                        break
                except Exception as e:
                    log(f"  ⚠️ 버튼 클릭 실패 ({selector}): {e}")
                    continue
            
            if not more_btn_found:
                log(f"  ⚠️ '네이버 가격비교 더보기' 버튼을 찾을 수 없음")
                log(f"  💡 이미 가격비교 페이지에 있거나, 검색 결과가 없을 수 있음")
            
        except Exception as e:
            log(f"  ❌ 더보기 클릭 오류: {e}", 'error')
        
        # 최종 URL 저장
        final_url = page.url
        log(f"  📍 최종 페이지: {final_url[:100]}...")
        
        # TODO: 상품 목록 파싱
        
        # 임시 테스트 데이터
        products = [
            {
                'name': f'{product_code} 테스트 상품',
                'price': '테스트',
                'mall': '테스트',
                'link': final_url
            }
        ]
        
        log(f"  📊 상품 정보: {len(products)}개 (테스트 데이터)")
        log(f"  🏁 검색 프로세스 완료!")
        
        if not use_existing_context:
            page.wait_for_timeout(2000)
            browser.close()
            log(f"  🔒 브라우저 종료")
        
        return {
            'success': True,
            'product_code': product_code,
            'products': products,
            'total': len(products),
            'final_url': final_url
        }
        
    except Exception as e:
        log(f"  ❌ 전체 오류 발생: {e}", 'error')
        
        if not use_existing_context and 'browser' in locals():
            try:
                browser.close()
            except:
                pass
        
        return {
            'success': False,
            'product_code': product_code,
            'error': str(e)
        }
def run_sourcing_search(product_codes, callback=None):
    """
    엑셀 상품 리스트를 포이즌에서 검색하여 가격 비교
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    results = []
    total = len(product_codes)  # ← products → product_codes
    
    # 전체 검색 시작 시간
    import time as time_module
    total_start_time = time_module.time()
    
    log(f"📋 총 {total}개 상품 비교 시작")
    
    import random
    import time
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            
            # 브라우저를 전역으로 저장 (중단 가능하도록)
            try:
                import sys
                if hasattr(sys.modules['__main__'], 'current_browser'):
                    sys.modules['__main__'].current_browser = browser
            except:
                pass
            
            context = browser.new_context()
            
            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                    log("✅ 쿠키 로드 완료")
                except:
                    log("⚠️ 쿠키 로드 실패")
            
            page = context.new_page()
            
            # 초기 설정: 한 번만 실행
            log("\n[1] 포이즌 접속 및 언어 설정")
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            wait_stable(page, 500)
            
            # 한국어 설정 (강화)
            korean_success = set_language_korean(page)
            
            if korean_success:
                # 언어 변경 후 페이지 새로고침
                log("  → 언어 설정 적용을 위해 페이지 새로고침")
                page.reload(wait_until="domcontentloaded")
                wait_stable(page, 500)
            
            # 상품 검색 페이지로 이동
            log("\n[2] 상품 검색 페이지 이동")
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)  # 1500ms → 1000ms (빠르게)
            
            # 검색창이 보일 때까지 대기
            try:
                page.wait_for_selector("input[type='text']", timeout=3000)  # 5000→3000
                log("  ✅ 검색 페이지 로딩 완료")
                
                # 한국어 확인 (placeholder 체크)
                try:
                    placeholder = page.evaluate("""
                        const input = document.querySelector('input[type="text"]');
                        return input ? input.placeholder : '';
                    """)
                    if placeholder:
                        log(f"  → 검색창 placeholder: {placeholder}")
                        if '상품' in placeholder or '商品' in placeholder:
                            log("  ✅ 한국어/중국어 UI 확인")
                except:
                    pass
            except:
                log("  ⚠️ 검색창 로딩 지연")
                wait_stable(page, 1000)
            
            # 각 상품 검색
            relogin_interval = random.randint(80, 100)  # 80~100개 랜덤
            log(f"  ℹ️ 재로그인 주기: {relogin_interval}개마다")
            
            for i, product in enumerate(products):
                # 중단 플래그 확인
                try:
                    import sys
                    if hasattr(sys.modules['__main__'], 'stop_flag') and sys.modules['__main__'].stop_flag:
                        log("⏹️ 사용자가 중단을 요청했습니다")
                        browser.close()
                        return {
                            'success': False,
                            'error': '사용자 중단'
                        }
                except:
                    pass
                
                # 80~100개마다 재로그인 (봇 감지 회피)
                if i > 0 and i % relogin_interval == 0:
                    log(f"\n⚠️ [{i}개 처리] 재로그인 진행 중... (봇 감지 회피)")
                    log(f"  → 모든 검색 일시 중지")
                    
                    # 다음 재로그인 주기 랜덤 설정
                    relogin_interval = random.randint(80, 100)
                    log(f"  → 다음 재로그인: {relogin_interval}개 후")
                    
                    # 현재 진행 중인 작업 완료 대기
                    wait_stable(page, 1000)
                    
                    # 페이지 새로고침 또는 재로그인
                    page.goto(LOGIN_URL, wait_until="domcontentloaded")
                    wait_stable(page, 500)
                    log(f"  → 언어 설정 중...")
                    set_language_korean(page)
                    
                    # 충분한 대기 시간 (안정화)
                    wait_stable(page, 1000)
                    
                    # 상품 검색 페이지로 다시 이동
                    log(f"  → 검색 페이지 이동 중...")
                    page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
                    wait_stable(page, 1000)
                    
                    # 검색창 로딩 대기
                    try:
                        page.wait_for_selector("input[type='text']", timeout=5000)
                        log(f"  ✅ 재로그인 완료")
                    except:
                        log(f"  ⚠️ 검색창 로딩 지연")
                        wait_stable(page, 1000)
                    
                    # 추가 안정화 대기 (누락 방지)
                    wait_stable(page, 1500)
                    log(f"  → 검색 재개...")
                
                log(f"\n[{i+1}/{total}] 🔍 {product['code']} 검색 중...")
                
                if callback:
                    callback(f"PROGRESS:{i+1}/{total}", 'progress')
                
                # 검색어 입력
                search_query = product['code'] if product['code'] else product['name']
                
                try:
                    # 검색창 찾기 (여러 선택자 시도)
                    search_input = None
                    selectors = [
                        "input[placeholder*='상품명']",
                        "input[placeholder*='商品']",
                        "input[placeholder*='搜索']",
                        "input[type='text']",
                        ".ant-input",
                        "input.ant-input"
                    ]
                    
                    for selector in selectors:
                        try:
                            if page.locator(selector).count() > 0:
                                search_input = page.locator(selector).first
                                log(f"  ✓ 검색창 발견: {selector}")
                                break
                        except:
                            continue
                    
                    if not search_input:
                        log(f"  ❌ 검색창을 찾을 수 없습니다")
                        continue
                    
                    # 검색창 클리어 (JavaScript)
                    try:
                        page.evaluate("""
                            const input = document.querySelector('input[placeholder*="상품명"]') || 
                                         document.querySelector('input[placeholder*="商品"]') ||
                                         document.querySelector('input[type="text"]') ||
                                         document.querySelector('.ant-input');
                            if (input) {
                                input.value = '';
                                input.focus();
                            }
                        """)
                        wait_stable(page, 100)
                    except:
                        pass
                    
                    # 검색창 클릭 후 타이핑 (사람처럼)
                    try:
                        search_input.click(force=True, timeout=3000)
                        wait_stable(page, 100)
                        
                        # 한 글자씩 타이핑 (48ms → 53ms, 10% 느리게)
                        search_input.type(search_query, delay=53)
                        log(f"  ✓ 검색어 입력: {search_query}")
                        wait_stable(page, 300)
                    except Exception as e:
                        log(f"  ⚠️ 타이핑 실패, JavaScript로 재시도")
                        
                        # JavaScript로 값 설정 (폴백)
                        page.evaluate(f"""
                            const input = document.querySelector('input[type="text"]');
                            if (input) {{
                                input.value = '{search_query}';
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                        """)
                        wait_stable(page, 300)
                    
                    # 검색 실행
                    search_executed = False
                    
                    # Enter 키로 검색 (가장 자연스러움)
                    try:
                        page.keyboard.press("Enter")
                        search_executed = True
                        log(f"  ✓ 검색 실행 (Enter)")
                    except:
                        pass
                    
                    if not search_executed:
                        # 검색 버튼 클릭
                        try:
                            page.click("button:has-text('검색 및 입찰')", timeout=2000)
                            search_executed = True
                            log(f"  ✓ 검색 실행 (버튼)")
                        except:
                            try:
                                page.click("button:has-text('搜索')", timeout=2000)
                                search_executed = True
                            except:
                                pass
                    
                    if not search_executed:
                        log(f"  ❌ 검색 실행 실패")
                        continue
                    
                    # 로딩 대기 - 테이블이 실제로 업데이트될 때까지
                    import time as time_module
                    search_start_time = time_module.time()
                    
                    wait_stable(page, 400)  # 600ms → 400ms
                    log(f"  ⏱️ 기본 대기: 400ms 완료")
                    
                    # 테이블 로딩 완료 대기 (최대 2초: 1초 × 2회)
                    result_found = False
                    for attempt in range(2):
                        attempt_start = time_module.time()
                        try:
                            page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=400)
                            
                            # 첫 번째 행의 상품번호 확인 (타임아웃 500ms)
                            first_row = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)").first
                            first_cell_text = first_row.locator("td").nth(2).inner_text(timeout=500)
                            
                            # 검색어가 결과에 포함되어 있는지 확인
                            if search_query.upper() in first_cell_text.upper():
                                elapsed = (time_module.time() - search_start_time) * 1000
                                log(f"  ✅ 검색 결과 일치 확인 (총 대기: {elapsed:.0f}ms)")
                                result_found = True
                                break
                            else:
                                # 일치하지 않으면 계속 대기 (0.5초 → 1초)
                                attempt_elapsed = (time_module.time() - attempt_start) * 1000
                                log(f"  ⏳ 시도 {attempt + 1}: 결과 불일치 ({attempt_elapsed:.0f}ms), 1초 대기...")
                                time.sleep(1.0)  # 0.5 → 1.0
                        except Exception as e:
                            attempt_elapsed = (time_module.time() - attempt_start) * 1000
                            log(f"  ⏳ 시도 {attempt + 1}: 테이블 로딩 대기 ({attempt_elapsed:.0f}ms)")
                            time.sleep(1.0)  # 0.5 → 1.0
                    
                    # 결과 없으면 "상품 없음"으로 추가
                    if not result_found:
                        total_elapsed = (time_module.time() - search_start_time) * 1000
                        log(f"  ⚠️ 검색 결과 없음 또는 불일치 (총 대기: {total_elapsed:.0f}ms)")
                        
                        # 상품 없음 데이터 추가
                        no_result_data = {
                            "이미지URL": "",
                            "상품번호": "상품 없음",
                            "제품명": "검색 결과 없음",
                            "SPU_ID": "",
                            "최근30일평균거래가": "-",
                            "중국노출": "-",
                            "중국노출_숫자": 0,
                            "중국시장최근30일판매량": 0,
                            "현지판매자최근30일판매량": 0,
                            "엑셀_상품번호": product['code'],
                            "엑셀_제품명": product['name'],
                            "엑셀_정가": product['original_price'],
                            "엑셀_할인가": product['sale_price'],
                            "엑셀_재고": product.get('stock', 0),
                            "가격차이": 0,
                        }
                        
                        results.append(no_result_data)
                        
                        if callback:
                            callback(f"DATA:{json.dumps(no_result_data, ensure_ascii=False)}", 'data')
                        
                        # 다음 상품으로 (1~2초 랜덤 딜레이)
                        delay = random.uniform(1.0, 2.0)
                        log(f"  ⏱️ 다음 상품 대기: {delay:.1f}초")
                        time.sleep(delay)
                        continue
                    
                    # 추가 랜덤 딜레이 (안정성) - 단축
                    random_delay = random.uniform(0.1, 0.15)  # 0.15~0.25 → 0.1~0.15
                    time.sleep(random_delay)
                    
                    # 첫 번째 결과 파싱
                    try:
                        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=2000)  # 3000→2000
                    except:
                        log(f"  ⚠️ 검색 결과 테이블 로딩 실패 (2초 타임아웃)")
                        continue
                    rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
                    
                    if rows.count() == 0:
                        log(f"  ⚠️ 검색 결과 없음: {search_query}")
                        continue
                    
                    # 첫 번째 행 파싱
                    row = rows.nth(0)
                    cells = row.locator("td")
                    
                    # 이미지 (키워드 검색과 동일한 로직)
                    img_url = ""
                    try:
                        imgs = cells.nth(1).locator("img")
                        if imgs.count() > 0:
                            img_url = imgs.first.get_attribute("src", timeout=1000)
                    except Exception as e:
                        log(f"  ⚠️ 이미지 추출 실패 (cells.nth(1)): {e}")
                        pass
                    
                    # Fallback: cells.nth(1)에서 못 찾으면 전체 행에서 찾기
                    if not img_url:
                        try:
                            imgs = row.locator("img")
                            if imgs.count() > 0:
                                img_url = imgs.first.get_attribute("src", timeout=1000)
                        except Exception as e:
                            log(f"  ⚠️ 이미지 추출 실패 (row fallback): {e}")
                            pass
                    
                    if img_url:
                        log(f"  📷 이미지 URL: {img_url[:80]}...")
                    else:
                        log(f"  ⚠️ 이미지 URL 없음")
                    
                    # 상품 정보
                    product_cell = cells.nth(2)
                    item_info = product_cell.inner_text(timeout=2000)
                    lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                    
                    style_id = ""
                    item_name = ""
                    spu_id = ""
                    
                    for idx, line in enumerate(lines):
                        line_clean = line.strip()
                        
                        if not style_id:
                            if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                                style_id = lines[idx + 1].strip()
                            elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                                style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                        
                        if not spu_id:
                            if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                                spu_id = lines[idx + 1].strip()
                            elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                                spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                        
                        if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                            item_name = line_clean
                    
                    # 가격 정보 (모두 타임아웃 추가)
                    avg_price_raw = cells.nth(5).inner_text(timeout=1000)
                    cn_exposure_raw = cells.nth(6).inner_text(timeout=1000)
                    cn_sales_raw = cells.nth(7).inner_text(timeout=1000)
                    local_sales_raw = cells.nth(8).inner_text(timeout=1000)
                    
                    # 숫자 추출
                    cn_exposure_num = extract_number(cn_exposure_raw)
                    cn_sales_num = extract_number(cn_sales_raw)
                    local_sales_num = extract_number(local_sales_raw)
                    
                    # 검색어와 결과 일치 검증
                    search_code = product['code'].upper().replace('-', '').replace(' ', '')
                    result_code = style_id.upper().replace('-', '').replace(' ', '')
                    
                    if search_code and result_code:
                        if search_code not in result_code and result_code not in search_code:
                            log(f"  ⚠️ 검색어 불일치: 검색={product['code']}, 결과={style_id}")
                            log(f"  → 다음 상품으로 넘어갑니다")
                            continue
                    
                    # 엑셀 데이터와 병합
                    comparison = {
                        "이미지URL": img_url,
                        "상품번호": style_id,
                        "제품명": item_name,
                        "SPU_ID": spu_id,
                        "최근30일평균거래가": avg_price_raw,
                        "중국노출": cn_exposure_raw,
                        "중국노출_숫자": cn_exposure_num,
                        "중국시장최근30일판매량": cn_sales_num,
                        "현지판매자최근30일판매량": local_sales_num,
                        "엑셀_상품번호": product['code'],
                        "엑셀_제품명": product['name'],
                        "엑셀_정가": product['original_price'],
                        "엑셀_할인가": product['sale_price'],
                        "엑셀_재고": product.get('stock', 0),
                        "가격차이": cn_exposure_num - product['sale_price'] if cn_exposure_num > 0 else 0,
                    }
                    
                    results.append(comparison)
                    
                    log(f"  ✅ 찾음: {style_id} | {item_name[:30] if item_name else ''}... | 중국노출: {cn_exposure_raw}")
                    
                    if callback:
                        callback(f"DATA:{json.dumps(comparison, ensure_ascii=False)}", 'data')
                    
                    # 상품 간 랜덤 딜레이 (사람처럼 보이게) - 1~2초
                    delay = random.uniform(1.0, 2.0)
                    log(f"  ⏱️ 다음 상품 대기: {delay:.1f}초")
                    time.sleep(delay)
                
                except Exception as e:
                    log(f"  ❌ 검색 오류: {e}")
                    continue
            
            # 브라우저 닫기
            browser.close()
            
            # 전역 참조 제거
            try:
                import sys
                if hasattr(sys.modules['__main__'], 'current_browser'):
                    sys.modules['__main__'].current_browser = None
            except:
                pass
            
            # 총 소요 시간 계산
            total_elapsed_sec = time_module.time() - total_start_time
            hours = int(total_elapsed_sec // 3600)
            minutes = int((total_elapsed_sec % 3600) // 60)
            seconds = int(total_elapsed_sec % 60)
            
            # 시간 문자열 생성
            time_str = ""
            if hours > 0:
                time_str = f"{hours}시간 {minutes}분 {seconds}초"
            elif minutes > 0:
                time_str = f"{minutes}분 {seconds}초"
            else:
                time_str = f"{seconds}초"
            
            log(f"\n✅ 비교 완료: {len(results)}/{total}")
            log(f"⏱️ 총 검색 시간: {time_str}")
    
    except Exception as e:
        log(f"❌ 전체 오류: {e}")
        # 오류 시에도 브라우저 참조 제거
        try:
            import sys
            if hasattr(sys.modules['__main__'], 'current_browser'):
                sys.modules['__main__'].current_browser = None
        except:
            pass
    
    # 엑셀 저장
    filename = save_comparison_to_excel(results, products)
    
    return {
        'success': True,
        'total_items': len(results),
        'file_path': filename
    }


def save_comparison_to_excel(results, original_products):
    """비교 결과를 엑셀로 저장"""
    now = datetime.now()
    filename = f"poizon_comparison_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가격비교"
    
    headers = ["이미지", "엑셀_상품번호", "상품번호", "제품명", "SPU_ID", 
               "엑셀_정가", "엑셀_할인가", "엑셀_재고", "포이즌_중국노출가", "가격차이",
               "평균거래가", "판매량", "현지판매량"]
    ws.append(headers)
    
    for data in results:
        ws.append([
            "",
            data.get("엑셀_상품번호", ""),
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("엑셀_정가", 0),
            data.get("엑셀_할인가", 0),
            data.get("엑셀_재고", 0),
            data.get("중국노출_숫자", 0),
            data.get("가격차이", 0),
            data.get("최근30일평균거래가", ""),
            data.get("중국시장최근30일판매량", 0),
            data.get("현지판매자최근30일판매량", 0),
        ])
    
    wb.save(filepath)
    log(f"💾 비교 결과 저장: {filepath}")
    
    return filename


if __name__ == "__main__":
    run()


###############################################################
# ============ 네이버 쇼핑 직접 URL 검색 ============== #
###############################################################

def search_naver_shopping_direct(product_code, callback=None, browser_context=None):
    """
    네이버 쇼핑 직접 URL 검색 (2번째 상품부터)
    검색창 사용 없이 URL로 바로 이동
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    def log(msg, level='info'):
        if LOG_CALLBACK:
            LOG_CALLBACK(msg, level)
    
    try:
        log(f"\n🔍 네이버 쇼핑 검색 시작: {product_code}")
        
        # 기존 브라우저 세션 사용
        page = browser_context['page']
        
        # 직접 검색 URL로 이동
        search_url = f"https://search.shopping.naver.com/search/all?query={product_code}"
        log(f"  📍 직접 URL 접속: {search_url[:80]}...")
        
        page.goto(search_url, wait_until="domcontentloaded", timeout=15000)
        
        # 페이지 로딩 대기
        wait_time = random.randint(3000, 5000)
        log(f"  ⏱️ 페이지 로딩 대기: {wait_time/1000:.1f}초")
        page.wait_for_timeout(wait_time)
        
        # 현재 URL 확인
        current_url = page.url
        log(f"  🔍 현재 URL: {current_url[:80]}...")
        
        if "security" in current_url.lower() or "error" in current_url.lower():
            log(f"  ❌ 접속 제한 감지")
            return {
                'success': False,
                'product_code': product_code,
                'error': '접속 제한'
            }
        
        log(f"  ✅ 검색 완료!")
        
        # TODO: 상품 목록 파싱
        
        # 임시 테스트 데이터
        products = [
            {
                'name': f'{product_code} 테스트 상품',
                'price': '테스트',
                'mall': '테스트',
                'link': current_url
            }
        ]
        
        log(f"  📊 상품 정보: {len(products)}개 (테스트 데이터)")
        log(f"  🏁 검색 완료!")
        
        return {
            'success': True,
            'product_code': product_code,
            'products': products,
            'total': len(products),
            'final_url': current_url
        }
        
    except Exception as e:
        log(f"  ❌ 오류 발생: {e}", 'error')
        return {
            'success': False,
            'product_code': product_code,
            'error': str(e)
        }


###############################################################
# ============ 구매처 검색 메인 함수 ============== #
###############################################################

def run_sourcing_for_products(product_codes, callback=None):
    """
    여러 상품의 구매처를 순차적으로 검색
    - 단일 브라우저 세션 사용 (로그인 유지)
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    def log(msg, level='info'):
        if LOG_CALLBACK:
            LOG_CALLBACK(msg, level)
    
    log(f"\n🚀 구매처 검색 시작")
    log(f"📦 총 {len(product_codes)}개 상품")
    
    results = []
    
    # 단일 브라우저 세션 시작
    with sync_playwright() as p:
        log(f"\n🌐 브라우저 시작 중...")
        
        browser = p.chromium.launch(
            headless=HEADLESS,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox',
                '--disable-web-security',
                '--disable-features=IsolateOrigins,site-per-process',
                '--disable-site-isolation-trials',
                '--incognito'  # 시크릿 모드
            ]
        )
        
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='ko-KR',
            timezone_id='Asia/Seoul',
            permissions=['geolocation'],
            geolocation={'latitude': 37.5665, 'longitude': 126.9780},  # 서울
            color_scheme='light',
            # 쿠키 비활성화 (시크릿 모드)
            accept_downloads=True,
            has_touch=False,
            is_mobile=False,
            java_script_enabled=True,
        )
        
        # 강력한 WebDriver 감지 우회 스크립트
        context.add_init_script("""
            // WebDriver 속성 제거
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            
            // Chrome 객체 추가
            window.chrome = {
                runtime: {},
                loadTimes: function() {},
                csi: function() {},
                app: {}
            };
            
            // Permissions 속성 덮어쓰기
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                    Promise.resolve({ state: Notification.permission }) :
                    originalQuery(parameters)
            );
            
            // Plugin 배열 추가
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            
            // Languages 설정
            Object.defineProperty(navigator, 'languages', {
                get: () => ['ko-KR', 'ko', 'en-US', 'en']
            });
            
            // Platform 설정
            Object.defineProperty(navigator, 'platform', {
                get: () => 'Win32'
            });
            
            // Hardware concurrency
            Object.defineProperty(navigator, 'hardwareConcurrency', {
                get: () => 8
            });
            
            // Device memory
            Object.defineProperty(navigator, 'deviceMemory', {
                get: () => 8
            });
            
            // User activation
            Object.defineProperty(navigator, 'userActivation', {
                get: () => ({
                    hasBeenActive: true,
                    isActive: true
                })
            });
            
            // Automation 관련 속성 제거
            delete navigator.__proto__.webdriver;
        """)
        
        page = context.new_page()
        
        # 먼저 네이버 메인으로 이동 (로그인 상태 확인용)
        log(f"\n🌐 네이버 접속 중...")
        page.goto("https://www.naver.com/", wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)
        
        # 네이버 로그인 (이미 로그인되어 있으면 스킵)
        log(f"\n🔐 네이버 로그인 확인...")
        if not perform_naver_login(page):
            log("❌ 네이버 로그인 실패", 'error')
            browser.close()
            return {
                'success': False,
                'error': '로그인 실패'
            }
        
        log(f"✅ 로그인 완료, 검색 시작\n")
        
        # === 첫 번째 상품: 네이버 메인에서 검색 ===
        if len(product_codes) > 0:
            first_code = product_codes[0]
            log(f"\n{'='*60}")
            log(f"[1/{len(product_codes)}] {first_code}")
            log(f"{'='*60}")
            log(f"  💡 첫 번째 상품: 네이버 메인에서 검색")
            
            # 검색 전 자연스러운 대기 (5-10초)
            natural_wait = random.randint(5000, 10000)
            log(f"  ⏱️ 자연스러운 대기: {natural_wait/1000:.1f}초 (봇 감지 회피)")
            page.wait_for_timeout(natural_wait)
            
            # STEP 1: 네이버 메인으로 이동
            log(f"\n  📍 STEP 1: 네이버 메인으로 이동...")
            try:
                page.goto("https://www.naver.com/", wait_until="domcontentloaded", timeout=15000)
                page.wait_for_timeout(random.randint(2000, 4000))
                log(f"  ✅ STEP 1 완료: 네이버 메인 로딩 완료")
            except Exception as e:
                log(f"  ❌ STEP 1 실패: {e}", 'error')
                browser.close()
                return {'success': False, 'error': f'메인 페이지 이동 실패: {e}'}
            
            # STEP 2: 메인 검색창 찾기
            log(f"\n  📍 STEP 2: 메인 검색창 찾기...")
            try:
                search_input = None
                selectors = ['input#query', 'input[name="query"]', 'input.search_input']
                
                for sel in selectors:
                    count = page.locator(sel).count()
                    log(f"     시도: {sel} → {count}개")
                    if count > 0:
                        search_input = sel
                        log(f"  ✅ STEP 2 완료: 검색창 발견 ({sel})")
                        break
                
                if not search_input:
                    log(f"  ❌ STEP 2 실패: 검색창을 찾을 수 없음", 'error')
                    # 스크린샷 저장
                    page.screenshot(path=f"/home/claude/naver_main_fail_{first_code}.png")
                    log(f"  📸 스크린샷 저장: /home/claude/naver_main_fail_{first_code}.png")
                    # 직접 URL로 우회
                    log(f"  🔀 직접 URL로 우회...")
                    page.goto(f"https://search.shopping.naver.com/search/all?query={first_code}")
                    page.wait_for_timeout(random.randint(3000, 5000))
                    # 다음으로 진행
                    results.append({'success': True, 'product_code': first_code, 'products': [], 'total': 0})
                    if len(product_codes) > 1:
                        delay = random.uniform(10.0, 20.0)
                        log(f"\n⏱️ 다음 상품까지 대기: {delay:.1f}초")
                        time.sleep(delay)
                else:
                    # STEP 3: 검색어 입력
                    log(f"\n  📍 STEP 3: 검색어 입력 ('{first_code}')...")
                    try:
                        # 클릭
                        page.click(search_input)
                        page.wait_for_timeout(random.randint(500, 1000))
                        log(f"     ✓ 검색창 클릭")
                        
                        # 기존 내용 지우기
                        page.fill(search_input, '')
                        page.wait_for_timeout(random.randint(300, 500))
                        log(f"     ✓ 기존 내용 지우기")
                        
                        # 타이핑
                        for i, char in enumerate(first_code, 1):
                            page.type(search_input, char, delay=random.randint(100, 200))
                            if i % 3 == 0:  # 3글자마다 로그
                                log(f"     → 입력 중: {first_code[:i]}")
                        
                        page.wait_for_timeout(random.randint(500, 1000))
                        log(f"  ✅ STEP 3 완료: 검색어 입력 완료")
                        
                        # STEP 4: 검색 실행
                        log(f"\n  📍 STEP 4: 검색 실행 (Enter)...")
                        page.press(search_input, 'Enter')
                        page.wait_for_timeout(random.randint(3000, 5000))
                        
                        current_url = page.url
                        log(f"     현재 URL: {current_url[:60]}...")
                        log(f"  ✅ STEP 4 완료: 검색 실행")
                        
                        # STEP 5: 쇼핑탭 클릭
                        log(f"\n  📍 STEP 5: 쇼핑탭 클릭...")
                        shopping_tabs = ['a:has-text("쇼핑")', 'a[href*="shopping"]']
                        tab_found = False
                        shopping_success = False
                        
                        for tab_sel in shopping_tabs:
                            count = page.locator(tab_sel).count()
                            log(f"     시도: {tab_sel} → {count}개")
                            if count > 0:
                                page.click(tab_sel)
                                page.wait_for_timeout(random.randint(3000, 5000))
                                final_url = page.url
                                log(f"     클릭 후 URL: {final_url[:60]}...")
                                
                                # 쇼핑 페이지로 이동했는지 확인
                                if "search.shopping.naver.com" in final_url:
                                    log(f"  ✅ STEP 5 완료: 쇼핑탭 이동 성공!")
                                    shopping_success = True
                                else:
                                    log(f"  ⚠️ 쇼핑 페이지로 이동 안 됨")
                                
                                tab_found = True
                                break
                        
                        if not tab_found:
                            log(f"  ⚠️ 쇼핑탭을 찾을 수 없음")
                        
                        # 쇼핑 페이지 이동 실패 시 직접 URL로
                        if not shopping_success:
                            log(f"  🔀 직접 쇼핑 검색 URL로 이동...")
                            shopping_url = f"https://search.shopping.naver.com/search/all?query={first_code}"
                            page.goto(shopping_url, wait_until="domcontentloaded", timeout=15000)
                            page.wait_for_timeout(random.randint(3000, 5000))
                            final_url = page.url
                            log(f"     최종 URL: {final_url[:60]}...")
                            log(f"  ✅ STEP 5 완료: 직접 URL 이동!")
                        
                        log(f"\n  🎉 메인 검색 전체 완료!")
                        
                    except Exception as e:
                        log(f"  ❌ STEP 3-5 오류: {e}", 'error')
                        import traceback
                        log(f"  상세: {traceback.format_exc()}")
                        # 우회
                        page.goto(f"https://search.shopping.naver.com/search/all?query={first_code}")
                        page.wait_for_timeout(random.randint(3000, 5000))
                    
                    # 첫 번째 상품 결과 저장
                    results.append({
                        'success': True,
                        'product_code': first_code,
                        'products': [],
                        'total': 0
                    })
                    
                    # 첫 검색 후 안정화 대기
                    if len(product_codes) > 1:
                        log(f"\n" + "="*60)
                        log(f"⏸️  첫 번째 검색 완료 - 안정화 대기")
                        log(f"="*60)
                        log(f"💡 네이버에서 검색 결과를 확인하세요")
                        log(f"💡 이상 감지 화면이 나오면 '멈춤' 버튼을 눌러주세요")
                        log(f"💡 정상이면 30초 후 자동으로 다음 상품 검색을 시작합니다")
                        log(f"\n⏱️  30초 대기 중...")
                        
                        # 5초마다 카운트다운
                        for remaining in range(30, 0, -5):
                            time.sleep(5)
                            # 중단 체크
                            import sys
                            if hasattr(sys.modules['__main__'], 'stop_requested'):
                                if sys.modules['__main__'].stop_requested:
                                    log(f"\n⏹️ 사용자가 검색을 중단했습니다")
                                    browser.close()
                                    return {
                                        'success': False,
                                        'error': '사용자 중단',
                                        'total_searched': len(results),
                                        'total_attempted': 1,
                                        'results': results
                                    }
                            log(f"⏱️  남은 시간: {remaining-5}초...")
                        
                        log(f"\n✅ 대기 완료, 다음 상품 검색 시작\n")
                    
            except Exception as e:
                log(f"  ❌ 메인 검색 전체 오류: {e}", 'error')
                import traceback
                log(f"  {traceback.format_exc()}")
                browser.close()
                return {'success': False, 'error': f'메인 검색 실패: {e}'}
        
        # 브라우저 컨텍스트 전달용
        browser_ctx = {'page': page, 'context': context, 'browser': browser}
        
        # 나머지 상품 검색 (2번째부터)
        for idx in range(2, len(product_codes) + 1):
            code = product_codes[idx - 1]
            
            # 중단 요청 체크
            import sys
            if hasattr(sys.modules['__main__'], 'stop_requested'):
                if sys.modules['__main__'].stop_requested:
                    log(f"\n⏹️ 사용자가 검색을 중단했습니다")
                    browser.close()
                    return {
                        'success': False,
                        'error': '사용자 중단',
                        'total_searched': len(results),
                        'total_attempted': idx - 1,
                        'results': results
                    }
            
            log(f"\n{'='*60}")
            log(f"[{idx}/{len(product_codes)}] {code}")
            log(f"{'='*60}")
            
            # 진행 상황 전송
            if callback:
                callback(f"PROGRESS:{idx}/{len(product_codes)}", 'progress')
            
            # 2번째부터는 직접 URL로 검색
            log(f"  💡 {idx}번째 상품: 직접 URL로 검색")
            result = search_naver_shopping_direct(code, callback, browser_context=browser_ctx)
            
            if result['success']:
                results.append(result)
                
                # 결과 데이터 전송
                if callback:
                    callback(f"DATA:{json.dumps(result, ensure_ascii=False)}", 'data')
            else:
                log(f"  ⚠️ 검색 실패: {result.get('error', '알 수 없는 오류')}", 'warning')
            
            # 상품 간 대기 (10-20초로 증가 - 접속 제한 회피)
            if idx < len(product_codes):
                delay = random.uniform(10.0, 20.0)
                log(f"\n⏱️ 다음 상품까지 대기: {delay:.1f}초 (접속 제한 회피)")
                time.sleep(delay)
        
        log(f"\n🔒 브라우저 종료 중...")
        browser.close()
    
    log(f"\n{'='*60}")
    log(f"✅ 구매처 검색 완료!")
    log(f"📊 총 {len(results)}개 상품 검색 성공 / {len(product_codes)}개 시도")
    log(f"{'='*60}")
    
    return {
        'success': True,
        'total_searched': len(results),
        'total_attempted': len(product_codes),
        'results': results
    }

"""
poizon_search.py 파일 맨 끝에 추가할 코드
(기존 코드 아래에 붙여넣기)
"""

###############################################################
# ============ 구매처 검색 결과 엑셀 저장 ============== #
###############################################################

def save_sourcing_results_to_excel(results, reason="정상종료"):
    """
    구매처 검색 결과를 엑셀로 저장
    
    Args:
        results: 검색 결과 리스트 (sourcing_search.py에서 전달)
        reason: 저장 사유 (정상종료, 사용자 중단, 오류 등)
    
    Returns:
        str: 저장된 파일 경로
    """
    if not results:
        log("💾 저장할 데이터가 없습니다.")
        return None
    
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sourcing_result_{timestamp}.xlsx"
    
    # 저장 경로 (현재 파일의 상위 outputs 폴더)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(current_dir, '..', 'outputs')
    
    # outputs 폴더 없으면 OUTPUT_DIR 사용
    if not os.path.exists(outputs_dir):
        outputs_dir = OUTPUT_DIR
    
    filepath = os.path.join(outputs_dir, filename)
    
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "구매처 검색 결과"
    
    # 헤더
    headers = ['#', '상품번호', '검색결과', '쇼핑몰', '상품명', '가격', '배송비', 'URL', '비고']
    ws.append(headers)
    
    # 데이터 입력
    row_num = 1
    for result in results:
        product_code = result.get('product_code', '')
        success = result.get('success', False)
        
        if success:
            products = result.get('products', [])
            
            if products:
                # 상품이 있으면 각 상품을 한 행씩
                for product in products:
                    ws.append([
                        row_num,
                        product_code,
                        '성공',
                        product.get('mall', ''),
                        product.get('name', ''),
                        product.get('price', ''),
                        product.get('shipping', ''),
                        product.get('link', ''),
                        ''
                    ])
                    row_num += 1
            else:
                # 검색 성공했지만 상품 없음
                ws.append([
                    row_num,
                    product_code,
                    '검색결과 없음',
                    '', '', '', '', '',
                    '네이버 쇼핑에서 상품을 찾을 수 없습니다'
                ])
                row_num += 1
        else:
            # 검색 실패
            error = result.get('error', '알 수 없는 오류')
            ws.append([
                row_num,
                product_code,
                '실패',
                '', '', '', '', '',
                f'오류: {error}'
            ])
            row_num += 1
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 열 너비 자동 조정
    column_widths = {
        'A': 8,   # #
        'B': 20,  # 상품번호
        'C': 15,  # 검색결과
        'D': 20,  # 쇼핑몰
        'E': 50,  # 상품명
        'F': 15,  # 가격
        'G': 12,  # 배송비
        'H': 60,  # URL
        'I': 30   # 비고
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 파일 저장
    wb.save(filepath)
    
    log(f"\n{'='*60}")
    log(f"✅ 엑셀 저장 완료!")
    log(f"{'='*60}")
    log(f"📁 파일명: {filename}")
    log(f"📂 경로: {filepath}")
    log(f"📊 저장 사유: {reason}")
    log(f"📦 총 {len(results)}개 상품 데이터 저장")
    log(f"{'='*60}\n")
    
    return filepath

"""
poizon_search.py 파일 맨 끝에 추가할 코드
(기존 코드 아래에 붙여넣기)
"""

###############################################################
# ============ 구매처 검색 결과 엑셀 저장 ============== #
###############################################################

def save_sourcing_results_to_excel(results, reason="정상종료"):
    """
    구매처 검색 결과를 엑셀로 저장
    
    Args:
        results: 검색 결과 리스트 (sourcing_search.py에서 전달)
        reason: 저장 사유 (정상종료, 사용자 중단, 오류 등)
    
    Returns:
        str: 저장된 파일 경로
    """
    if not results:
        log("💾 저장할 데이터가 없습니다.")
        return None
    
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sourcing_result_{timestamp}.xlsx"
    
    # 저장 경로 (현재 파일의 상위 outputs 폴더)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(current_dir, '..', 'outputs')
    
    # outputs 폴더 없으면 OUTPUT_DIR 사용
    if not os.path.exists(outputs_dir):
        outputs_dir = OUTPUT_DIR
    
    filepath = os.path.join(outputs_dir, filename)
    
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "구매처 검색 결과"
    
    # 헤더
    headers = ['#', '상품번호', '검색결과', '쇼핑몰', '상품명', '가격', '배송비', 'URL', '비고']
    ws.append(headers)
    
    # 데이터 입력
    row_num = 1
    for result in results:
        product_code = result.get('product_code', '')
        success = result.get('success', False)
        
        if success:
            products = result.get('products', [])
            
            if products:
                # 상품이 있으면 각 상품을 한 행씩
                for product in products:
                    ws.append([
                        row_num,
                        product_code,
                        '성공',
                        product.get('mall', ''),
                        product.get('name', ''),
                        product.get('price', ''),
                        product.get('shipping', ''),
                        product.get('link', ''),
                        ''
                    ])
                    row_num += 1
            else:
                # 검색 성공했지만 상품 없음
                ws.append([
                    row_num,
                    product_code,
                    '검색결과 없음',
                    '', '', '', '', '',
                    '네이버 쇼핑에서 상품을 찾을 수 없습니다'
                ])
                row_num += 1
        else:
            # 검색 실패
            error = result.get('error', '알 수 없는 오류')
            ws.append([
                row_num,
                product_code,
                '실패',
                '', '', '', '', '',
                f'오류: {error}'
            ])
            row_num += 1
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 열 너비 자동 조정
    column_widths = {
        'A': 8,   # #
        'B': 20,  # 상품번호
        'C': 15,  # 검색결과
        'D': 20,  # 쇼핑몰
        'E': 50,  # 상품명
        'F': 15,  # 가격
        'G': 12,  # 배송비
        'H': 60,  # URL
        'I': 30   # 비고
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 파일 저장
    wb.save(filepath)
    
    log(f"\n{'='*60}")
    log(f"✅ 엑셀 저장 완료!")
    log(f"{'='*60}")
    log(f"📁 파일명: {filename}")
    log(f"📂 경로: {filepath}")
    log(f"📊 저장 사유: {reason}")
    log(f"📦 총 {len(results)}개 상품 데이터 저장")
    log(f"{'='*60}\n")
    
    return filepath