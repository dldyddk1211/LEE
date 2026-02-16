import os
import sys
import json
import random
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import requests
from io import BytesIO

# 로그 콜백 함수
LOG_CALLBACK = None

def log(message, level='info'):
    """터미널 + GUI 로그 출력"""
    print(message)
    if LOG_CALLBACK:
        try:
            LOG_CALLBACK(message, level)
        except:
            pass

############################
# ===== USER CONFIG ===== #
############################

MODE = "REAL"
SEARCH_KEYWORD = "나이키"
HEADLESS = False
POIZON_ID = "sionejj@naver.com"
POIZON_PW = "wnaoddl1!"
LOGIN_URL = "https://seller.poizon.com/"
GOODS_SEARCH_URL = "https://seller.poizon.com/main/goods/search"

# 네이버 로그인 설정
NAVER_ID = ""  # 네이버 아이디 입력
NAVER_PW = ""  # 네이버 비밀번호 입력
NAVER_COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "naver_cookies.json")

LOG_DIR = "logs"
SHOT_DIR = "shots"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_data")
COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "poizon_cookies.json")
TEST_MAX_PAGES = 2
REAL_MAX_PAGES = 20

############################


def setup_logging():
    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"runlog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    log(f"LOG PATH: {log_path}")
    log(f"MODE: {MODE}")
    log(f"SEARCH_KEYWORD: {SEARCH_KEYWORD}")
    return log_path


def safe_screenshot(page, name: str):
    os.makedirs(SHOT_DIR, exist_ok=True)
    path = os.path.join(SHOT_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{name}.png")
    try:
        page.screenshot(path=path, full_page=True)
        log(f"📸 스크린샷 저장: {path}")
    except Exception as e:
        log(f"⚠️ 스크린샷 실패: {e}", 'error')


def wait_for_inputs(page):
    try:
        page.wait_for_load_state("networkidle", timeout=6000)
    except Exception:
        pass
    try:
        page.wait_for_function("document.querySelectorAll('input').length > 0", timeout=6000)
    except Exception:
        pass


def fill_first(page, selectors, value, step_name="fill"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.fill(sel, value)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def click_first(page, selectors, step_name="click"):
    last_err = None
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=2000)
            page.click(sel)
            log(f"✅ {step_name} OK: {sel}", 'success')
            return sel
        except Exception as e:
            last_err = e
    raise last_err


def wait_stable(page, ms=600):
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(ms)


def set_language_korean(page):
    """언어 한국어로 변경"""
    log("\n[1-1] 언어 한국어로 변경 중...")
    try:
        english_found = False
        
        try:
            if page.locator("text=English").count() > 0:
                english_found = True
                log("  → English 버튼 발견 (text)")
        except:
            pass
        
        if not english_found:
            try:
                if page.locator("button:has-text('English')").count() > 0:
                    english_found = True
                    log("  → English 버튼 발견 (button)")
            except:
                pass
        
        if not english_found:
            log("  ✅ 이미 한국어 설정됨", 'success')
            return True
        
        try:
            page.click("text=English", timeout=5000)
        except:
            try:
                page.click("button:has-text('English')", timeout=5000)
            except:
                log("  ⚠️ English 버튼 클릭 실패", 'warning')
                return False
        
        log("  → English 버튼 클릭 완료")
        wait_stable(page, 500)
        
        try:
            page.click("text=한국어", timeout=5000)
        except:
            try:
                page.click("button:has-text('한국어')", timeout=5000)
            except:
                try:
                    page.evaluate("""
                        const koreanBtn = [...document.querySelectorAll('*')]
                            .find(el => el.textContent.includes('한국어'));
                        if (koreanBtn) koreanBtn.click();
                    """)
                except:
                    log("  ⚠️ 한국어 버튼 클릭 실패", 'warning')
                    return False
        
        log("  → 한국어 선택 완료")
        wait_stable(page, 800)
        
        try:
            if page.locator("text=한국어").count() > 0:
                log("  ✅ 한국어 변경 완료", 'success')
                return True
        except:
            pass
        
        log("  ✅ 언어 변경 완료 (확인 불가)", 'success')
        return True
        
    except Exception as e:
        log(f"  ⚠️ 언어 변경 실패: {e}", 'error')
        return False


def try_sort_descending(page):
    log("\n[4-1] 중국 시장 최근 30일 판매량 내림차순 정렬")
    try:
        sales_th = page.locator("th").nth(12)
        icon = sales_th.locator("span.anticon")
        
        if icon.count() > 0:
            icon.last.click()
            page.wait_for_timeout(600)
            page.locator("text=내림차순").first.click()
            page.wait_for_timeout(400)
            page.locator("button:has-text('확인')").first.click()
            page.wait_for_timeout(1000)
            log("  ✅ 내림차순 정렬 완료", 'success')
            return True
    except Exception as e:
        log(f"  ⚠️ 정렬 실패: {e}", 'error')
    return False


def download_image(url):
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            return BytesIO(response.content)
    except Exception:
        pass
    return None


def extract_number(text):
    """텍스트에서 숫자만 추출"""
    if not text:
        return 0
    
    import re
    numbers = re.findall(r'[\d,]+', str(text))
    if numbers:
        return int(numbers[0].replace(',', ''))
    return 0


def scrape_current_page(page):
    rows_data = []
    try:
        page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
        rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
        count = rows.count()
        log(f"  → 현재 페이지 행 수: {count}")

        for i in range(count):
            try:
                row = rows.nth(i)
                cells = row.locator("td")
                
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception:
                    pass
                
                if not img_url:
                    try:
                        imgs = row.locator("img")
                        if imgs.count() > 0:
                            img_url = imgs.first.get_attribute("src")
                    except Exception:
                        pass

                try:
                    product_cell = cells.nth(2)
                    item_info = product_cell.inner_text()
                    lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                    
                    style_id = ""
                    item_name = ""
                    spu_id = ""
                    
                    for idx, line in enumerate(lines):
                        line_clean = line.strip()
                        
                        if not style_id:
                            if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                                style_id = lines[idx + 1].strip()
                            elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                                style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                        
                        if not spu_id:
                            if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                                spu_id = lines[idx + 1].strip()
                            elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                                spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                        
                        if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                            item_name = line_clean
                    
                    log(f"  ✨ 최종 결과 - 상품번호:'{style_id}' / 제품명:'{item_name}' / SPU:'{spu_id}'")
                
                except Exception as e:
                    log(f"    상품 정보 파싱 오류: {e}", 'error')
                    style_id = ""
                    item_name = ""
                    spu_id = ""

                brand_category = cells.nth(3).inner_text().replace("\n", " / ")
                status = cells.nth(4).inner_text()
                avg_price = cells.nth(5).inner_text()
                cn_exposure = cells.nth(6).inner_text()
                cn_sales_raw = cells.nth(7).inner_text()
                local_sales_raw = cells.nth(8).inner_text()
                
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)

                rows_data.append({
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "상태": status,
                    "최근30일평균거래가": avg_price,
                    "중국노출": cn_exposure,
                    "중국시장최근30일판매량": cn_sales_num,
                    "중국시장최근30일판매량_원본": cn_sales_raw,
                    "현지판매자최근30일판매량": local_sales_num,
                    "현지판매자최근30일판매량_원본": local_sales_raw,
                })
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"DATA:{json.dumps(rows_data[-1], ensure_ascii=False)}", 'data')
                    except:
                        pass
                
                log(f"    [{i+1}/{count}] ✓ {style_id} | {item_name[:30]}... | 판매량:{cn_sales_num}")

            except Exception as e:
                log(f"    [{i+1}/{count}] ✗ 파싱 실패: {e}", 'error')
                continue

    except Exception as e:
        log(f"⚠️ scrape_current_page 오류: {e}", 'error')

    return rows_data


def save_to_excel(all_data, keyword):
    now = datetime.now()
    filename = f"poizon_{keyword}_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검색결과"
    
    headers = ["이미지", "상품번호", "제품명", "SPU_ID", "상태", 
               "최근30일평균거래가", "중국노출", "중국시장최근30일판매량", "현지판매자최근30일판매량"]
    ws.append(headers)
    ws.row_dimensions[1].height = 20

    row_num = 2
    for data in all_data:
        ws.append([
            "",
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("상태", ""),
            data.get("최근30일평균거래가", ""),
            data.get("중국노출", ""),
            data.get("중국시장최근30일판매량", ""),
            data.get("현지판매자최근30일판매량", ""),
        ])

        img_url = data.get("이미지URL", "")
        if img_url:
            try:
                img_data = download_image(img_url)
                if img_data:
                    img = XLImage(img_data)
                    img.width = 60
                    img.height = 60
                    ws.add_image(img, f'A{row_num}')
                    ws.row_dimensions[row_num].height = 50
            except Exception:
                pass
        row_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 25

    wb.save(filepath)
    log(f"💾 엑셀 저장 완료: {filepath} ({len(all_data)}행)", 'success')
    return filepath


def run(keyword=None, max_pages=None, callback=None, skip_login=False):
    global LOG_CALLBACK, SEARCH_KEYWORD, REAL_MAX_PAGES
    
    import time as time_module
    total_start_time = time_module.time()
    
    if callback:
        LOG_CALLBACK = callback
    
    if keyword:
        SEARCH_KEYWORD = keyword
    
    if max_pages:
        REAL_MAX_PAGES = int(max_pages)
    
    setup_logging()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context()
        
        need_login = True
        if skip_login and os.path.exists(COOKIE_FILE):
            try:
                with open(COOKIE_FILE, 'r') as f:
                    cookies = json.load(f)
                context.add_cookies(cookies)
                log("\n[1] 저장된 쿠키 로드 완료", 'success')
                need_login = False
            except Exception as e:
                log(f"  ⚠️ 쿠키 로드 실패: {e}", 'error')
                need_login = True
        
        page = context.new_page()

        try:
            if need_login:
                log("\n[1] 접속: " + LOGIN_URL)
                page.goto(LOGIN_URL, wait_until="domcontentloaded")
                wait_stable(page, 500)

                lang_success = set_language_korean(page)
                if not lang_success:
                    raise Exception("언어 변경 실패")
                
                log("\n[2] 로그인 시작")
                wait_for_inputs(page)
                page.locator("input").nth(0).fill(POIZON_ID)
                log("  → ID 입력 완료")
                page.locator("input").nth(1).fill(POIZON_PW)
                log("  → PW 입력 완료")

                try:
                    click_first(page, ["button:has-text('로그인')"], "로그인")
                except Exception:
                    page.locator("input").nth(1).press("Enter")

                wait_stable(page, 400)
                log("  ✅ 로그인 완료", 'success')
                
                try:
                    cookies = context.cookies()
                    with open(COOKIE_FILE, 'w') as f:
                        json.dump(cookies, f)
                    log("  💾 쿠키 저장 완료", 'success')
                except Exception as e:
                    log(f"  ⚠️ 쿠키 저장 실패: {e}", 'error')
            else:
                log("\n[2] 쿠키로 로그인 건너뛰기")
                page.goto(LOGIN_URL, wait_until="domcontentloaded")
                wait_stable(page, 500)
                
                lang_success = set_language_korean(page)
                if not lang_success:
                    log("  ⚠️ 언어 변경 실패, 계속 진행", 'error')

            log("\n[3] 상품 검색 페이지 이동")
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)

            log(f"\n[4] 키워드 입력: {SEARCH_KEYWORD}")
            wait_for_inputs(page)
            fill_first(page, ["input[placeholder*='상품명']"], SEARCH_KEYWORD, "키워드")

            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except Exception:
                page.keyboard.press("Enter")

            wait_stable(page, 1600)
            try_sort_descending(page)

            max_pages = REAL_MAX_PAGES
            log(f"\n[5] 수집 시작 (최대 {max_pages}페이지)")
            all_data = []
            page_num = 1

            while page_num <= max_pages:
                log(f"\n  📄 페이지 {page_num} 수집 중...")
                
                if LOG_CALLBACK:
                    try:
                        LOG_CALLBACK(f"PROGRESS:{page_num}/{max_pages}", 'progress')
                    except:
                        pass
                
                page_data = scrape_current_page(page)
                all_data.extend(page_data)
                log(f"  ✅ 페이지 {page_num} 완료: {len(page_data)}개 수집 (누계: {len(all_data)}개)", 'success')

                if page_num >= max_pages:
                    log(f"\n  ⏹ 최대 페이지({max_pages}) 도달")
                    break

                try:
                    next_btn = page.locator("li.ant-pagination-next:not(.ant-pagination-disabled) button")
                    if next_btn.count() == 0 or not next_btn.first.is_enabled():
                        log("\n  ⏹ 마지막 페이지 도달")
                        break

                    next_btn.first.click()
                    log(f"  → 다음 페이지로 이동 중...")
                    wait_stable(page, 1500)
                    page_num += 1
                except Exception as e:
                    log(f"\n  ⏹ 다음 페이지 오류: {e}", 'error')
                    break

            log(f"\n[6] 총 {len(all_data)}개 수집 완료", 'success')
            excel_path = save_to_excel(all_data, SEARCH_KEYWORD)
            safe_screenshot(page, "done")
            log("\n=== 완료 ===", 'success')
            
            total_elapsed_sec = time_module.time() - total_start_time
            hours = int(total_elapsed_sec // 3600)
            minutes = int((total_elapsed_sec % 3600) // 60)
            seconds = int(total_elapsed_sec % 60)
            
            time_str = ""
            if hours > 0:
                time_str = f"{hours}시간 {minutes}분 {seconds}초"
            elif minutes > 0:
                time_str = f"{minutes}분 {seconds}초"
            else:
                time_str = f"{seconds}초"
            
            log(f"⏱️ 총 검색 시간: {time_str}", 'success')
            
            return {
                'success': True,
                'data': all_data,
                'total_items': len(all_data),
                'pages': page_num,
                'file_path': os.path.basename(excel_path)
            }

        except Exception as e:
            log(f"⛔ Error: {e}", 'error')
            safe_screenshot(page, "error")
            return {'success': False, 'error': str(e)}
        finally:
            page.wait_for_timeout(1000)
            browser.close()


def perform_login():
    """접속 확인용 - 로그인 후 쿠키 저장"""
    global LOG_CALLBACK
    
    try:
        if os.path.exists(COOKIE_FILE):
            import time
            file_age = time.time() - os.path.getmtime(COOKIE_FILE)
            if file_age < 24 * 3600:
                log("♻️ 이미 로그인되어 있습니다 (쿠키 유효)", 'success')
                return {
                    'success': True,
                    'message': '✅ 이미 접속되어 있습니다 (쿠키 사용)'
                }
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            context = browser.new_context()
            page = context.new_page()
            
            log("\n[1] POIZON 접속 중...", 'info')
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            wait_stable(page, 500)
            
            lang_success = set_language_korean(page)
            if not lang_success:
                raise Exception("언어 변경 실패")
            
            log("\n[2] 로그인 진행 중...", 'info')
            wait_for_inputs(page)
            page.locator("input").nth(0).fill(POIZON_ID)
            log("  → ID 입력 완료")
            page.locator("input").nth(1).fill(POIZON_PW)
            log("  → PW 입력 완료")

            try:
                click_first(page, ["button:has-text('로그인')"], "로그인")
            except Exception:
                page.locator("input").nth(1).press("Enter")

            wait_stable(page, 400)
            log("  ✅ 로그인 완료", 'success')
            
            cookies = context.cookies()
            with open(COOKIE_FILE, 'w') as f:
                json.dump(cookies, f)
            log("  💾 쿠키 저장 완료", 'success')
            
            page.wait_for_timeout(1000)
            browser.close()
            
            log("\n✅ POIZON 접속 완료! 이제 데이터 수집을 시작할 수 있습니다", 'success')
            
            return {
                'success': True,
                'message': '✅ POIZON 접속 성공'
            }
            
    except Exception as e:
        log(f"⛔ 접속 오류: {e}", 'error')
        return {
            'success': False,
            'message': f'접속 실패: {str(e)}'
        }


def run_poizon_from_gui(keyword, max_pages=20, callback=None, skip_login=False):
    """GUI에서 호출하는 함수"""
    return run(
        keyword=keyword, 
        max_pages=max_pages, 
        callback=callback,
        skip_login=skip_login
    )


###############################################################
# ============ 엑셀 비교 기능 ============== #
###############################################################

def compare_product_price(product_code, product_name, callback=None):
    """단일 상품을 포이즌에서 검색하여 가격 정보 반환"""
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            context = browser.new_context()
            
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, 'r') as f:
                        cookies = json.load(f)
                    context.add_cookies(cookies)
                    log("쿠키 로드 완료")
                except:
                    pass
            
            page = context.new_page()
            
            search_query = product_code if product_code else product_name
            log(f"🔍 검색: {search_query}")
            
            page.goto(GOODS_SEARCH_URL, wait_until="domcontentloaded")
            wait_stable(page, 1000)
            
            wait_for_inputs(page)
            fill_first(page, ["input[placeholder*='상품명']"], search_query, "키워드")
            
            try:
                click_first(page, ["button:has-text('검색 및 입찰')"], "검색")
            except:
                page.keyboard.press("Enter")
            
            wait_stable(page, 1600)
            
            try:
                page.wait_for_selector(".ant-table-tbody tr:not(.ant-table-measure-row)", timeout=8000)
                rows = page.locator(".ant-table-tbody tr:not(.ant-table-measure-row)")
                
                if rows.count() == 0:
                    log(f"  ⚠️ 검색 결과 없음: {search_query}")
                    browser.close()
                    return None
                
                row = rows.nth(0)
                cells = row.locator("td")
                
                img_url = ""
                try:
                    imgs = cells.nth(1).locator("img")
                    if imgs.count() > 0:
                        img_url = imgs.first.get_attribute("src")
                except Exception as e:
                    log(f"  이미지 추출 실패: {e}")
                    pass
                
                product_cell = cells.nth(2)
                item_info = product_cell.inner_text()
                lines = [l.strip() for l in item_info.split("\n") if l.strip()]
                
                style_id = ""
                item_name = ""
                spu_id = ""
                
                for idx, line in enumerate(lines):
                    line_clean = line.strip()
                    
                    if not style_id:
                        if line_clean in ["상품 번호:", "상품번호:", "货号:", "번호:"] and idx + 1 < len(lines):
                            style_id = lines[idx + 1].strip()
                        elif ("상품번호" in line_clean or "货号" in line_clean or "번호" in line_clean) and line_clean not in ["상품 번호:", "상품번호:", "货号:", "번호:"]:
                            style_id = line_clean.replace("상품번호:", "").replace("상품번호：", "").replace("상품번호", "").replace("상품 번호:", "").replace("상품 번호：", "").replace("상품 번호", "").replace("货号:", "").replace("货号：", "").replace("货号", "").replace("번호:", "").replace("번호：", "").replace("번호", "").strip()
                    
                    if not spu_id:
                        if line_clean in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："] and idx + 1 < len(lines):
                            spu_id = lines[idx + 1].strip()
                        elif "SPU" in line_clean.upper() and line_clean not in ["SPU_ID:", "SPU_ID：", "SPU ID:", "SPU:", "SPU："]:
                            spu_id = line_clean.replace("SPU_ID：", "").replace("SPU_ID:", "").replace("SPU ID:", "").replace("SPU_ID", "").replace("SPU ID", "").replace("SPU:", "").replace("SPU：", "").replace("SPU", "").strip()
                    
                    if not item_name and line_clean and line_clean != style_id and "상품번호" not in line_clean and "货号" not in line_clean and "SPU" not in line_clean.upper() and "번호" not in line_clean and line_clean not in [":", "："]:
                        item_name = line_clean
                
                avg_price_raw = cells.nth(5).inner_text()
                cn_exposure_raw = cells.nth(6).inner_text()
                cn_sales_raw = cells.nth(7).inner_text()
                local_sales_raw = cells.nth(8).inner_text()
                
                cn_exposure_num = extract_number(cn_exposure_raw)
                cn_sales_num = extract_number(cn_sales_raw)
                local_sales_num = extract_number(local_sales_raw)
                
                result = {
                    "이미지URL": img_url,
                    "상품번호": style_id,
                    "제품명": item_name,
                    "SPU_ID": spu_id,
                    "최근30일평균거래가": avg_price_raw,
                    "중국노출": cn_exposure_raw,
                    "중국노출_숫자": cn_exposure_num,
                    "중국시장최근30일판매량": cn_sales_num,
                    "현지판매자최근30일판매량": local_sales_num,
                }
                
                log(f"  ✅ 찾음: {style_id} | {item_name[:30] if item_name else ''}... | 중국노출: {cn_exposure_raw}")
                
                browser.close()
                return result
                
            except Exception as e:
                log(f"  ❌ 파싱 오류: {e}")
                browser.close()
                return None
                
    except Exception as e:
        log(f"❌ 검색 오류: {e}")
        return None
def run_excel_comparison(products, callback=None):
    """
    엑셀 리스트의 여러 상품을 포이즌에서 검색하여 비교
    
    Args:
        products: 엑셀에서 읽은 상품 리스트 (dict의 list)
        callback: 진행상황 콜백 함수
    
    Returns:
        dict: {'success': True/False, 'total_items': int, 'file_path': str, 'error': str}
    """
    global LOG_CALLBACK
    
    if callback:
        LOG_CALLBACK = callback
    
    try:
        results = []
        total = len(products)
        
        log(f"\n🔍 총 {total}개 상품 비교 시작", 'info')
        
        for idx, product in enumerate(products, 1):
            # 진행상황 전송
            if callback:
                callback(f"PROGRESS:{idx}/{total}", 'progress')
            
            product_code = product.get('상품번호', '') or product.get('code', '')
            product_name = product.get('상품명', '') or product.get('name', '')
            
            if not product_code and not product_name:
                log(f"[{idx}/{total}] ⚠️ 상품번호/상품명 없음, 건너뜀", 'warning')
                continue
            
            log(f"\n[{idx}/{total}] 🔍 검색: {product_code} - {product_name}", 'info')
            
            # 포이즌에서 검색
            poizon_data = compare_product_price(product_code, product_name, callback=callback)
            
            if poizon_data:
                # 엑셀 데이터 + 포이즌 데이터 결합
                combined = {**product, **poizon_data}
                results.append(combined)
                
                # 실시간 결과 전송
                if callback:
                    try:
                        callback(f"PRODUCT_RESULT:{json.dumps({'product_code': product_code, 'products': [combined]}, ensure_ascii=False)}", 'data')
                    except:
                        pass
                
                log(f"  ✅ 검색 완료: {poizon_data.get('상품명', '')}", 'success')
            else:
                log(f"  ⚠️ 검색 결과 없음", 'warning')
                # 검색 실패해도 원본 데이터는 포함
                combined = {**product, '상품명': '검색 결과 없음'}
                results.append(combined)
        
        # 엑셀 저장
        if results:
            filepath = save_comparison_to_excel(results, products)
            
            log(f"\n✅ 비교 완료!", 'success')
            log(f"📄 총 {len(results)}개 상품 처리", 'success')
            log(f"💾 파일 저장: {filepath}", 'success')
            
            return {
                'success': True,
                'total_items': len(results),
                'file_path': os.path.basename(filepath),
                'results': results
            }
        else:
            return {
                'success': False,
                'error': '검색 결과가 없습니다'
            }
            
    except Exception as e:
        log(f"❌ 비교 중 오류: {e}", 'error')
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }

def save_comparison_to_excel(results, original_products):
    """비교 결과를 엑셀로 저장"""
    now = datetime.now()
    filename = f"poizon_comparison_{now.strftime('%y%m%d_%H%M')}.xlsx"
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가격비교"
    
    headers = ["이미지", "엑셀_상품번호", "상품번호", "제품명", "SPU_ID", 
               "엑셀_정가", "엑셀_할인가", "엑셀_재고", "포이즌_중국노출가", "가격차이",
               "평균거래가", "판매량", "현지판매량"]
    ws.append(headers)
    
    for data in results:
        ws.append([
            "",
            data.get("엑셀_상품번호", ""),
            data.get("상품번호", ""),
            data.get("제품명", ""),
            data.get("SPU_ID", ""),
            data.get("엑셀_정가", 0),
            data.get("엑셀_할인가", 0),
            data.get("엑셀_재고", 0),
            data.get("중국노출_숫자", 0),
            data.get("가격차이", 0),
            data.get("최근30일평균거래가", ""),
            data.get("중국시장최근30일판매량", 0),
            data.get("현지판매자최근30일판매량", 0),
        ])
    
    wb.save(filepath)
    log(f"💾 비교 결과 저장: {filepath}")
    
    return filename


###############################################################
# ============ 구매처 검색 결과 엑셀 저장 ============== #
###############################################################

def save_sourcing_results_to_excel(results, reason="정상종료"):
    """
    구매처 검색 결과를 엑셀로 저장
    (sourcing_search.py의 Selenium 버전에서 호출)
    """
    if not results:
        log("💾 저장할 데이터가 없습니다.")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sourcing_result_{timestamp}.xlsx"
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(current_dir, '..', 'outputs')
    
    if not os.path.exists(outputs_dir):
        outputs_dir = OUTPUT_DIR
    
    filepath = os.path.join(outputs_dir, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "구매처 검색 결과"
    
    headers = ['#', '상품번호', '검색결과', '쇼핑몰', '상품명', '가격', '배송비', 'URL', '비고']
    ws.append(headers)
    
    row_num = 1
    for result in results:
        product_code = result.get('product_code', '')
        success = result.get('success', False)
        
        if success:
            products = result.get('products', [])
            
            if products:
                for product in products:
                    ws.append([
                        row_num,
                        product_code,
                        '성공',
                        product.get('mall', ''),
                        product.get('name', ''),
                        product.get('price', ''),
                        product.get('shipping', '무료배송'),
                        product.get('link', ''),
                        ''
                    ])
                    row_num += 1
            else:
                ws.append([
                    row_num,
                    product_code,
                    '검색결과 없음',
                    '', '', '', '', '',
                    '네이버 쇼핑에서 상품을 찾을 수 없습니다'
                ])
                row_num += 1
        else:
            error = result.get('error', '알 수 없는 오류')
            ws.append([
                row_num,
                product_code,
                '실패',
                '', '', '', '', '',
                f'오류: {error}'
            ])
            row_num += 1
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    column_widths = {
        'A': 8, 'B': 20, 'C': 15, 'D': 20, 'E': 50,
        'F': 15, 'G': 12, 'H': 60, 'I': 30
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(filepath)
    
    log(f"\n{'='*60}")
    log(f"✅ 엑셀 저장 완료!")
    log(f"{'='*60}")
    log(f"📁 파일명: {filename}")
    log(f"📂 경로: {filepath}")
    log(f"📊 저장 사유: {reason}")
    log(f"📦 총 {len(results)}개 상품 데이터 저장")
    log(f"{'='*60}\n")
    
    return filepath


if __name__ == "__main__":
    run()