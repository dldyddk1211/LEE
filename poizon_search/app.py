# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Windows cp949 인코딩 문제 방지
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
import sys
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# .env 파일 로딩 (있으면 환경변수로 적용)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
import os
try:
    _env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if os.path.exists(_env_path):
        with open(_env_path) as _f:
            for _line in _f:
                _line = _line.strip()
                if _line and not _line.startswith('#') and '=' in _line:
                    _k, _v = _line.split('=', 1)
                    os.environ[_k.strip()] = _v.strip()
except Exception:
    pass

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 표준 라이브러리
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
import os
import sys
import json
import time
import uuid
import threading
import queue
import asyncio
from datetime import datetime
from config import paths

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Flask 관련
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
from flask import (
    Flask,
    render_template,
    render_template_string,
    request,
    Response,
    send_file,
    jsonify,
    stream_with_context
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# openpyxl 관련
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter



# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 전역 변수 선언 ✅ 여기!
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# POIZON
browser = None
page = None
context = None

# 무신사 ✅
musinsa_browser = None
musinsa_page = None
musinsa_context = None

# 크림 ✅
kream_browser = None
kream_page = None
kream_context = None

# 기타
log_queue = queue.Queue()


# kream_search 모듈 import
try:
    from kream_data import kream_search
    print("✅ kream_search 모듈 로드 성공")


except ImportError:
    try:
        import kream_data.kream_search as kream_search
        print("✅ kream_search 모듈 로드 성공 (직접 import)")
    except ImportError:
        import sys
        kream_data_path = os.path.join(os.path.dirname(__file__), 'kream_data')
        if os.path.exists(kream_data_path):
            sys.path.insert(0, kream_data_path)
            try:
                import kream_search
                print("✅ kream_search 모듈 로드 성공 (sys.path)")
            except ImportError:
                print("⚠️ kream_search 모듈을 찾을 수 없습니다")
                kream_search = None

try:
    from musinsa_data import musinsa_search
    print("✅ musinsa_search 모듈 로드 성공")
except ImportError:
    try:
        import musinsa_data.musinsa_search as musinsa_search
        print("✅ musinsa_search 모듈 로드 성공 (직접 import)")
    except ImportError:
        import sys
        musinsa_data_path = os.path.join(os.path.dirname(__file__), 'musinsa_data')
        if os.path.exists(musinsa_data_path):
            sys.path.insert(0, musinsa_data_path)
            try:
                import musinsa_search
                print("✅ musinsa_search 모듈 로드 성공 (sys.path)")
            except ImportError:
                print("⚠️ musinsa_search 모듈을 찾을 수 없습니다")
                musinsa_search = None

app = Flask(__name__)

app.secret_key = 'inventory_secret_key_2024!'  # 아무 문자열이나 가능


# ✅ 스케줄러 Blueprint 등록
try:
    from scheduler_data.scheduler_api import scheduler_bp, save_task_to_history
    app.register_blueprint(scheduler_bp)
    print("✅ 스케줄러 등록 성공")
except Exception as e:
    print(f"⚠️ 스케줄러 로드 실패: {e}")
    # 대비용 더미 함수
    def save_task_to_history(task_data):
        print("⚠️ 스케줄러 없이 실행 중")
        return None

# ✅ 재고관리 Blueprint 등록
try:
    from inventory_data.inventory_api import inventory_bp
    app.register_blueprint(inventory_bp)
    print("✅ 재고관리 등록 성공")
except Exception as e:
    print(f"⚠️ 재고관리 로드 실패: {e}")

# ✅ 해외소싱 Blueprint 등록
try:
    from overseas_data.overseas_api import overseas_bp
    app.register_blueprint(overseas_bp)
    print("✅ 해외소싱 등록 성공")
except Exception as e:
    print(f"⚠️ 해외소싱 로드 실패: {e}")

# 캐시 비활성화
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True

@app.after_request
def add_no_cache_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# 전역 변수
log_queue = queue.Queue()
result_data = {}
stop_flag = False
current_browser = None
is_working = False
work_start_time = None
work_type = None
estimated_items = 0
current_items = 0
stop_requested = False

# ==========================================
# 콜백 함수
# ==========================================

def log_callback(message, level='info'):
    global current_items
    
    if message.startswith("PROGRESS:"):
        parts = message.split(":")
        if len(parts) >= 2:
            progress_part = parts[1]
            page_info = None
            total_pages = None
            
            if "|PAGE:" in progress_part:
                main_part, page_part = progress_part.split("|PAGE:")
                current, total = main_part.split("/")
                page_info, total_pages = page_part.split("/")
            else:
                current, total = progress_part.split("/")
            
            current_items = int(current)
            log_data = {
                'type': 'progress',
                'current': int(current),
                'total': int(total)
            }
            
            if page_info and total_pages:
                log_data['page'] = int(page_info)
                log_data['total_pages'] = int(total_pages)
            
            log_queue.put(log_data)

    elif message.startswith("TOTAL_COUNT:"):
        try:
            total = int(message.split(":")[1])
            log_queue.put({'type': 'progress', 'current': 0, 'total': total})
        except:
            pass

    elif message.startswith("PRODUCT_START:"):
        product_code = message.split(":", 1)[1]
        log_queue.put({
            'type': 'product_start',
            'product_code': product_code
        })

    elif message.startswith("PRODUCT_RESULT:"):
        try:
            json_str = message.split(":", 1)[1]
            data = json.loads(json_str)
            log_queue.put({
                'type': 'product_result',
                'product_code': data['product_code'],
                'products': data['products']
            })
        except Exception as e:
            print(f"PRODUCT_RESULT 처리 오류: {e}")

    elif message.startswith("DATA:"):
        try:
            item = json.loads(message.split(":", 1)[1])
            log_queue.put({'type': 'data', 'item': item})
        except Exception as e:
            print(f"DATA 처리 오류: {e}")

    else:
        log_queue.put({
            'type': 'log',
            'message': message,
            'level': level
        })

# ==========================================
# 백그라운드 작업
# ==========================================
def run_scraper(keyword, max_pages, skip_login=False):
    """백그라운드에서 스크래핑 실행"""
    global result_data, stop_flag, is_working, work_start_time, work_type, estimated_items, current_items
    
    is_working = True
    import time as time_module
    work_start_time = time_module.time()
    work_type = 'scraping'
    estimated_items = max_pages
    current_items = 0
    stop_flag = False
    
    try:
        from poizon_data.poizon_search import run_poizon_from_gui
        
        result = run_poizon_from_gui(
            keyword=keyword,
            max_pages=max_pages,
            callback=log_callback,
            skip_login=skip_login
        )
        
        if result.get('success'):
            # ✅ 스케줄러에 저장!
            try:
                end_time = time_module.time()
                duration_seconds = int(end_time - work_start_time)
                
                task_data = {
                    'keyword': keyword,
                    'mode': 'keyword',
                    'collected_count': result.get('total_items', 0),
                    'kream_count': 0,
                    'duration_seconds': duration_seconds,
                    'data': result.get('data', [])
                }
                
                task_id = save_task_to_history(task_data)
                if task_id:
                    print(f"💾 스케줄러 저장 완료: {task_id}")
                else:
                    print(f"⚠️ 스케줄러 저장 실패")
                
            except Exception as e:
                print(f"⚠️ 스케줄러 저장 오류: {e}")
                import traceback
                traceback.print_exc()
            
            # 기존 complete 메시지
            log_queue.put({
                'type': 'complete',
                'total_items': result.get('total_items', 0),
                'pages': result.get('pages', 0),
                'file_path': result.get('file_path', ''),
                'data': result.get('data', [])
            })
            result_data['file_path'] = result.get('file_path', '')
        else:
            log_queue.put({
                'type': 'error',
                'message': result.get('error', '알 수 없는 오류')
            })
    except Exception as e:
        if not stop_flag:
            log_queue.put({
                'type': 'error',
                'message': str(e)
            })
    finally:
        is_working = False
        work_start_time = None
        work_type = None
        estimated_items = 0
        current_items = 0

def run_comparison(products):
    """리스트 비교 작업"""
    global result_data, stop_flag, current_browser, is_working, work_start_time, work_type, estimated_items, current_items
    
    is_working = True
    import time as time_module
    work_start_time = time_module.time()
    work_type = 'comparison'
    estimated_items = len(products)
    current_items = 0
    stop_flag = False
    
    try:
        from poizon_data.poizon_search import run_excel_comparison
        
        result = run_excel_comparison(products, callback=log_callback)
        
        if result.get('success'):
            # ✅ 스케줄러에 저장!
            try:
                end_time = time_module.time()
                duration_seconds = int(end_time - work_start_time)
                
                task_data = {
                    'keyword': f'리스트비교_{len(products)}개',
                    'mode': 'compare',
                    'collected_count': result.get('total_items', 0),
                    'kream_count': 0,
                    'duration_seconds': duration_seconds,
                    'data': result.get('results', [])  # results 필드 사용
                }
                
                task_id = save_task_to_history(task_data)
                if task_id:
                    print(f"💾 스케줄러 저장 완료: {task_id}")
                else:
                    print(f"⚠️ 스케줄러 저장 실패")
                
            except Exception as e:
                print(f"⚠️ 스케줄러 저장 오류: {e}")
                import traceback
                traceback.print_exc()
            
            # 텔레그램 알림
            try:
                from utils.telegram import send_telegram_async
                elapsed_min = int((time_module.time() - work_start_time) // 60)
                elapsed_sec = int((time_module.time() - work_start_time) % 60)
                send_telegram_async(
                    f"✅ <b>리스트 비교 완료</b>\n"
                    f"• 상품 수: {result.get('total_items', 0)}개\n"
                    f"• 소요 시간: {elapsed_min}분 {elapsed_sec}초"
                )
            except Exception:
                pass

            # 기존 complete 메시지
            log_queue.put({
                'type': 'complete',
                'total_items': result.get('total_items', 0),
                'file_path': result.get('file_path', '')
            })
            result_data['file_path'] = result.get('file_path', '')
        else:
            log_queue.put({
                'type': 'error',
                'message': result.get('error', '알 수 없는 오류')
            })
    except Exception as e:
        if not stop_flag:
            log_queue.put({
                'type': 'error',
                'message': str(e)
            })
    finally:
        current_browser = None
        is_working = False
        work_start_time = None
        work_type = None
        estimated_items = 0
        current_items = 0
            

# ==========================================
# 무신사 검색 실행 함수
# ==========================================

def run_musinsa_search(keyword, max_items, search_mode='keyword'):
    """무신사 검색 실행 (스레드)"""
    global stop_flag, is_working, work_start_time, work_type, estimated_items, current_items
    
    print(f"\n{'='*60}")
    print(f"🟤 run_musinsa_search 시작")
    print(f"  keyword: {keyword}")
    print(f"  max_items: {max_items}")
    print(f"  max_items 타입: {type(max_items)}")
    print(f"{'='*60}\n")
    
    is_working = True
    import time as time_module
    work_start_time = time_module.time()
    work_type = 'musinsa'
    
    if max_items == 'max':
        estimated_items = 1000
    else:
        try:
            estimated_items = int(max_items)
        except:
            estimated_items = 10
    
    current_items = 0
    stop_flag = False
    
    try:
        result = musinsa_search.search_musinsa(   # ← keyword_detail → search_musinsa
            keyword=keyword,
            max_items=max_items,
            search_mode=search_mode,
            callback=log_callback
        )
                
        print(f"\n{'='*60}")
        print(f"✅ run_musinsa_search 완료")
        print(f"  성공: {result.get('success')}")
        print(f"  수집 개수: {result.get('total_items', 0)}")
        print(f"{'='*60}\n")
        
        if result.get('success'):
            try:
                end_time = time_module.time()
                duration_seconds = int(end_time - work_start_time)
                task_data = {
                    'keyword': f'무신사_{keyword}',
                    'mode': 'musinsa',
                    'collected_count': result.get('total_items', 0),
                    'kream_count': 0,
                    'duration_seconds': duration_seconds,
                    'data': result.get('results', [])
                }
                task_id = save_task_to_history(task_data)
                if task_id:
                    print(f"💾 스케줄러 저장 완료: {task_id}")
                else:
                    print(f"⚠️ 스케줄러 저장 실패")
            except Exception as e:
                print(f"⚠️ 스케줄러 저장 오류: {e}")
                import traceback
                traceback.print_exc()
            
            # 텔레그램 알림
            try:
                from utils.telegram import send_telegram_async
                elapsed_min = int((time_module.time() - work_start_time) // 60)
                elapsed_sec = int((time_module.time() - work_start_time) % 60)
                send_telegram_async(
                    f"✅ <b>무신사 검색 완료</b>\n"
                    f"• 키워드: {keyword}\n"
                    f"• 수집 수: {result.get('total_items', 0)}개\n"
                    f"• 소요 시간: {elapsed_min}분 {elapsed_sec}초"
                )
            except Exception:
                pass

            log_queue.put({
                'type': 'complete',
                'total_items': result.get('total_items', 0),
                'data': result.get('results', [])
            })
        else:
            log_queue.put({
                'type': 'error',
                'message': result.get('error', '알 수 없는 오류')
            })

    except Exception as e:
        print(f"\n❌ run_musinsa_search 오류: {e}")
        import traceback
        traceback.print_exc()
        if not stop_flag:
            log_queue.put({
                'type': 'error',
                'message': str(e)
            })
    finally:
        is_working = False
        work_start_time = None
        work_type = None
        estimated_items = 0
        current_items = 0


# ==========================================
# 무신사 → 크림 → 포이즌 통합 검색
# ==========================================

def run_full_search(keyword, max_items, search_mode='keyword'):
    """무신사 검색 후 자동으로 크림 → 포이즌 순서로 연속 검색"""
    global stop_flag, is_working, work_start_time, work_type, estimated_items, current_items

    import time as time_module
    from utils.telegram import send_telegram_async

    is_working = True
    work_start_time = time_module.time()
    stop_flag = False
    while not log_queue.empty():
        log_queue.get()

    total_start = time_module.time()
    musinsa_count = 0
    kream_count   = 0
    poizon_count  = 0
    product_codes = []

    # ────────────────────────────────────
    # 1단계: 무신사 검색
    # ────────────────────────────────────
    try:
        work_type = 'musinsa'
        estimated_items = int(max_items) if str(max_items).isdigit() else 100
        current_items = 0

        label = f'랭킹 TOP {max_items}' if search_mode == 'ranking' else f'{keyword} ({max_items}개)'
        log_queue.put({'type': 'log', 'message': f'🟤 [1/3] 무신사 검색 시작: {label}'})
        result = musinsa_search.search_musinsa(keyword=keyword, max_items=max_items, search_mode=search_mode, callback=log_callback)

        if result.get('success'):
            items = result.get('results', [])
            musinsa_count = len(items)
            # 유효한 품번만 추출
            product_codes = [
                it.get('product_code', '').strip()
                for it in items
                if it.get('product_code', '').strip() not in ('', '-')
            ]
            elapsed = int(time_module.time() - total_start)
            send_telegram_async(
                f"✅ <b>[1/3] 무신사 검색 완료</b>\n"
                f"• 키워드: {keyword}\n"
                f"• 수집: {musinsa_count}개 (품번 {len(product_codes)}개)\n"
                f"• 경과: {elapsed//60}분 {elapsed%60}초\n"
                f"→ 크림 검색 시작..."
            )
            log_queue.put({'type': 'log', 'message': f'✅ 무신사 완료 {musinsa_count}개 → 크림 검색 시작'})
        else:
            err = result.get('error', '알 수 없는 오류')
            send_telegram_async(f"❌ <b>무신사 오류</b>\n{err}")
            log_queue.put({'type': 'error', 'message': f'무신사 오류: {err}'})
            return

    except Exception as e:
        send_telegram_async(f"❌ <b>무신사 오류</b>\n{e}")
        log_queue.put({'type': 'error', 'message': f'무신사 오류: {e}'})
        return

    if stop_flag or not product_codes:
        if not product_codes:
            send_telegram_async('⚠️ 품번이 없어 크림/포이즌 검색을 건너뜁니다')
        is_working = False
        return

    # ────────────────────────────────────
    # 2단계: 크림 검색
    # ────────────────────────────────────
    try:
        work_type = 'kream'
        estimated_items = len(product_codes)
        current_items = 0

        log_queue.put({'type': 'log', 'message': f'🛒 [2/3] 크림 검색 시작: {len(product_codes)}개 품번'})
        kream_result = kream_search.search_kream_products_batch(product_codes, callback=log_callback)

        if kream_result.get('success'):
            kream_data = kream_result.get('results', {})
            kream_count = len(kream_data)
            elapsed = int(time_module.time() - total_start)
            send_telegram_async(
                f"✅ <b>[2/3] 크림 검색 완료</b>\n"
                f"• 결과: {kream_count}개\n"
                f"• 경과: {elapsed//60}분 {elapsed%60}초\n"
                f"→ 포이즌 검색 시작..."
            )
            log_queue.put({'type': 'log', 'message': f'✅ 크림 완료 {kream_count}개 → 포이즌 검색 시작'})
        else:
            err = kream_result.get('error', '알 수 없는 오류')
            send_telegram_async(f"❌ <b>크림 오류</b>\n{err}\n→ 포이즌 검색은 계속 진행합니다")
            log_queue.put({'type': 'log', 'message': f'⚠️ 크림 오류: {err}'})

    except Exception as e:
        send_telegram_async(f"❌ <b>크림 오류</b>\n{e}\n→ 포이즌 검색은 계속 진행합니다")
        log_queue.put({'type': 'log', 'message': f'⚠️ 크림 오류: {e}'})

    if stop_flag:
        is_working = False
        return

    # ────────────────────────────────────
    # 3단계: 포이즌 검색
    # ────────────────────────────────────
    try:
        from poizon_data.poizon_search import search_single_product
        work_type = 'poizon'
        estimated_items = len(product_codes)
        current_items = 0

        log_queue.put({'type': 'log', 'message': f'🔴 [3/3] 포이즌 검색 시작: {len(product_codes)}개 품번'})
        poizon_data = {}

        for idx, code in enumerate(product_codes):
            if stop_flag:
                break
            current_items = idx + 1
            log_queue.put({'type': 'log', 'message': f'  포이즌 [{idx+1}/{len(product_codes)}] {code}'})
            try:
                pr = search_single_product(code)
                if pr:
                    poizon_data[code] = pr
                    poizon_count += 1
            except Exception as e:
                log_queue.put({'type': 'log', 'message': f'  ⚠️ 포이즌 {code} 오류: {e}'})

        elapsed = int(time_module.time() - total_start)
        send_telegram_async(
            f"✅ <b>[3/3] 포이즌 검색 완료</b>\n"
            f"• 결과: {poizon_count}개\n"
            f"• 경과: {elapsed//60}분 {elapsed%60}초"
        )
        log_queue.put({'type': 'log', 'message': f'✅ 포이즌 완료 {poizon_count}개'})

    except Exception as e:
        send_telegram_async(f"❌ <b>포이즌 오류</b>\n{e}")
        log_queue.put({'type': 'log', 'message': f'⚠️ 포이즌 오류: {e}'})

    # ────────────────────────────────────
    # 최종 요약
    # ────────────────────────────────────
    total_elapsed = int(time_module.time() - total_start)
    send_telegram_async(
        f"🎉 <b>전체 검색 완료</b>\n"
        f"• 키워드: {keyword}\n"
        f"• 무신사: {musinsa_count}개\n"
        f"• 크림: {kream_count}개\n"
        f"• 포이즌: {poizon_count}개\n"
        f"• 총 소요: {total_elapsed//60}분 {total_elapsed%60}초"
    )
    log_queue.put({
        'type': 'complete',
        'total_items': musinsa_count,
        'message': f'전체 완료 - 무신사 {musinsa_count} / 크림 {kream_count} / 포이즌 {poizon_count}'
    })

    is_working = False
    work_type = None
    estimated_items = 0
    current_items = 0


# ==========================================
# 라우트
# ==========================================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/check_login')
def check_login():
    """POIZON 로그인"""
    print("\n" + "="*60)
    print("🔍 /check_login 호출")
    print("="*60)
    
    try:
        from poizon_data.poizon_search import perform_login
        print("✅ perform_login import 성공")
        
        result = perform_login()
        print(f"✅ perform_login 실행 완료: {result}")
        
        return jsonify({
            'logged_in': result.get('success', False),
            'message': result.get('message', '')
        })
        
    except ImportError as e:
        print(f"❌ ImportError: {e}")
        return jsonify({
            'logged_in': False,
            'message': f'모듈을 찾을 수 없습니다: {str(e)}'
        }), 500
    except Exception as e:
        print(f"❌ Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'logged_in': False,
            'message': f'오류: {str(e)}'
        }), 500

@app.route('/check_status')
def check_status():
    global is_working, work_start_time, work_type, estimated_items, current_items
    
    if not is_working:
        return jsonify({'working': False, 'message': '서버 사용 가능'})
    
    import time
    elapsed_seconds = int(time.time() - work_start_time) if work_start_time else 0
    remaining_minutes = 1
    
    if current_items > 0:
        avg_time = elapsed_seconds / current_items
        remaining = max(0, estimated_items - current_items)
        remaining_minutes = max(1, int(avg_time * remaining) // 60)
    
    work_type_kr = '수집' if work_type == 'scraping' else '비교'
    
    return jsonify({
        'working': True,
        'message': f'서버 작업 중 ({work_type_kr})',
        'work_type': work_type,
        'progress': f'{current_items}/{estimated_items}',
        'elapsed_minutes': elapsed_seconds // 60,
        'remaining_minutes': remaining_minutes
    })

@app.route('/start')
def start_search():
    mode = request.args.get('mode', 'poizon')
    keyword = request.args.get('keyword', '')
    skip_login = request.args.get('skip_login', 'false') == 'true'
    
    # ✅ 무신사 검색 모드 추가
    search_mode = request.args.get('search_mode', 'keyword')  # 'keyword' 또는 'ranking'

    # ✅ max_pages 처리 (POIZON용)
    max_pages_str = request.args.get('max_pages', '1')
    try:
        if max_pages_str and max_pages_str != 'NaN':
            max_pages = int(max_pages_str)
        else:
            max_pages = 1
    except ValueError:
        max_pages = 1

    # ✅ max_items 처리 (무신사용)
    max_items_str = request.args.get('max_items', 'max')
    if max_items_str == 'max':
        max_items = 'max'  # ✅ 문자열 그대로
    else:
        try:
            max_items = int(max_items_str)
        except ValueError:
            max_items = 'max'
    
    print(f"\n{'='*60}")
    print(f"📡 /start 라우트 호출:")
    print(f"  mode: {mode}")
    print(f"  keyword: {keyword}")
    print(f"  search_mode: {search_mode}")
    print(f"  max_pages: {max_pages}")
    print(f"  max_items: {max_items} (타입: {type(max_items)})")
    print(f"  skip_login: {skip_login}")
    print(f"{'='*60}\n")
    
    # mode 처리
    if mode == 'keyword':
        mode = 'poizon'
    
    # 큐 비우기
    while not log_queue.empty():
        log_queue.get()
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 모드별 처리
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    if mode == 'poizon':
        # POIZON 검색 (max_pages 사용)
        print(f"🟣 POIZON 스레드 시작: keyword={keyword}, max_pages={max_pages}")
        thread = threading.Thread(target=run_scraper, args=(keyword, max_pages, skip_login))
        thread.daemon = True
        thread.start()
        
    elif mode == 'musinsa':
        # ✅ 무신사 검색 (브라우저는 musinsa_search 모듈 내부에서 관리)
        print(f"🟤 무신사 스레드 시작: keyword={keyword}, max_items={max_items}, search_mode={search_mode}")
        thread = threading.Thread(target=run_musinsa_search, args=(keyword, max_items, search_mode))
        thread.daemon = True
        thread.start()
    elif mode == 'popular':
        # ✅ 포이즌 신규 인기 상품
        max_popular = int(request.args.get('max_items', '200'))
        brand = request.args.get('brand', '')
        print(f"🔥 인기상품 스레드 시작: max_items={max_popular}, brand={brand}")
        thread = threading.Thread(target=run_popular_scraper, args=(max_popular, brand))
        thread.daemon = True
        thread.start()

    elif mode == 'kream':
        print("🛒 크림 모드 (준비 중)")
        log_queue.put({'type': 'error', 'message': '🛒 크림 검색 기능은 준비 중입니다'})

    else:
        # 기본: POIZON
        print(f"⚠️ 알 수 없는 모드 '{mode}' → POIZON으로 처리")
        thread = threading.Thread(target=run_scraper, args=(keyword, max_pages, skip_login))
        thread.daemon = True
        thread.start()
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SSE 이벤트 생성
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def generate():
        """SSE 이벤트 생성"""
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print("📡 SSE generate 시작")
        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        
        try:
            event_count = 0
            
            while True:
                try:
                    # ✅ 1초 타임아웃으로 큐에서 데이터 가져오기
                    data = log_queue.get(timeout=1)
                    event_count += 1
                    
                    print(f"📩 [{event_count}] SSE 이벤트: {data.get('type', 'unknown')}")
                    
                    # ✅ complete 이벤트일 때 mode 정보 추가
                    if data.get('type') == 'complete':
                        data['mode'] = mode
                        print(f"  ✅ complete 이벤트 (mode: {mode})")
                    
                    # ✅ JSON 직렬화
                    json_data = json.dumps(data, ensure_ascii=False)
                    yield f"data: {json_data}\n\n"
                    
                    # ✅ complete 또는 error면 종료
                    if data.get('type') in ['complete', 'error']:
                        print(f"  🔚 SSE 스트림 종료 (type: {data.get('type')})")
                        break
                        
                except queue.Empty:
                    # ✅ heartbeat
                    yield ": ping\n\n"
                    
        except GeneratorExit:
            print("⚠️ SSE 클라이언트 연결 종료 (GeneratorExit)")
            
        except Exception as e:
            print(f"❌ SSE generate 오류: {e}")
            import traceback
            traceback.print_exc()
            
            error_msg = {
                'type': 'error',
                'message': f'서버 오류: {str(e)}'
            }
            
            try:
                json_data = json.dumps(error_msg, ensure_ascii=False)
                yield f"data: {json_data}\n\n"
            except:
                yield f"data: {{'type':'error','message':'서버 오류'}}\n\n"
        
        finally:
            print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
            print("📡 SSE generate 종료")
            print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    
    # ✅ Response 생성
    response = Response(generate(), mimetype='text/event-stream')
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Connection'] = 'keep-alive'
    
    return response

# ==========================================
# 포이즌 신규 인기 상품
# ==========================================

def run_popular_scraper(max_items=200, brand=''):
    """백그라운드에서 인기 상품 스크래핑"""
    global result_data, stop_flag, is_working, work_start_time, work_type, estimated_items, current_items

    is_working = True
    import time as time_module
    work_start_time = time_module.time()
    work_type = 'popular'
    estimated_items = max_items
    current_items = 0
    stop_flag = False

    try:
        from poizon_data.poizon_search import run_popular_products
        print(f"  [popular] run_popular_products imported OK, brand={brand}")

        result = run_popular_products(
            max_items=max_items,
            brand=brand,
            callback=log_callback
        )
        print(f"  [popular] result: success={result.get('success')}, items={result.get('total_items', 0)}")

        if result.get('success'):
            # 스케줄러에 저장
            try:
                end_time = time_module.time()
                duration_seconds = int(end_time - work_start_time)

                task_data = {
                    'keyword': '신규인기상품',
                    'mode': 'popular',
                    'collected_count': result.get('total_items', 0),
                    'kream_count': 0,
                    'duration_seconds': duration_seconds,
                    'data': result.get('data', [])
                }

                task_id = save_task_to_history(task_data)
                if task_id:
                    print(f"  [popular] saved: {task_id}")
            except Exception as e:
                print(f"  [popular] save error: {e}")

            log_queue.put({
                'type': 'complete',
                'mode': 'popular',
                'total_items': result.get('total_items', 0),
                'file_path': result.get('file_path', ''),
                'data': result.get('data', [])
            })
        else:
            err_msg = result.get('error', '알 수 없는 오류')
            print(f"  [popular] FAIL: {err_msg}")
            log_queue.put({
                'type': 'error',
                'message': err_msg
            })
    except Exception as e:
        import traceback
        print(f"  [popular] EXCEPTION: {e}")
        traceback.print_exc()
        if not stop_flag:
            log_queue.put({
                'type': 'error',
                'message': str(e)
            })
    finally:
        is_working = False
        work_start_time = None
        work_type = None
        estimated_items = 0
        current_items = 0


@app.route('/start_popular_search')
def start_popular_search():
    """포이즌 신규 인기 상품 검색 시작 (SSE)"""
    global is_working

    max_items_str = request.args.get('max_items', '200')
    try:
        max_items = int(max_items_str)
    except ValueError:
        max_items = 200

    print(f"\n{'='*60}")
    print(f"  [popular] start: max_items={max_items}, is_working={is_working}")
    print(f"{'='*60}\n")

    # 이미 작업 중이면 에러
    if is_working:
        def error_gen():
            yield f"data: {json.dumps({'type':'error','message':'이미 다른 작업이 진행 중입니다. 완료 후 다시 시도하세요.'}, ensure_ascii=False)}\n\n"
        return Response(error_gen(), mimetype='text/event-stream',
                       headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    # 큐 비우기
    while not log_queue.empty():
        log_queue.get()

    thread = threading.Thread(target=run_popular_scraper, args=(max_items,))
    thread.daemon = True
    thread.start()

    def generate():
        try:
            event_count = 0
            while True:
                try:
                    data = log_queue.get(timeout=1)
                    event_count += 1

                    if data.get('type') == 'complete':
                        data['mode'] = 'popular'

                    json_data = json.dumps(data, ensure_ascii=False)
                    yield f"data: {json_data}\n\n"

                    if data.get('type') in ['complete', 'error']:
                        print(f"  [popular] SSE end (type={data.get('type')}, events={event_count})")
                        break
                except queue.Empty:
                    yield ": ping\n\n"
        except GeneratorExit:
            print("  [popular] SSE client disconnected")
        except Exception as e:
            print(f"  [popular] SSE error: {e}")
            yield f"data: {json.dumps({'type':'error','message':str(e)}, ensure_ascii=False)}\n\n"

    response = Response(
        stream_with_context(generate()),
        mimetype='text/event-stream'
    )
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    response.headers['Connection'] = 'keep-alive'
    return response


@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '파일이 없습니다'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다'})

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return jsonify({'success': False, 'error': '빈 파일입니다'})

        first_row = all_rows[0]

        # 최대 컬럼 수 (빈 열 제외)
        max_cols = max(len(r) for r in all_rows) if all_rows else 0

        # 1행 텍스트를 헤더 라벨로 반환
        headers = [str(cell).strip() if cell is not None else '' for cell in first_row]

        # 1행이 헤더인지 데이터인지 자동 판별
        # 판별 기준: 1행에 숫자(가격/재고)가 있으면 데이터, 텍스트만 있으면 헤더
        first_row_has_number = False
        for cell in first_row:
            if cell is not None and isinstance(cell, (int, float)):
                first_row_has_number = True
                break

        is_header = not first_row_has_number
        data_start = 1 if is_header else 0

        # 엑셀 원본 그대로 행 데이터 구성 (컬럼 변환 없음)
        products = []
        for row in all_rows[data_start:]:
            if not any(row):
                continue
            cells = []
            for cell in row:
                if cell is None:
                    cells.append('')
                else:
                    cells.append(str(cell).strip())
            products.append(cells)

        return jsonify({
            'success': True,
            'count': len(products),
            'products': products,
            'headers': headers[:max_cols],
            'is_header': is_header,
            'col_count': max_cols,
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/convert_product_format', methods=['POST'])
def convert_product_format():
    """원본 엑셀에서 품번 추출 + 중복 제거 → 비교용 엑셀 생성"""
    try:
        import re
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '파일이 없습니다'})

        file = request.files['file']
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 모든 셀에서 품번 패턴 추출
        raw_codes = []
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val is None:
                    continue
                text = str(cell_val).strip()
                if not text:
                    continue
                # 패턴: 02-415445-102-  260 → 가운데 품번 415445-102 추출
                m = re.match(r'^\d{2}-(\d{4,}-\d{3})', text)
                if m:
                    raw_codes.append(m.group(1))
                else:
                    # 이미 품번 형태인 경우 (예: 415445-102)
                    m2 = re.match(r'^(\d{4,}-\d{3})$', text)
                    if m2:
                        raw_codes.append(m2.group(1))

        if not raw_codes:
            return jsonify({'success': False, 'error': '품번을 찾을 수 없습니다. 02-XXXXXX-XXX 형식이 필요합니다.'})

        original_count = len(raw_codes)
        # 중복 제거 (순서 유지)
        seen = set()
        unique_codes = []
        for code in raw_codes:
            if code not in seen:
                seen.add(code)
                unique_codes.append(code)

        # 비교용 엑셀 생성
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = '품번리스트'
        out_ws.append(['상품번호', '제품명', '정가', '할인가', '재고'])
        for code in unique_codes:
            out_ws.append([code, '', 0, 0, 0])

        # 컬럼 너비
        out_ws.column_dimensions['A'].width = 18
        out_ws.column_dimensions['B'].width = 30

        from datetime import datetime as _dt
        timestamp = _dt.now().strftime('%Y%m%d_%H%M%S')
        filename = f'품번변환_{timestamp}.xlsx'

        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, filename)
        out_wb.save(filepath)

        return jsonify({
            'success': True,
            'filename': filename,
            'original_count': original_count,
            'count': len(unique_codes),
            'codes': unique_codes
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/download_converted/<filename>')
def download_converted(filename):
    """변환된 엑셀 다운로드"""
    try:
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': '잘못된 파일명'}), 400
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs', filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=filename)
        return jsonify({'error': '파일 없음'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/compare_prices', methods=['POST'])
def compare_prices():
    products = request.json.get('products', [])
    
    while not log_queue.empty():
        log_queue.get()
    
    thread = threading.Thread(target=run_comparison, args=(products,))
    thread.daemon = True
    thread.start()
    
    def generate():
        while True:
            try:
                data = log_queue.get(timeout=1)
                yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
                
                if data.get('type') in ['complete', 'error']:
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'ping'})}\n\n"
    
    return Response(generate(), mimetype='text/event-stream')

@app.route('/download/<path:filename>')
def download(filename):
    try:
        # outputs/ 폴더 우선 확인 (download_excel로 저장된 파일)
        file_path_outputs = os.path.join(paths.get('outputs'), filename)
        if os.path.exists(file_path_outputs):
            return send_file(file_path_outputs, as_attachment=True, download_name=filename)

        # poizon_data/output_data/ 폴더 확인 (포이즌 스크래퍼 파일)
        file_path = os.path.join(os.path.dirname(__file__), 'poizon_data', 'output_data', filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)

        return jsonify({'error': '파일을 찾을 수 없습니다'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/shutdown_kream_browser')
def shutdown_kream_browser():
    try:
        if kream_search:
            kream_search.close_kream_browser()
        return jsonify({'success': True}), 200
    except Exception as e:
        print(f"⚠️ 크림 브라우저 종료 오류 (무시): {e}")
        return jsonify({'success': True}), 200

@app.route('/shutdown_browser')
def shutdown_browser():
    try:
        from poizon_data.poizon_search import close_browser_session
        close_browser_session()
        return jsonify({'success': True, 'message': 'POIZON 브라우저 종료 완료'}), 200
    except Exception as e:
        print(f"⚠️ POIZON 브라우저 종료 중 오류 (무시): {e}")
        return jsonify({'success': True, 'message': '이미 종료됨 또는 오류 무시'}), 200

@app.route('/stop', methods=['POST'])
def stop():
    global stop_flag, current_browser
    
    stop_flag = True
    
    try:
        from poizon_data import poizon_search
        poizon_search.stop_flag = True
    except:
        pass
    
    if current_browser:
        try:
            current_browser.close()
        except:
            pass
        current_browser = None
    
    return jsonify({'success': True})

# ==========================================
# 크림 검색
# ==========================================

@app.route('/kream_popup')
def kream_popup():
    """크림 검색 팝업 (자동 시작 지원)"""
    session_id = request.args.get('session_id', None)
    
    if session_id and hasattr(app, 'kream_search_sessions'):
        if session_id in app.kream_search_sessions:
            product_codes = app.kream_search_sessions[session_id]['product_codes']
            return render_template('kream_popup.html', 
                                   auto_start=True, 
                                   product_codes=product_codes)
    
    return render_template('kream_popup.html')

@app.route('/kream_login', methods=['GET', 'POST'])
def kream_login():
    try:
        if kream_search is None:
            return jsonify({'success': False, 'error': 'kream_search 모듈이 없습니다'})

        # GET = 쿠키만 체크 (빠름), POST = 실제 로그인
        if request.method == 'GET':
            cookie_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'kream_data', 'kream_cookies.json')
            if os.path.exists(cookie_file):
                import time as _t
                age = _t.time() - os.path.getmtime(cookie_file)
                if age < 24 * 3600 and os.path.getsize(cookie_file) > 10:
                    return jsonify({'success': True, 'message': '쿠키 유효 (로그인 스킵)'})
            return jsonify({'success': False, 'message': '쿠키 없음 또는 만료'})

        result = kream_search.login_kream()
        return jsonify({'success': result, 'message': '로그인 성공' if result else '로그인 실패'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/search_kream_product', methods=['POST'])
def search_kream_product():
    try:
        if kream_search is None:
            return jsonify({'success': False, 'error': 'kream_search 모듈이 없습니다'})
        
        product_code = request.json.get('product_code', '')
        if not product_code:
            return jsonify({'success': False, 'error': '상품 코드가 없습니다'})
        
        result = kream_search.search_kream_product_detail(product_code)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/start_kream_search', methods=['POST'])
def start_kream_search():
    """크림 검색 시작 - 인라인 방식"""
    try:
        if kream_search is None:
            return jsonify({
                'success': False, 
                'error': 'kream_search 모듈이 로드되지 않았습니다.'
            }), 500
        
        data = request.json
        product_codes = data.get('product_codes', [])
        
        if not product_codes:
            return jsonify({'success': False, 'error': '상품 코드가 없습니다'})
        
        task_id = str(uuid.uuid4())
        
        if not hasattr(app, 'kream_tasks'):
            app.kream_tasks = {}
        
        app.kream_tasks[task_id] = {
            'product_codes': product_codes,
            'status': 'pending'
        }
        
        return jsonify({
            'success': True,
            'task_id': task_id,
            'total_products': len(product_codes)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    

@app.route('/kream_search_progress/<task_id>')
def kream_search_progress(task_id):
    """크림 검색 진행 상황 SSE"""
    
    if not hasattr(app, 'kream_tasks') or task_id not in app.kream_tasks:
        def error_gen():
            yield f"event: error\ndata: {json.dumps({'error': 'Task not found'})}\n\n"
        return Response(error_gen(), mimetype='text/event-stream')
    
    task = app.kream_tasks[task_id]
    product_codes = task['product_codes']
    
    progress_queue = queue.Queue()
    
    def run_background():
        try:
            kream_search.background_kream_search(task_id, product_codes, progress_queue)
        except Exception as e:
            print(f"❌ 백그라운드 검색 오류: {e}")
            import traceback
            traceback.print_exc()
            progress_queue.put({'event': 'error', 'data': {'error': str(e)}})
    
    thread = threading.Thread(target=run_background)
    thread.daemon = True
    thread.start()
    
    def generate():
        try:
            while True:
                try:
                    msg = progress_queue.get(timeout=30)
                    
                    event_type = msg.get('event', 'message')
                    event_data = msg.get('data', {})
                    
                    yield f"event: {event_type}\ndata: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                    
                    if event_type in ['complete', 'error']:
                        if task_id in app.kream_tasks:
                            del app.kream_tasks[task_id]
                        break
                    
                except queue.Empty:
                    yield f"event: ping\ndata: {json.dumps({'status': 'alive'})}\n\n"
                    
        except Exception as e:
            print(f"❌ SSE 스트림 오류: {e}")
            import traceback
            traceback.print_exc()
            yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"
    
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'}
    )

@app.route('/export_kream_to_excel', methods=['POST'])
def export_kream_to_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        data = request.json.get('data', [])
        if not data:
            return jsonify({'success': False, 'error': '데이터가 없습니다'})
        
        wb = Workbook()
        ws = wb.active
        ws.title = "크림 검색 결과"
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        
        headers = ['순번', '상품번호', '제품명', '평균거래가', '중국노출가', 
                   '중국판매량', '현업자판매량', '크림평균가', '크림판매량', '비교']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        
        for row_idx, item in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=str(item.get('순번', row_idx - 1)))
            ws.cell(row=row_idx, column=2, value=str(item.get('상품번호', '')))
            ws.cell(row=row_idx, column=3, value=str(item.get('제품명', '')))
            ws.cell(row=row_idx, column=4, value=str(item.get('평균거래가', '') or item.get('최근30일평균거래가', '')))
            ws.cell(row=row_idx, column=5, value=str(item.get('중국노출가', '') or item.get('중국노출', '')))
            ws.cell(row=row_idx, column=6, value=str(item.get('중국판매량', '') or item.get('중국시장최근30일판매량', '')))
            ws.cell(row=row_idx, column=7, value=str(item.get('현업자판매량', '') or item.get('현지판매자최근30일판매량', '')))
            ws.cell(row=row_idx, column=8, value=str(item.get('크림평균가', '')))
            ws.cell(row=row_idx, column=9, value=str(item.get('크림판매량', '')))
            ws.cell(row=row_idx, column=10, value=str(item.get('비교', '')))
        
        column_widths = [8, 15, 40, 15, 15, 12, 15, 15, 12, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'kream_search_{timestamp}.xlsx'
        
        output_dir = paths.get('outputs')
        os.makedirs(output_dir, exist_ok=True)

        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        wb.close()
        
        return jsonify({'success': True, 'filename': filename, 'count': len(data)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download_kream/<filename>')
def download_kream(filename):
    try:
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': '잘못된 파일명입니다'}), 400
        
        file_path = os.path.join(os.path.dirname(__file__), 'kream_data', 'output_data', filename)
        
        if os.path.exists(file_path):
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({'error': '파일을 찾을 수 없습니다'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==========================================
# 구매처 검색
# ==========================================

@app.route('/start_sourcing', methods=['POST'])
def start_sourcing():
    """크림 검색 시작"""
    product_codes = request.json.get('product_codes', [])
    
    if not product_codes:
        return jsonify({'success': False, 'error': '상품번호가 없습니다'})
    
    session_id = str(uuid.uuid4())
    
    if not hasattr(app, 'kream_search_sessions'):
        app.kream_search_sessions = {}
    
    app.kream_search_sessions[session_id] = {
        'product_codes': product_codes,
        'status': 'pending'
    }
    
    return jsonify({
        'success': True,
        'session_id': session_id,
        'popup_url': f'/kream_popup?session_id={session_id}'
    })


@app.route('/shutdown_musinsa_browser')
def shutdown_musinsa_browser():
    try:
        if musinsa_search:
            musinsa_search.close_musinsa_browser()
        return jsonify({'success': True}), 200
    except Exception as e:
        print(f"⚠️ 무신사 브라우저 종료 오류 (무시): {e}")
        return jsonify({'success': True}), 200


# ✅ 무신사 로그인 라우트 추가
@app.route('/check_musinsa_login')
def check_musinsa_login():
    """무신사 로그인"""
    print("\n" + "="*60)
    print("🔍 /check_musinsa_login 호출")
    print("="*60)
    
    try:
        if musinsa_search is None:
            return jsonify({
                'logged_in': False,
                'message': 'musinsa_search 모듈을 찾을 수 없습니다'
            }), 500
        
        result = musinsa_search.login_musinsa()
        print(f"✅ login_musinsa 실행 완료: {result}")
        
        return jsonify({
            'logged_in': result.get('success', False),
            'message': result.get('message', '')
        })
        
    except Exception as e:
        print(f"❌ Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'logged_in': False,
            'message': f'오류: {str(e)}'
        }), 500
    
# ✅ POIZON 로그인 라우트 추가
@app.route('/check_poizon_login')
def check_poizon_login():
    """POIZON 로그인"""
    print("\n" + "="*60)
    print("🔍 /check_poizon_login 호출")
    print("="*60)
    
    try:
        from poizon_data.poizon_search import perform_login
        print("✅ perform_login import 성공")
        
        result = perform_login()
        print(f"✅ perform_login 실행 완료: {result}")
        
        return jsonify({
            'logged_in': result.get('success', False),
            'message': result.get('message', '')
        })
        
    except ImportError as e:
        print(f"❌ ImportError: {e}")
        return jsonify({
            'logged_in': False,
            'message': f'모듈을 찾을 수 없습니다: {str(e)}'
        }), 500
    except Exception as e:
        print(f"❌ Exception: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'logged_in': False,
            'message': f'오류: {str(e)}'
        }), 500


# ==========================================
# 무신사→포이즌 검색 API
# ==========================================

@app.route('/start_poizon_search', methods=['POST'])
def start_poizon_search():
    """무신사 상품번호로 포이즌 검색 시작"""
    try:
        data = request.json
        product_codes = data.get('product_codes', [])
        
        if not product_codes:
            return jsonify({'success': False, 'error': '상품번호가 없습니다'}), 400
        
        task_id = str(uuid.uuid4())
        
        if not hasattr(app, 'poizon_tasks'):
            app.poizon_tasks = {}
        
        app.poizon_tasks[task_id] = {
            'product_codes': product_codes,
            'status': 'pending',
            'queue': queue.Queue()
        }
        
        return jsonify({
            'success': True,
            'task_id': task_id,
            'total_products': len(product_codes)
        })
        
    except Exception as e:
        print(f"❌ start_poizon_search 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/poizon_search_progress/<task_id>')
def poizon_search_progress(task_id):
    """포이즌 검색 진행 상황 SSE"""
    
    if not hasattr(app, 'poizon_tasks') or task_id not in app.poizon_tasks:
        def error_gen():
            yield f"event: error\ndata: {json.dumps({'error': 'Task not found'})}\n\n"
        return Response(error_gen(), mimetype='text/event-stream')
    
    task = app.poizon_tasks[task_id]
    product_codes = task['product_codes']
    progress_queue = task['queue']
    
    def run_background():
        try:
            from poizon_data import poizon_search
            poizon_search.search_multiple_products(product_codes, progress_queue)
        except Exception as e:
            print(f"❌ 백그라운드 검색 오류: {e}")
            import traceback
            traceback.print_exc()
            progress_queue.put({'event': 'error', 'data': {'error': str(e)}})
    
    thread = threading.Thread(target=run_background)
    thread.daemon = True
    thread.start()
    
    def generate():
        try:
            while True:
                try:
                    msg = progress_queue.get(timeout=30)
                    
                    event_type = msg.get('event', 'message')
                    event_data = msg.get('data', {})
                    
                    yield f"event: {event_type}\ndata: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                    
                    if event_type in ['complete', 'error']:
                        if task_id in app.poizon_tasks:
                            del app.poizon_tasks[task_id]
                        break
                    
                except queue.Empty:
                    yield f"event: ping\ndata: {json.dumps({'status': 'alive'})}\n\n"
                    
        except Exception as e:
            print(f"❌ SSE 스트림 오류: {e}")
            import traceback
            traceback.print_exc()
            yield f"event: error\ndata: {json.dumps({'error': str(e)})}\n\n"
    
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive'
        }
    )

@app.route('/api/send_best', methods=['POST'])
def api_send_best():
    """필터 결과 베스트 10을 텔레그램으로 전송"""
    try:
        from utils.telegram import send_telegram_async
        from datetime import datetime as _dt
        data = request.get_json()
        mode      = data.get('mode', '')
        auto_type = data.get('auto_type', '')   # 'kream' | 'poizon' | 'jordan' | ''
        items     = data.get('items', [])[:10]

        now = _dt.now().strftime('%m/%d %H:%M')
        mode_label = {
            'poizon': '포이즌', 'musinsa': '무신사', 'compare': '리스트비교',
            'musinsa_kream': '무신사', 'compare_kream': '리스트비교',
            'musinsa_poizon': '무신사', 'compare_poizon': '리스트비교',
        }.get(mode, mode)

        # 조건 만족 상품 없음 메시지
        if auto_type in ('kream_empty', 'poizon_empty'):
            label = '🛒 크림' if auto_type == 'kream_empty' else '🟣 포이즌'
            send_telegram_async(
                f'{label} <b>베스트 ({mode_label}) — {now}</b>\n'
                f'⚠️ 조건 만족 상품 없음 (판매량 50개+, 마진 5,000원+)'
            )
            return jsonify({'ok': True})

        # 헤더 (auto_type에 따라 구분)
        if auto_type == 'kream':
            header = f'🛒 <b>크림 자동 베스트 TOP{len(items)} ({mode_label}) — {now}</b>\n<i>판매량 100개+, 마진 10,000원+</i>\n'
        elif auto_type == 'poizon':
            header = f'🟣 <b>포이즌 자동 베스트 TOP{len(items)} ({mode_label}) — {now}</b>\n<i>판매량 100개+, 마진 10,000원+</i>\n'
        elif auto_type == 'jordan':
            header = f'👟 <b>조던 베스트 TOP{len(items)} ({mode_label}) — {now}</b>\n'
        else:
            header = f'🏆 <b>베스트 {len(items)}위 ({mode_label}) — {now}</b>\n'

        lines = [header]

        for i, d in enumerate(items, 1):
            code = d.get('상품번호') or d.get('엑셀_상품번호') or d.get('product_code') or '-'
            name = d.get('제품명') or d.get('name') or '-'
            name = name[:20] + '…' if len(str(name)) > 20 else name

            if auto_type == 'kream':
                kd = int(d.get('크림비교', 0) or 0)
                ks = int(d.get('크림판매량', 0) or 0)
                lines.append(
                    f'{i}. <code>{code}</code> {name}\n'
                    f'   크림마진: <b>+{kd:,}원</b> | 판매량: <b>{ks:,}개</b>'
                )
            elif auto_type == 'poizon':
                pd_ = int(d.get('포이즌비교', 0) or 0)
                ps  = (int(d.get('포이즌중국판매량', 0) or 0) +
                       int(d.get('포이즌현업자판매량', 0) or 0))
                lines.append(
                    f'{i}. <code>{code}</code> {name}\n'
                    f'   포이즌마진: <b>+{pd_:,}원</b> | 판매량: <b>{ps:,}개</b>'
                )
            elif mode == 'poizon':
                diff  = int(d.get('크림비교', 0) or 0)
                ksale = int(d.get('크림판매량', 0) or 0)
                lines.append(
                    f'{i}. <code>{code}</code> {name}\n'
                    f'   포이즌-크림: <b>+{diff:,}원</b> | 크림판매량: {ksale:,}개'
                )
            else:
                kd  = int(d.get('크림비교', 0)  or 0)
                pd_ = int(d.get('포이즌비교', 0) or 0)
                ks  = int(d.get('크림판매량', 0) or 0)
                ps  = (int(d.get('포이즌중국판매량', 0) or 0) +
                       int(d.get('포이즌현업자판매량', 0) or 0))
                lines.append(
                    f'{i}. <code>{code}</code> {name}\n'
                    f'   크림차: <b>{kd:+,}원</b> ({ks:,}개) | '
                    f'포이즌차: <b>{pd_:+,}원</b> ({ps:,}개)'
                )

        send_telegram_async('\n'.join(lines))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/download_excel', methods=['POST'])
def download_excel():
    """현재 테이블 그대로 엑셀 저장"""
    try:
        data = request.json
        mode = data.get('mode', 'poizon')
        keyword = data.get('keyword', '검색결과')
        collected_data = data.get('data', [])
        
        if not collected_data:
            return jsonify({'success': False, 'error': '데이터가 없습니다'})
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        mode_names = {
            'poizon': 'POIZON',
            'musinsa': '무신사',
            'compare': '리스트비교',
            'kream': '크림'
        }
        mode_name = mode_names.get(mode, 'DATA')
        
        safe_keyword = keyword.replace('/', '_').replace('\\', '_')[:30]
        filename = f"{mode_name}_{safe_keyword}_{timestamp}.xlsx"
        
        output_dir = paths.get('outputs')
        os.makedirs(output_dir, exist_ok=True)

        filepath = os.path.join(output_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "수집결과"
        
        if mode == 'compare':
            headers = [
                '#', '이미지', '상품번호', '제품명', '정가', '할인가', '재고',
                '크림평균가', '크림비교', '크림판매량',
                '중국노출가', '포이즌비교', '중국판매량', '현업자판매량'
            ]
        elif mode == 'poizon':
            headers = [
                '#', '이미지', '상품번호', '제품명', '평균거래가',
                '중국노출가', '중국판매량', '현업자판매량',
                '크림평균가', '크림비교'
            ]
        elif mode == 'musinsa':
            headers = [
                '#', '이미지', '상품번호', '제품명', '최대혜택가',
                '크림평균가', '크림비교', '크림판매량',
                '포이즌노출가', '포이즌비교', '중국판매량', '현업자판매량'
            ]
        else:
            headers = [
                '#', '이미지', '상품번호', '제품명', '평균거래가',
                '중국노출가', '판매량'
            ]
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for idx, item in enumerate(collected_data, 1):
            if mode == 'compare':
                row = [
                    idx,
                    item.get('포이즌이미지URL', item.get('크림이미지URL', item.get('이미지URL', '-'))),
                    item.get('엑셀_상품번호', item.get('상품번호', '-')),
                    item.get('제품명', '-'),
                    item.get('엑셀_정가', '-'),
                    item.get('엑셀_할인가', '-'),
                    item.get('엑셀_재고', '-'),
                    item.get('크림평균가', '-'),
                    item.get('크림비교', '-'),
                    item.get('크림판매량', '-'),
                    item.get('포이즌노출가', item.get('중국노출', '-')),
                    item.get('포이즌비교', '-'),
                    item.get('포이즌중국판매량', 0),
                    item.get('포이즌현업자판매량', 0)
                ]
            elif mode == 'poizon':
                row = [
                    idx,
                    item.get('이미지URL', item.get('image_url', '-')),
                    item.get('상품번호', '-'),
                    item.get('제품명', '-'),
                    item.get('최근30일평균거래가', '-'),
                    item.get('중국노출', '-'),
                    item.get('중국시장최근30일판매량', 0),
                    item.get('현지판매자최근30일판매량', 0),
                    item.get('크림평균가', '-'),
                    item.get('크림비교', '-')
                ]
            elif mode == 'musinsa':
                row = [
                    idx,
                    item.get('이미지URL', item.get('image_url', '-')),
                    item.get('product_code', '-'),
                    item.get('name', '-'),
                    item.get('price', '-'),
                    item.get('크림평균가', '-'),
                    item.get('크림비교', '-'),
                    item.get('크림판매량', '-'),
                    item.get('포이즌노출가', '-'),
                    item.get('포이즌비교', '-'),
                    item.get('포이즌중국판매량', 0),
                    item.get('포이즌현업자판매량', 0)
                ]
            else:
                row = [
                    idx,
                    item.get('이미지URL', item.get('image_url', '-')),
                    item.get('상품번호', '-'),
                    item.get('제품명', '-'),
                    item.get('최근30일평균거래가', '-'),
                    item.get('중국노출', '-'),
                    item.get('중국시장최근30일판매량', 0)
                ]
            
            ws.append(row)
        
        column_widths = {
            '#': 8, '이미지': 50, '상품번호': 20, '제품명': 30,
            '정가': 15, '할인가': 15, '재고': 10,
            '평균거래가': 15, '최대혜택가': 15,
            '중국노출가': 15, '포이즌노출가': 15,
            '크림평균가': 15, '크림비교': 15, '포이즌비교': 15,
            '중국판매량': 15, '현업자판매량': 15, '크림판매량': 15, '판매량': 15
        }
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            width = column_widths.get(header, 15)
            ws.column_dimensions[col_letter].width = width
        
        image_col = headers.index('이미지') + 1 if '이미지' in headers else None
        
        if image_col:
            for row_num in range(2, len(collected_data) + 2):
                cell = ws.cell(row=row_num, column=image_col)
                url_value = cell.value
                
                if url_value and url_value != '-' and isinstance(url_value, str) and url_value.startswith('http'):
                    cell.hyperlink = url_value
                    cell.value = '🔗 이미지'
                    cell.font = Font(color="0563C1", underline="single")
        
        wb.save(filepath)

        if os.path.exists(filepath):
            print(f"✅ 파일 저장 성공: {filepath}")
        else:
            print(f"❌ 파일 저장 실패: {filepath}")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'count': len(collected_data)
        })
        
    except Exception as e:
        print(f"엑셀 다운로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download/<filename>')
def download_file(filename):
    """파일 다운로드"""
    try:
        filepath = os.path.join(paths.get('outputs'), filename)
        
        if not os.path.exists(filepath):
            print(f"❌ 파일 없음: {filepath}")
            return jsonify({'error': '파일을 찾을 수 없습니다'}), 404
        
        print(f"✅ 파일 전송: {filepath}")
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ 다운로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/search_wholesale_product', methods=['POST'])
def search_wholesale_product():
    """도매 오더폼 - 포이즌에서 단건 상품 검색 (이미지 + 품명)"""
    try:
        code = request.json.get('code', '').strip()
        if not code:
            return jsonify({'success': False, 'error': '품번 없음'})
        from poizon_data import poizon_search as ps
        result = ps.search_single_product(code)
        if result:
            return jsonify({'success': True, 'img_url': result.get('img_url', ''), 'name': result.get('name', '')})
        return jsonify({'success': False, 'error': '검색 결과 없음 (포이즌 쿠키 확인 필요)'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==========================================
# 설정 (계정 관리)
# ==========================================

SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'settings.json')

def load_settings():
    """설정 파일 로드 (없으면 기본값)"""
    default = {
        'naver': {
            'accounts': [
                {'id': 'dldyddk1211', 'pw': 'dhkdl4213!'},
                {'id': '', 'pw': ''}
            ],
            'selected': 0
        },
        'poizon': {
            'accounts': [
                {'id': 'sionejj@naver.com', 'pw': 'wnaoddl1!'},
                {'id': '', 'pw': ''}
            ],
            'selected': 0
        },
        'kream': {
            'accounts': [
                {'id': 'yaglobal@naver.com', 'pw': 'dyddk1309!'},
                {'id': '', 'pw': ''}
            ],
            'selected': 0
        },
        'musinsa': {
            'accounts': [
                {'id': 'yaglobal', 'pw': 'dyddk1309!'},
                {'id': '', 'pw': ''}
            ],
            'selected': 0
        }
    }
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return default

def save_settings(data):
    """설정 파일 저장"""
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_active_account(site):
    """특정 사이트의 현재 선택된 계정 반환"""
    settings = load_settings()
    site_data = settings.get(site, {})
    accounts = site_data.get('accounts', [])
    selected = site_data.get('selected', 0)
    if selected < len(accounts) and accounts[selected].get('id'):
        return accounts[selected]
    # 선택된 계정이 없으면 첫 번째 유효 계정
    for acc in accounts:
        if acc.get('id'):
            return acc
    return {'id': '', 'pw': ''}

@app.route('/settings')
def settings_page():
    return render_template('settings.html')

# ==========================================
# 입고 체크
# ==========================================

@app.route('/receiving')
def receiving_page():
    return render_template('receiving.html')

@app.route('/api/receiving/save', methods=['POST'])
def receiving_save():
    """입고 스캔 데이터 저장"""
    try:
        data = request.get_json()
        items = data.get('items', [])
        receiving_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'local_data', 'receiving.json')
        os.makedirs(os.path.dirname(receiving_file), exist_ok=True)
        with open(receiving_file, 'w', encoding='utf-8') as f:
            json.dump({'items': items, 'updated': datetime.now().isoformat()}, f, ensure_ascii=False, indent=2)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/receiving/export', methods=['POST'])
def receiving_export():
    """입고 데이터 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        data = request.get_json()
        items = data.get('items', [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '입고체크'

        # 헤더
        headers = ['No.', '상품번호', '수량', '최초스캔', '최종스캔']
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 데이터
        total_qty = 0
        for idx, item in enumerate(items, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=item.get('code', ''))
            qty = item.get('qty', 1)
            ws.cell(row=idx+1, column=3, value=qty)
            ws.cell(row=idx+1, column=4, value=item.get('firstScan', ''))
            ws.cell(row=idx+1, column=5, value=item.get('lastScan', ''))
            total_qty += qty

        # 합계 행
        sum_row = len(items) + 2
        ws.cell(row=sum_row, column=1, value='합계')
        ws.cell(row=sum_row, column=1).font = Font(bold=True)
        ws.cell(row=sum_row, column=2, value=f'{len(items)}품목')
        ws.cell(row=sum_row, column=2).font = Font(bold=True)
        ws.cell(row=sum_row, column=3, value=total_qty)
        ws.cell(row=sum_row, column=3).font = Font(bold=True)

        # 컬럼 너비
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'입고체크_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings', methods=['GET'])
def get_settings():
    return jsonify({'status': 'ok', 'settings': load_settings()})

@app.route('/api/settings', methods=['POST'])
def update_settings():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': '데이터 없음'})

        save_settings(data)

        # 각 모듈의 전역 변수 업데이트
        try:
            naver_acc = get_active_account('naver')
            if naver_acc['id']:
                import poizon_data.poizon_search as ps
                ps.NAVER_ID = naver_acc['id']
                ps.NAVER_PW = naver_acc['pw']
        except Exception as e:
            print(f"⚠️ 네이버 계정 업데이트 실패: {e}")

        try:
            poizon_acc = get_active_account('poizon')
            if poizon_acc['id']:
                import poizon_data.poizon_search as ps
                ps.POIZON_ID = poizon_acc['id']
                ps.POIZON_PW = poizon_acc['pw']
        except Exception as e:
            print(f"⚠️ 포이즌 계정 업데이트 실패: {e}")

        try:
            kream_acc = get_active_account('kream')
            if kream_acc['id']:
                import kream_data.kream_search as ks
                ks.KREAM_EMAIL = kream_acc['id']
                ks.KREAM_PASSWORD = kream_acc['pw']
        except Exception as e:
            print(f"⚠️ 크림 계정 업데이트 실패: {e}")

        try:
            musinsa_acc = get_active_account('musinsa')
            if musinsa_acc['id']:
                import musinsa_data.musinsa_search as ms
                ms.MUSINSA_ID = musinsa_acc['id']
                ms.MUSINSA_PASSWORD = musinsa_acc['pw']
        except Exception as e:
            print(f"⚠️ 무신사 계정 업데이트 실패: {e}")

        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 데이터 경로 설정 API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route('/api/data_paths', methods=['GET'])
def get_data_paths():
    """현재 데이터 경로 정보 반환"""
    info = paths.get_settings()
    info['paths'] = {k: v for k, v in paths.PATHS.items()}
    return jsonify(info)

@app.route('/api/data_paths', methods=['POST'])
def update_data_paths():
    """데이터 루트 경로 변경"""
    try:
        data = request.get_json()
        new_root = data.get('data_root', '').strip()
        if not new_root:
            return jsonify({'success': False, 'error': '경로를 입력해주세요'})

        paths.set_data_root(new_root)
        paths.ensure_dirs()
        return jsonify({'success': True, 'data_root': paths.DATA_ROOT})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/data_paths/reset', methods=['POST'])
def reset_data_paths():
    """데이터 경로를 기본값으로 초기화"""
    try:
        paths.set_data_root('')
        paths.ensure_dirs()
        return jsonify({'success': True, 'data_root': paths.DATA_ROOT})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/data_paths/browse', methods=['POST'])
def browse_data_path():
    """폴더 선택 다이얼로그 열기 (Windows: 탐색기, Mac: Finder)"""
    import threading

    result = {'path': None, 'error': None}
    initial_dir = request.get_json().get('initial_dir', '') if request.get_json() else ''

    def open_dialog():
        try:
            import platform
            system = platform.system()

            if system == 'Windows':
                import subprocess
                # PowerShell의 폴더 선택 다이얼로그 사용
                ps_script = (
                    'Add-Type -AssemblyName System.Windows.Forms; '
                    '$f = New-Object System.Windows.Forms.FolderBrowserDialog; '
                    '$f.Description = "데이터 저장 경로를 선택하세요"; '
                    '$f.ShowNewFolderButton = $true; '
                )
                if initial_dir:
                    ps_script += f'$f.SelectedPath = "{initial_dir}"; '
                ps_script += (
                    'if ($f.ShowDialog() -eq "OK") { $f.SelectedPath } else { "" }'
                )
                proc = subprocess.run(
                    ['powershell', '-Command', ps_script],
                    capture_output=True, text=True, timeout=120
                )
                selected = proc.stdout.strip()
                if selected:
                    result['path'] = selected

            elif system == 'Darwin':
                import subprocess
                # AppleScript로 Finder 폴더 선택
                script = 'tell application "System Events" to activate\n'
                if initial_dir:
                    script += f'set defaultPath to POSIX file "{initial_dir}"\n'
                    script += 'set selectedFolder to choose folder with prompt "데이터 저장 경로를 선택하세요" default location defaultPath\n'
                else:
                    script += 'set selectedFolder to choose folder with prompt "데이터 저장 경로를 선택하세요"\n'
                script += 'return POSIX path of selectedFolder'
                proc = subprocess.run(
                    ['osascript', '-e', script],
                    capture_output=True, text=True, timeout=120
                )
                selected = proc.stdout.strip()
                if selected:
                    result['path'] = selected.rstrip('/')

            else:
                result['error'] = 'Unsupported OS'

        except Exception as e:
            result['error'] = str(e)

    # 메인 스레드에서 다이얼로그 열기 (별도 스레드로 실행하되 완료 대기)
    t = threading.Thread(target=open_dialog)
    t.start()
    t.join(timeout=120)

    if result['error']:
        return jsonify({'success': False, 'error': result['error']})
    if result['path']:
        return jsonify({'success': True, 'path': result['path']})
    return jsonify({'success': False, 'cancelled': True})


@app.route('/proxy_image')
def proxy_image():
    """외부 이미지 프록시 (CORS 우회용)"""
    import requests as req
    url = request.args.get('url', '')
    if not url:
        return '', 404
    try:
        r = req.get(url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
        return Response(r.content, content_type=r.headers.get('Content-Type', 'image/jpeg'))
    except Exception:
        return '', 404


# ✅ 서식 Blueprint 등록 (forms_data/forms_api.py)
try:
    import sys
    _forms_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'forms_data')
    if _forms_data_path not in sys.path:
        sys.path.insert(0, _forms_data_path)
    from forms_api import forms_bp
    app.register_blueprint(forms_bp)
    print("✅ 서식/양식 등록 성공")
except Exception as e:
    print(f"⚠️ 서식/양식 로드 실패: {e}")

# 구글 시트 자동 동기화 시작
try:
    from inventory_data.sheets_sync import sync_once, start_sync_background
    start_sync_background()
    print("✅ 구글 시트 자동 동기화 시작")
except Exception as e:
    print(f"⚠️ 시트 동기화 시작 실패: {e}")

# 텔레그램 명령어 폴링 시작
try:
    from utils.telegram import register_handler, start_polling, send_telegram_async
    import time as _time

    def _cmd_status(_):
        if is_working:
            elapsed = int(_time.time() - (work_start_time or _time.time()))
            m, s = elapsed // 60, elapsed % 60
            wt = {'scraping': '포이즌 검색', 'comparison': '리스트 비교', 'musinsa': '무신사 검색'}.get(work_type, work_type)
            return (f'🔄 <b>작업 중</b>: {wt}\n'
                    f'• 진행: {current_items} / {estimated_items}개\n'
                    f'• 경과: {m}분 {s}초')
        return '✅ 현재 대기 중 (작업 없음)'

    def _cmd_stop(_):
        global stop_flag
        if not is_working:
            return '⚠️ 현재 실행 중인 작업이 없습니다'
        stop_flag = True
        return '🛑 중지 신호를 보냈습니다'

    def _cmd_musinsa(args):
        global stop_flag
        if is_working:
            return '⚠️ 이미 작업 중입니다. /stop 으로 먼저 중지하세요'
        if not args:
            return '사용법: /musinsa [키워드] [개수]\n예) /musinsa 나이키 100\n\n무신사 → 크림 → 포이즌 순서로 자동 검색합니다'
        keyword = args[0]
        max_items = int(args[1]) if len(args) > 1 and args[1].isdigit() else 100
        t = threading.Thread(target=run_full_search, args=(keyword, max_items), daemon=True)
        t.start()
        return (f'🔍 <b>통합 검색 시작</b>\n'
                f'• 키워드: {keyword}\n'
                f'• 최대: {max_items}개\n'
                f'• 순서: 무신사 → 크림 → 포이즌\n'
                f'각 단계 완료 시 알림을 보내드립니다')

    def _cmd_ranking(args):
        if is_working:
            return '⚠️ 이미 작업 중입니다. /stop 으로 먼저 중지하세요'
        max_items = int(args[0]) if args and args[0].isdigit() else 100
        t = threading.Thread(target=run_musinsa_search, args=('', max_items, 'ranking'), daemon=True)
        t.start()
        return (f'📊 <b>무신사 랭킹 검색 시작</b>\n'
                f'• 최대: {max_items}개\n'
                f'• 완료 시 알림을 보내드립니다')

    def _cmd_sync(_):
        def _do():
            from inventory_data.sheets_sync import sync_once
            count = sync_once()
            send_telegram_async(f'✅ 구글 시트 동기화 완료: {count}건')
        threading.Thread(target=_do, daemon=True).start()
        return '🔄 구글 시트 동기화를 시작합니다...'

    def _cmd_deploy(args):
        """배포 확인 후 실행. /deploy ok → pull+재시작, /deploy force ok → 강제"""
        import subprocess

        force = len(args) > 0 and args[0].lower() == 'force'
        confirmed = args[-1].lower() == 'ok' if args else False

        # 확인 없이 /deploy 또는 /deploy force → 안내만
        if not confirmed:
            if force:
                return ('⚠️ <b>강제 배포 확인 필요</b>\n'
                        '로컬 변경사항이 모두 삭제됩니다.\n\n'
                        '실행하려면: /deploy force ok')
            return ('🚀 <b>배포 확인 필요</b>\n'
                    'GitHub 최신 코드를 받아 서버를 재시작합니다.\n\n'
                    '실행하려면: /deploy ok\n'
                    '강제 실행: /deploy force ok')

        # /deploy ok 또는 /deploy force ok → 실제 실행
        def _do():
            try:
                repo_dir = os.path.dirname(os.path.abspath(__file__))
                if force:
                    cmds = [
                        ['git', 'fetch', 'origin'],
                        ['git', 'reset', '--hard', 'origin/main'],
                        ['git', 'clean', '-fd'],
                    ]
                    outputs = []
                    for cmd in cmds:
                        r = subprocess.run(cmd, cwd=repo_dir, capture_output=True, text=True, timeout=60)
                        outputs.append((r.stdout + r.stderr).strip())
                    output = '\n'.join(o for o in outputs if o)
                    send_telegram_async(f'⚠️ <b>강제 업데이트 완료</b>\n<code>{output}</code>\n\n🔄 서버 재시작 중...')
                else:
                    result = subprocess.run(
                        ['git', 'pull', 'origin', 'main'],
                        cwd=repo_dir, capture_output=True, text=True, timeout=60
                    )
                    output = result.stdout.strip() or result.stderr.strip()
                    if result.returncode != 0:
                        send_telegram_async(
                            f'❌ pull 실패:\n<code>{output}</code>\n\n'
                            f'강제 실행: /deploy force ok'
                        )
                        return
                    send_telegram_async(f'📦 <b>Git Pull 완료</b>\n<code>{output}</code>\n\n🔄 서버 재시작 중...')
                _time.sleep(2)
                os.execv(sys.executable, [sys.executable] + sys.argv)
            except Exception as e:
                send_telegram_async(f'❌ 배포 실패: {e}')

        threading.Thread(target=_do, daemon=True).start()
        return '✅ 배포 시작합니다...'

    def _cmd_rank(args):
        """무신사 랭킹 → 크림 → 포이즌 통합 검색"""
        if is_working:
            return '⚠️ 이미 작업 중입니다. /stop 으로 먼저 중지하세요'
        max_items = int(args[0]) if args and args[0].isdigit() else 100
        t = threading.Thread(target=run_full_search, args=('', max_items, 'ranking'), daemon=True)
        t.start()
        return (f'📊 <b>랭킹 통합 검색 시작</b>\n'
                f'• 최대: {max_items}개\n'
                f'• 순서: 무신사 랭킹 → 크림 → 포이즌\n'
                f'각 단계 완료 시 알림을 보내드립니다')

    def _cmd_help(_):
        return ('📋 <b>사용 가능한 명령어</b>\n\n'
                '/status — 현재 작업 상태 확인\n'
                '/stop — 진행 중인 작업 중지\n'
                '/musinsa [키워드] [개수]\n'
                '  → 무신사 → 크림 → 포이즌 통합 검색\n'
                '  예) /musinsa 나이키 100\n'
                '/ranking [개수]\n'
                '  → 무신사 랭킹 TOP N 수집만\n'
                '  예) /ranking 100\n'
                '/rank [개수]\n'
                '  → 무신사 랭킹 → 크림 → 포이즌 통합 검색\n'
                '  예) /rank 100\n'
                '/sync — 구글 시트 강제 동기화\n'
                '/deploy — 배포 안내 (확인 필요)\n'
                '/deploy ok — GitHub pull 후 서버 재시작\n'
                '/deploy force ok — 강제 덮어쓰기 후 재시작\n'
                '/help — 도움말')

    register_handler('/status',  _cmd_status)
    register_handler('/stop',    _cmd_stop)
    register_handler('/musinsa', _cmd_musinsa)
    register_handler('/ranking', _cmd_ranking)
    register_handler('/rank',    _cmd_rank)
    register_handler('/sync',    _cmd_sync)
    register_handler('/deploy',  _cmd_deploy)
    register_handler('/help',    _cmd_help)
    register_handler('/start',   _cmd_help)

    start_polling()
    print('✅ 텔레그램 명령어 폴링 시작')
except Exception as e:
    print(f'⚠️ 텔레그램 폴링 시작 실패: {e}')

# ==========================================
# NAS 백업 API + 자동 백업 스케줄러
# ==========================================

@app.route('/api/backup/run', methods=['POST'])
def run_backup():
    """수동 NAS 백업 실행"""
    try:
        result = paths.backup_to_nas()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/backup/status', methods=['GET'])
def backup_status():
    """백업 상태 조회"""
    try:
        status = paths.get_backup_status()
        return jsonify(status)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/backup/settings', methods=['POST'])
def save_backup_settings():
    """백업 설정 저장 (시간, NAS 계정)"""
    try:
        data = request.get_json()
        paths.save_backup_settings(
            backup_hour=data.get('backup_hour', 3),
            backup_minute=data.get('backup_minute', 0),
            nas_credentials=data.get('nas_credentials')
        )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/telegram/settings', methods=['GET'])
def get_telegram_settings():
    """텔레그램 설정 조회"""
    try:
        cfg = paths.get_telegram_settings()
        return jsonify({'success': True, 'telegram': cfg})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/telegram/settings', methods=['POST'])
def save_telegram_settings():
    """텔레그램 설정 저장"""
    try:
        data = request.get_json()
        paths.save_telegram_settings({
            'bot_token': data.get('bot_token', '').strip(),
            'chat_id': data.get('chat_id', '').strip(),
            'anthropic_api_key': data.get('anthropic_api_key', '').strip(),
        })
        # 텔레그램 모듈 설정 갱신
        try:
            from utils.telegram import reload_config
            reload_config()
        except Exception:
            pass
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/telegram/test', methods=['POST'])
def test_telegram():
    """텔레그램 연결 테스트"""
    try:
        from utils.telegram import send_telegram
        ok = send_telegram('[Test] 설정 페이지에서 연결 테스트', use_html=False)
        return jsonify({'success': ok})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def _start_daily_backup():
    """설정된 시간에 NAS 자동 백업 스케줄러"""
    import time as _t
    from datetime import datetime as _dt, timedelta

    def _schedule():
        while True:
            hour, minute = paths.get_backup_schedule()
            now = _dt.now()
            target = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if target <= now:
                target += timedelta(days=1)
            wait_sec = (target - now).total_seconds()
            print(f"  [backup] next: {target.strftime('%Y-%m-%d %H:%M')} ({int(wait_sec//3600)}h later)")
            _t.sleep(wait_sec)

            # NAS 마운트 확인 및 자동 마운트 시도 (Mac)
            if not paths.is_nas_connected() and paths.IS_PRODUCTION:
                try:
                    creds = paths.get_nas_credentials()
                    if creds.get('ip') and creds.get('user') and creds.get('password'):
                        import subprocess
                        nas_share = creds.get('share', 'LEE')
                        mount_point = f"/Volumes/{nas_share}"
                        subprocess.run(['mkdir', '-p', mount_point], capture_output=True)
                        subprocess.run([
                            'mount_smbfs',
                            f"//{creds['user']}:{creds['password']}@{creds['ip']}/{nas_share}",
                            mount_point
                        ], capture_output=True, timeout=30)
                except Exception as e:
                    print(f"  [backup] NAS auto-mount failed: {e}")

            try:
                result = paths.backup_to_nas()
                if result['success']:
                    files_str = ', '.join(f['file'] for f in result['files'])
                    print(f"  [backup] OK: {files_str}")
                    try:
                        from utils.telegram import send_telegram_async
                        send_telegram_async(f"<b>[auto backup] NAS OK</b>\nfiles: {files_str}\ntime: {result['timestamp']}")
                    except Exception:
                        pass
                else:
                    print(f"  [backup] FAIL: {result.get('error', 'unknown')}")
            except Exception as e:
                print(f"  [backup] ERROR: {e}")

    t = threading.Thread(target=_schedule, daemon=True)
    t.start()

# Mac 운영서버에서만 자동 백업 실행
if paths.IS_PRODUCTION:
    _start_daily_backup()

# ==========================================
# 서버 시작
# ==========================================

if __name__ == '__main__':
    app.run(
        debug=True,
        host='0.0.0.0',
        port=3001,
        use_reloader=False  # ← 재시작 시 포트 충돌 방지
    )