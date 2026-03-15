"""
overseas_api.py
해외소싱 Flask Blueprint — 스크래핑/업로드/SSE 로그
멀티 클라이언트 브로드캐스트 방식 (데스크탑/태블릿/모바일 동시 수신)
"""

import asyncio
import json
import logging
import os
import threading
import time
from collections import deque
from datetime import datetime

from flask import Blueprint, Response, jsonify, request

from overseas_data.xebio_search import (
    scrape_xebio, set_app_status, force_close_browser,
    load_latest_products, save_products, OUTPUT_DIR,
)
from overseas_data.site_config import get_site, get_category, SITES

logger = logging.getLogger(__name__)

overseas_bp = Blueprint("overseas", __name__, url_prefix="/overseas")

# ── 브로드캐스트 로그 시스템 ────────────────────
# 최근 500개 로그를 링 버퍼에 유지 (새 클라이언트 접속 시 히스토리 전송)
_log_buffer = deque(maxlen=500)
_log_lock = threading.Lock()
_log_id = 0  # 각 로그에 고유 ID 부여 (클라이언트가 마지막 수신 ID 이후만 받기)

# ── 상태 딕셔너리 ──────────────────────────────
status = {
    "scraping": False,
    "last_scrape": None,
    "product_count": 0,
    "paused": False,
    "stop_requested": False,
}

# xebio_search에 상태 주입
set_app_status(status)


def push_log(msg: str):
    """로그를 브로드캐스트 버퍼에 추가 (모든 클라이언트가 수신)"""
    global _log_id
    logger.info(f"[overseas] {msg}")
    with _log_lock:
        _log_id += 1
        _log_buffer.append({"id": _log_id, "msg": msg, "ts": time.time()})


# ── 스크래핑 실행 (스레드) ──────────────────────

def _run_scrape(site_id, category_id, keyword, pages, brand_code=""):
    """백그라운드 스레드에서 비동기 스크래핑 실행"""
    status["scraping"] = True
    status["stop_requested"] = False
    status["paused"] = False

    push_log(f"🚀 스크래핑 시작: {site_id} › {category_id}")

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        products = loop.run_until_complete(
            scrape_xebio(
                status_callback=push_log,
                site_id=site_id,
                category_id=category_id,
                keyword=keyword,
                pages=pages,
                brand_code=brand_code,
            )
        )
        loop.close()

        if products:
            status["product_count"] = len(products)
            status["last_scrape"] = datetime.now().isoformat()
            push_log(f"✅ 스크래핑 완료! {len(products):,}개 상품 수집")
        else:
            push_log("⚠️ 수집된 상품이 없습니다")
    except Exception as e:
        push_log(f"❌ 스크래핑 오류: {e}")
        logger.exception(e)
    finally:
        status["scraping"] = False
        status["paused"] = False


# =============================================
# API 라우트
# =============================================

@overseas_bp.route("/scrape", methods=["POST"])
def manual_scrape():
    """수동 스크래핑 실행"""
    if status["scraping"]:
        return jsonify({"ok": False, "message": "이미 실행 중입니다"})

    data = request.json or {}
    site_id = data.get("site_id", "xebio")
    category_id = data.get("category_id", "sale")
    keyword = data.get("keyword", "")
    pages = data.get("pages", "")
    brand_code = data.get("brand_code", "")

    thread = threading.Thread(
        target=_run_scrape,
        args=(site_id, category_id, keyword, pages, brand_code),
        daemon=True,
    )
    thread.start()

    desc = f"{site_id} › {category_id}"
    if brand_code:
        from overseas_data.site_config import get_brands
        brand_name = get_brands(site_id).get(brand_code, brand_code)
        desc += f" › {brand_name}"
    if keyword:
        desc += f" [{keyword}]"
    if pages:
        desc += f" (p.{pages})"
    return jsonify({"ok": True, "message": f"스크래핑 시작됨 ({desc})"})


@overseas_bp.route("/auto", methods=["POST"])
def auto_run():
    """자동 실행 (스크래핑 + 후처리)"""
    if status["scraping"]:
        return jsonify({"ok": False, "message": "이미 실행 중입니다"})

    data = request.json or {}
    site_id = data.get("site_id", "xebio")
    category_id = data.get("category_id", "sale")
    keyword = data.get("keyword", "")
    brand_code = data.get("brand_code", "")

    thread = threading.Thread(
        target=_run_scrape,
        args=(site_id, category_id, keyword, "", brand_code),
        daemon=True,
    )
    thread.start()
    return jsonify({"ok": True, "message": "자동 실행 시작됨"})


@overseas_bp.route("/pause", methods=["POST"])
def pause_scrape():
    """일시정지 토글"""
    if not status["scraping"]:
        return jsonify({"ok": False, "message": "실행 중인 작업이 없습니다"})

    if status["paused"]:
        # 재개
        status["paused"] = False
        push_log("▶️ 재개 — 수집을 계속합니다!")
        return jsonify({"ok": True, "paused": False, "message": "재개됨"})
    else:
        # 일시정지
        status["paused"] = True
        push_log("⏸️ 일시정지 요청 — 현재 상품 수집 완료 후 멈춥니다...")
        return jsonify({"ok": True, "paused": True, "message": "일시정지 요청됨"})


@overseas_bp.route("/reset", methods=["POST"])
def reset_all():
    """리셋: 수집 중단 + 브라우저 강제 종료 + 데이터 삭제"""
    import glob
    import shutil

    status["stop_requested"] = True
    status["paused"] = False

    # 브라우저 강제 종료
    def close_browser():
        try:
            asyncio.run(force_close_browser())
            push_log("🔄 브라우저 종료 완료")
        except Exception as e:
            logger.debug(f"브라우저 종료 오류: {e}")
    threading.Thread(target=close_browser, daemon=True).start()

    # output 데이터 삭제
    for f in glob.glob(os.path.join(OUTPUT_DIR, "*.json")):
        try:
            os.remove(f)
        except Exception:
            pass

    # 지연 초기화
    def delayed_reset():
        time.sleep(1.5)
        status.update({
            "scraping": False,
            "last_scrape": None,
            "product_count": 0,
            "paused": False,
            "stop_requested": False,
        })
        push_log("✅ 리셋 완료 — 초기 상태로 돌아갔습니다")
    threading.Thread(target=delayed_reset, daemon=True).start()

    push_log("🔄 리셋 요청 — 모든 데이터 삭제 및 초기화")
    return jsonify({"ok": True, "message": "리셋 완료"})


@overseas_bp.route("/status", methods=["GET"])
def get_status():
    """현재 상태 반환"""
    return jsonify({
        "scraping": status["scraping"],
        "paused": status["paused"],
        "last_scrape": status["last_scrape"],
        "product_count": status["product_count"],
    })


@overseas_bp.route("/products", methods=["GET"])
def get_products():
    """수집된 상품 목록 반환"""
    products = load_latest_products()
    return jsonify(products)


@overseas_bp.route("/logs/stream")
def log_stream():
    """SSE 실시간 로그 — 멀티 클라이언트 브로드캐스트
    각 클라이언트가 독립적으로 마지막 수신 ID를 추적하므로
    데스크탑/태블릿/모바일 모두 동일한 로그를 동시에 수신합니다.
    """
    def generate():
        last_id = 0
        # 접속 시 최근 로그 50개 히스토리 전송
        with _log_lock:
            recent = list(_log_buffer)[-50:]
        for entry in recent:
            yield f"data: {json.dumps({'msg': entry['msg']}, ensure_ascii=False)}\n\n"
            last_id = entry["id"]

        # 이후 실시간 폴링
        heartbeat_counter = 0
        while True:
            time.sleep(1)
            heartbeat_counter += 1
            with _log_lock:
                new_msgs = [e for e in _log_buffer if e["id"] > last_id]
            if new_msgs:
                for entry in new_msgs:
                    yield f"data: {json.dumps({'msg': entry['msg']}, ensure_ascii=False)}\n\n"
                    last_id = entry["id"]
                heartbeat_counter = 0
            elif heartbeat_counter >= 30:
                # 30초간 새 로그 없으면 heartbeat
                yield f"data: {json.dumps({'msg': '.'})}\n\n"
                heartbeat_counter = 0

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@overseas_bp.route("/translate", methods=["POST"])
def translate_products():
    """수집된 상품 한국어 번역 (이미 번역된 상태이므로 재번역)"""
    from overseas_data.translator import translate_ja_ko, translate_brand

    products = load_latest_products()
    if not products:
        return jsonify({"ok": False, "message": "수집된 상품이 없습니다"})

    count = 0
    for p in products:
        if p.get("name") and not p.get("name_ko"):
            p["name_ko"] = translate_ja_ko(p["name"])
            count += 1
        if p.get("brand") and not p.get("brand_ko"):
            p["brand_ko"] = translate_brand(p["brand"])

    save_products(products)
    push_log(f"🌐 번역 완료: {count}개 상품")
    return jsonify({"ok": True, "message": f"{count}개 상품 번역 완료", "count": count})


@overseas_bp.route("/download/excel", methods=["GET"])
def download_excel():
    """수집된 상품을 엑셀로 다운로드"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        products = load_latest_products()
        if not products:
            return jsonify({"ok": False, "message": "다운로드할 상품이 없습니다"})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "해외소싱 상품"

        # 헤더
        headers = ["No", "브랜드", "상품명(원문)", "상품명(한국어)", "품번",
                    "가격(엔)", "재고", "링크", "수집일시"]
        header_fill = PatternFill(start_color="D4A843", end_color="D4A843", fill_type="solid")
        header_font = Font(bold=True, size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터
        for i, p in enumerate(products, 1):
            ws.cell(row=i + 1, column=1, value=i)
            ws.cell(row=i + 1, column=2, value=p.get("brand_ko", p.get("brand", "")))
            ws.cell(row=i + 1, column=3, value=p.get("name", ""))
            ws.cell(row=i + 1, column=4, value=p.get("name_ko", ""))
            ws.cell(row=i + 1, column=5, value=p.get("product_code", ""))
            ws.cell(row=i + 1, column=6, value=p.get("price_jpy", 0))
            ws.cell(row=i + 1, column=7, value="O" if p.get("in_stock") else "X")
            ws.cell(row=i + 1, column=8, value=p.get("link", ""))
            ws.cell(row=i + 1, column=9, value=p.get("scraped_at", ""))

        # 컬럼 너비
        widths = [6, 14, 40, 40, 16, 12, 6, 50, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 저장
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"overseas_products_{ts}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)

        from flask import send_file
        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        logger.exception(e)
        return jsonify({"ok": False, "message": f"엑셀 생성 오류: {e}"})


@overseas_bp.route("/sites", methods=["GET"])
def get_sites():
    """사이트/카테고리 목록 반환"""
    result = []
    for site_id, site in SITES.items():
        cats = []
        for cat_id, cat in site["categories"].items():
            cats.append({"id": cat_id, "name": cat["name"]})
        brands = []
        for code, name in site.get("brands", {}).items():
            brands.append({"code": code, "name": name})
        result.append({
            "id": site_id,
            "name": site["name"],
            "categories": cats,
            "brands": brands,
        })
    return jsonify(result)
